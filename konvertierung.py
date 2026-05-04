#!/usr/bin/env python3
"""
Konvertierungsskript: Export-CSVs -> Orca Abholer Excel
Aufruf: python konvertierung.py <eingabe.csv> [ausgabe.xlsx]

Unterstuetzte Eingabeformate:
  1. Wawican          (bestellung-export-wawican-*.csv)
  2. Cannabis Apotheke (Cannabis Apotheke_*.csv)
  3. Bestellexport    (Bestellexport_*.csv)
"""

import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TARGET_COLUMNS = [
    'Paket-Barcode', 'Bestellnummer', 'Datum', 'Vorname', 'Name',
    'Ziel-Kiosk', 'Status', 'Bestellwert', 'Versicher.', 'Email',
    'Lieferung-Adresse', 'Lieferung', 'Notizen', 'Kontrollstatus',
    'Zahlung', 'Rezept'
]


def read_csv_auto(path):
    """Liest CSV mit automatischer Erkennung von Encoding und Trennzeichen."""
    for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
        for sep in [',', ';', '\t']:
            try:
                df = pd.read_csv(path, encoding=enc, sep=sep, dtype=str)
                if df.shape[1] > 3:
                    return df
            except Exception:
                pass
    raise ValueError(f"Konnte '{path}' nicht lesen. Bitte Dateiformat prüfen.")


def detect_format(df):
    cols = set(df.columns)
    if 'Best.-Nr.' in cols:
        return 'cannabis'
    if 'OrderNumber' in cols and 'Billing_FirstName' in cols:
        return 'bestellexport'
    if 'Id' in cols and 'Reservierungsdatum' in cols:
        return 'wawican'
    return None


def last4(val):
    s = str(val).strip()
    return s[-4:] if len(s) >= 4 else s


def extract_kiosk(text):
    """Extrahiert den Apothekennamen (erstes Wort) aus der Adresse."""
    if pd.isna(text) or str(text).strip() == '':
        return ''
    return str(text).split()[0]


def clean(val):
    """Gibt leeren String statt NaN/None zurück."""
    if pd.isna(val) if not isinstance(val, str) else False:
        return ''
    s = str(val).strip()
    return '' if s in ('nan', 'None') else s


def convert_wawican(df):
    out = pd.DataFrame(index=range(len(df)))
    out['Paket-Barcode'] = df['Id'].apply(clean)
    out['Bestellnummer'] = df['Id'].apply(lambda x: last4(clean(x)))
    out['Datum'] = df['Reservierungsdatum'].apply(clean)
    out['Vorname'] = ''
    nachname = df['Nachname'].apply(clean)
    vorname = df['Vorname'].apply(clean)
    out['Name'] = nachname + ', ' + vorname
    out['Name'] = out['Name'].str.strip(', ')
    out['Ziel-Kiosk'] = df['Abholort'].apply(extract_kiosk)
    out['Status'] = df['Status'].apply(clean)
    out['Bestellwert'] = df['Rechnungsbetrag'].apply(clean)
    out['Versicher.'] = df['Versichertenstatus'].apply(clean)
    out['Email'] = ''
    out['Lieferung-Adresse'] = ''
    out['Lieferung'] = df['Lieferart'].apply(clean)
    out['Notizen'] = ''
    out['Kontrollstatus'] = 'Offen'
    zahlung = df['Zahlungsstatus'].apply(clean)
    zahlung_fallback = df['Zahlungsart'].apply(clean)
    out['Zahlung'] = zahlung.where(zahlung != '', zahlung_fallback)
    out['Rezept'] = ''
    return out


def convert_cannabis(df):
    out = pd.DataFrame(index=range(len(df)))
    out['Paket-Barcode'] = df['Best.-Nr.'].apply(clean)
    out['Bestellnummer'] = df['Best.-Nr.'].apply(lambda x: last4(clean(x)))
    out['Datum'] = df['Datum'].apply(clean)
    out['Vorname'] = ''
    out['Name'] = df['Name'].apply(clean)
    out['Ziel-Kiosk'] = df['Abholadresse'].apply(extract_kiosk)
    out['Status'] = df['Status'].apply(clean)
    bestellwert = df['Bestellwert'].apply(clean)
    out['Bestellwert'] = bestellwert.str.replace('€', '', regex=False).str.strip()
    out['Versicher.'] = df['Versicherung'].apply(clean)
    out['Email'] = ''
    out['Lieferung-Adresse'] = df['Lieferung-Adresse'].apply(clean)
    out['Lieferung'] = df['Lieferung'].apply(clean)
    out['Notizen'] = ''
    out['Kontrollstatus'] = 'Offen'
    out['Zahlung'] = df['Zahlung'].apply(clean)
    out['Rezept'] = df['Rezept'].apply(clean)
    return out


def convert_bestellexport(df):
    out = pd.DataFrame(index=range(len(df)))
    out['Paket-Barcode'] = df['OrderNumber'].apply(clean)
    out['Bestellnummer'] = df['OrderNumber'].apply(lambda x: last4(clean(x)))
    out['Datum'] = df['DateOfOrder'].apply(clean)
    out['Vorname'] = ''
    nachname = df['Billing_LastName'].apply(clean)
    vorname = df['Billing_FirstName'].apply(clean)
    out['Name'] = nachname + ', ' + vorname
    out['Name'] = out['Name'].str.strip(', ')
    out['Ziel-Kiosk'] = df['Pharmacy'].apply(extract_kiosk)
    out['Status'] = df['Status'].apply(clean)
    out['Bestellwert'] = df['Total'].apply(clean)
    out['Versicher.'] = ''
    out['Email'] = df['UserEmail'].apply(clean)
    strasse = df['ShippingAddress_Street'].apply(clean)
    hausnr = df['ShippingAddress_HouseNumber'].apply(clean)
    plz = df['ShippingAddress_Zip'].apply(clean)
    ort = df['ShippingAddress_City'].apply(clean)
    out['Lieferung-Adresse'] = strasse + ' ' + hausnr + ', ' + plz + ' ' + ort
    out['Lieferung-Adresse'] = out['Lieferung-Adresse'].str.strip(', ')
    out['Lieferung'] = df['DeliveryOption'].apply(clean)
    out['Notizen'] = ''
    out['Kontrollstatus'] = 'Offen'
    out['Zahlung'] = df['PaymentStatus'].apply(clean)
    out['Rezept'] = ''
    return out


def write_excel(result, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Skriptumtopfen'

    header_fill = PatternFill('solid', start_color='4F81BD', end_color='4F81BD')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [22, 14, 16, 10, 28, 13, 16, 13, 13, 28, 28, 12, 16, 15, 12, 10]

    for col_idx, col_name in enumerate(TARGET_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    ws.row_dimensions[1].height = 30

    data_font = Font(name='Arial', size=10)
    data_align = Alignment(vertical='center')

    for enum_idx, (_, row) in enumerate(result.iterrows(), 2):
        for col_idx, col_name in enumerate(TARGET_COLUMNS, 1):
            val = row[col_name]
            cell = ws.cell(row=enum_idx, column=col_idx, value=val if val != '' else None)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border

    ws.freeze_panes = 'A2'
    wb.save(output_path)


def convert(input_path, output_path=None):
    if output_path is None:
        from datetime import datetime
        datum = datetime.today().strftime('%d.%m.%Y')
        folder = os.path.dirname(input_path)
        output_path = os.path.join(folder, f'Orca_Abholer_{datum}.xlsx')

    print(f"Lese: {input_path}")
    df = read_csv_auto(input_path)
    fmt = detect_format(df)

    if fmt is None:
        print(f"FEHLER: Unbekanntes Eingabeformat.")
        print(f"Gefundene Spalten: {list(df.columns)}")
        sys.exit(1)

    print(f"Format erkannt: {fmt} ({len(df)} Zeilen)")

    if fmt == 'wawican':
        result = convert_wawican(df)
    elif fmt == 'cannabis':
        result = convert_cannabis(df)
    elif fmt == 'bestellexport':
        result = convert_bestellexport(df)

    write_excel(result, output_path)
    print(f"Fertig: {output_path}")
    return output_path


if __name__ == '__main__':
    try:
        if len(sys.argv) < 2:
            print("Verwendung: CSV-Datei auf dieses Skript ziehen (Drag & Drop)")
            print("oder: python konvertierung.py <eingabe.csv>")
        else:
            input_path = sys.argv[1]
            output_path = sys.argv[2] if len(sys.argv) >= 3 else None
            convert(input_path, output_path)
    except Exception as e:
        print(f"\nFEHLER: {e}")
    finally:
        input("\nDrücke Enter zum Beenden...")
