# -*- coding: utf-8 -*-
"""Sofort-Export: Tour 2 Excel im selben Format wie Bombadils _upload_tourliste.

Holt frische Daten aus OrcaScan, weist Touren live zu (T1 aus tour_zeiten.json,
T2 = alle verpackten Pakete die nicht in T1 sind), schreibt die Excel ins
TOURLISTEN_DIR (W:\...) sowie als Backup in den Bombadil-Ordner.
"""
import sys, json, tempfile, importlib.util
from pathlib import Path
from datetime import datetime

BASE_DIR    = Path(r"C:\Users\Abfuellung 15\Documents\Marcels Skripts\Bombadil")
BOMBADIL_PY = BASE_DIR / "Bombadil.py"

print("Lade Bombadil-Modul ...")
src = BOMBADIL_PY.read_text(encoding="utf-8")
src = src.replace("\ncheck_for_update()\n",
                  "\npass  # disabled by export script\n", 1)
src = src.replace("BASE_DIR      = Path(__file__).resolve().parent",
                  f'BASE_DIR      = Path(r"{BASE_DIR}")', 1)
tmp_path = Path(tempfile.gettempdir()) / "bombadil_tour2.py"
tmp_path.write_text(src, encoding="utf-8")

import os
os.chdir(BASE_DIR)
sys.path.insert(0, str(BASE_DIR))
spec = importlib.util.spec_from_file_location("bombadil_lib", tmp_path)
B    = importlib.util.module_from_spec(spec)
spec.loader.exec_module(B)
print("  OK")

print("Hole Abholer_DB aus OrcaScan ...")
abholer_df = B.fetch_abholer_orca()
print(f"  {len(abholer_df)} Zeilen")

print("Hole Tagesbote-Sheet aus OrcaScan ...")
tagesbote_df = B.fetch_sheet_orca(B.ORCA_TAGESBOTE_SHEET_ID)
print(f"  {len(tagesbote_df)} Zeilen")

print("Berechne PU heute ...")
rows, diag = B.compute_pickup_heute(abholer_df, tagesbote_df, t2_cutoff=None)
print(f"  {len(rows)} Pakete")

# Tour-Zuweisung (gleiche Logik wie Bombadil)
heute = datetime.now().strftime("%Y-%m-%d")
# WICHTIG: Der Stempel kann auf dem NAS liegen (Kollegen nutzen die NAS-Version),
# waehrend der Master lokal keine tour_zeiten hat. Daher BEIDE Orte pruefen und
# den mit den meisten T1-Barcodes nehmen (= der echte Stempel-Stand).
import json as _json2, datetime as _dt2
_heute_iso = _dt2.date.today().isoformat()
_kandidaten = []
try:
    _kandidaten.append(B._load_tour_zeiten())  # lokal (Master)
except Exception:
    pass
for _p in [Path(rf"W:\Dokumentenaustausch\Tagesskripte\Bombadil\tour_zeiten\tour_zeiten_{_heute_iso}.json"),
           Path(rf"W:\Dokumentenaustausch\Tagesskripte\Bombadil\tour_zeiten_{_heute_iso}.json")]:
    try:
        if _p.exists():
            _kandidaten.append(_json2.loads(_p.read_text(encoding="utf-8")))
    except Exception:
        pass
# den Kandidaten mit den meisten t1_barcodes nehmen (echter Stempel-Stand)
tz = max(_kandidaten, key=lambda d: len(d.get("t1_barcodes") or []), default={})
t1_bc   = set(tz.get("t1_barcodes") or [])
print(f"T1-Barcodes geladen: {len(t1_bc)}  (t1={tz.get('t1')}, t2={tz.get('t2')})  aus {len(_kandidaten)} Quelle(n)")
t2_gesetzt = bool(tz.get("t2"))
for r in rows:
    bc = r["barcode"]
    status = r.get("tb_status", "").lower()
    if status == "offen":
        r["tour"] = ""
    elif bc in t1_bc:
        r["tour"] = "T1"
    elif status == "verpackt":
        r["tour"] = "T2"   # forciert auch wenn t2_gesetzt False ist
    else:
        r["tour"] = ""

t2_rows = [r for r in rows if r.get("tour") == "T2"]
print(f"T2-Pakete: {len(t2_rows)}")
if not t2_rows:
    print("Keine T2-Pakete zum Exportieren. Abbruch.")
    sys.exit(1)

# Excel im Bombadil-Tourlisten-Format schreiben
import pandas as pd

def _last4(v):
    s = str(v).strip()
    return s[-4:] if len(s) >= 4 else s

df = pd.DataFrame([{
    "Paket-Barcode":     r["barcode"],
    "Bestellnummer":     _last4(r["barcode"]),
    "Datum":             r["scan_datum"],
    "Vorname":           "",
    "Name":              r["name"],
    "Ziel-Kiosk":        r["zielkiosk"],
    "Status":            r["tb_status"],
    "Bestellwert":       "",
    "Versicher.":        "",
    "Email":             "",
    "Lieferung-Adresse": "",
    "Lieferung":         "",
    "Notizen":           "",
    "Kontrollstatus":    r["tb_status"],
    "Zahlung":           "",
    "Rezept":            "",
} for r in t2_rows])

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

wb = Workbook()
ws = wb.active
cols = list(df.columns)
thin   = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
col_w  = [22, 14, 16, 10, 28, 13, 16, 13, 13, 28, 28, 12, 16, 15, 12, 10]
for ci, cn in enumerate(cols, 1):
    c = ws.cell(row=1, column=ci, value=cn)
    c.fill = PatternFill('solid', start_color='4F81BD', end_color='4F81BD')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(ci)].width = col_w[ci - 1] if ci <= len(col_w) else 15
ws.row_dimensions[1].height = 30
for ri, (_, row) in enumerate(df.iterrows(), 2):
    for ci, cn in enumerate(cols, 1):
        val = row[cn]
        if isinstance(val, float) and math.isnan(val):
            val = None
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(vertical='center')
        c.border = border
ws.freeze_panes = 'A2'

heute_de = datetime.now().strftime("%d.%m.%Y")
filename = f"Orca_Abholer_{heute_de}B.xlsx"

# 1. Versuche W:\TOURLISTEN_DIR
saved_paths = []
try:
    B.TOURLISTEN_DIR.mkdir(parents=True, exist_ok=True)
    pfad_nas = B.TOURLISTEN_DIR / filename
    wb.save(str(pfad_nas))
    saved_paths.append(str(pfad_nas))
except Exception as e:
    print(f"NAS-Speichern fehlgeschlagen: {e}")

# 2. Backup lokal
pfad_local = BASE_DIR / filename
wb.save(str(pfad_local))
saved_paths.append(str(pfad_local))

print()
print(">> Gespeichert:")
for p in saved_paths:
    print(f"   {p}")
print(f"  {len(t2_rows)} Pakete in T2")
