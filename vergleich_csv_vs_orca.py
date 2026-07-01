"""
Vergleicht CSV-Tagessummen (CAF Priority List = Wahrheit)
mit OrcaScan-Daten (NAS-Archiv + Google Drive) fuer Express + Normal.
"""
import pickle, sys, io
sys.stdout.reconfigure(encoding='utf-8')
import os
import pandas as pd
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BOMBADIL_DIR  = r"C:\Users\Abfuellung 15\Documents\Marcels Skripts\Bombadil"
CSV_DATEI     = r"C:\Users\Abfuellung 15\Downloads\CAF Priority List(Report Versand).csv"
NAS_EXPRESS   = r"W:\Dokumentenaustausch\Tagesskripte\Bombadil\Archiv\DHL Express"
NAS_NORMAL    = r"W:\Dokumentenaustausch\Tagesskripte\Bombadil\Archiv\DHL Normal"
OUTPUT        = r"C:\Users\Abfuellung 15\Downloads\vergleich_csv_vs_orca.xlsx"


# ── 1. CSV-Tagessummen ───────────────────────────────────────────────────────

def load_csv_tagessummen():
    df = pd.read_csv(CSV_DATEI, encoding='latin-1', sep=';', index_col=0)
    df = df.drop(columns=[c for c in df.columns if 'Unnamed' in str(c)], errors='ignore')
    summen = df.apply(pd.to_numeric, errors='coerce').sum(axis=0)
    summen.index = pd.to_datetime(summen.index, format='%d.%m.%Y', errors='coerce')
    summen = summen[summen.index.notna() & (summen > 0)].sort_index()
    print(f"CSV: {len(summen)} Tage, von {summen.index.min().date()} bis {summen.index.max().date()}")
    return summen


# ── 2. OrcaScan-Daten laden (NAS + Drive) ───────────────────────────────────

def lade_nas(folder):
    rows = []
    if not os.path.isdir(folder):
        return pd.DataFrame()
    for f in os.listdir(folder):
        if not f.lower().endswith('.xlsx'):
            continue
        try:
            xls = pd.read_excel(os.path.join(folder, f), sheet_name=None)
            for sheet in xls.values():
                sheet.columns = [str(c).strip() for c in sheet.columns]
                bc  = next((c for c in sheet.columns if 'barcode' in c.lower()), None)
                sc  = next((c for c in sheet.columns if 'scan' in c.lower() or 'date' in c.lower()), None)
                if bc and sc:
                    sub = sheet[[bc, sc]].copy()
                    sub.columns = ['Barcode', 'Datum']
                    sub['Barcode'] = sub['Barcode'].astype(str).str.strip()
                    sub['Datum']   = pd.to_datetime(sub['Datum'], errors='coerce').dt.date
                    rows.append(sub.dropna())
        except Exception:
            pass
    if not rows:
        return pd.DataFrame()
    df = pd.concat(rows, ignore_index=True).drop_duplicates(subset=['Barcode'])
    return df


def lade_drive():
    with open(f"{BOMBADIL_DIR}\\token.json", "rb") as fh:
        creds = pickle.load(fh)
    service = build("drive", "v3", credentials=creds, cache_discovery=False)
    rows = []
    for prefix in ["Express", "Normal"]:
        q = f"trashed=false and name contains 'DHL_{prefix}' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
        files = service.files().list(q=q, fields="files(id,name)", pageSize=200).execute().get("files", [])
        print(f"  Drive {prefix}: {len(files)} Dateien")
        for f in files:
            try:
                req = service.files().get_media(fileId=f['id'])
                buf = io.BytesIO()
                dl  = MediaIoBaseDownload(buf, req)
                done = False
                while not done:
                    _, done = dl.next_chunk()
                buf.seek(0)
                xls = pd.read_excel(buf, sheet_name=None)
                for sheet in xls.values():
                    sheet.columns = [str(c).strip() for c in sheet.columns]
                    bc = next((c for c in sheet.columns if 'barcode' in c.lower()), None)
                    sc = next((c for c in sheet.columns if 'scan' in c.lower() or 'date' in c.lower()), None)
                    if bc and sc:
                        sub = sheet[[bc, sc]].copy()
                        sub.columns = ['Barcode', 'Datum']
                        sub['Barcode'] = sub['Barcode'].astype(str).str.strip()
                        sub['Datum']   = pd.to_datetime(sub['Datum'], errors='coerce').dt.date
                        rows.append(sub.dropna())
            except Exception as e:
                print(f"    Fehler {f['name']}: {e}")
    if not rows:
        return pd.DataFrame()
    return pd.concat(rows, ignore_index=True).drop_duplicates(subset=['Barcode'])


# ── 3. Vergleich + Excel ─────────────────────────────────────────────────────

def schreibe_excel(vergleich_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Vergleich"

    FILL_HEAD = PatternFill("solid", start_color="2C3E50")
    FILL_OK   = PatternFill("solid", start_color="C6EFCE")
    FILL_WARN = PatternFill("solid", start_color="FFEB9C")
    FILL_ERR  = PatternFill("solid", start_color="FFC7CE")
    FONT_HEAD = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    FONT_NORM = Font(name="Arial", size=10)
    FONT_BOLD = Font(name="Arial", size=10, bold=True)
    ALIGN_C   = Alignment(horizontal="center", vertical="center")
    ALIGN_L   = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Datum", "Wochentag", "CSV Gesamt", "OrcaScan Gesamt", "Differenz", "Fehlend %", "Status"]
    widths  = [14, 12, 14, 18, 12, 12, 22]

    ws.merge_cells("A1:G1")
    ws["A1"] = "Vergleich: CAF Priority List (Wahrheit) vs. OrcaScan (Express + Normal)"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="2C3E50")
    ws.row_dimensions[1].height = 22

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = FILL_HEAD
        cell.font = FONT_HEAD
        cell.alignment = ALIGN_C
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    WOCHENTAGE = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

    for r, row in enumerate(vergleich_df.itertuples(index=False), start=3):
        csv_val  = int(row.CSV)
        orca_val = int(row.OrcaScan)
        diff     = orca_val - csv_val
        pct      = round((csv_val - orca_val) / csv_val * 100, 1) if csv_val > 0 else 0
        wt       = WOCHENTAGE[pd.Timestamp(row.Datum).weekday()]

        if abs(pct) <= 5:
            fill   = FILL_OK
            status = "OK"
        elif abs(pct) <= 20:
            fill   = FILL_WARN
            status = f"Abweichung {pct}%"
        else:
            fill   = FILL_ERR
            status = f"Grosse Luecke {pct}%"

        vals = [str(row.Datum), wt, csv_val, orca_val, diff, f"{pct}%", status]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill      = fill
            cell.font      = FONT_BOLD if abs(pct) > 5 else FONT_NORM
            cell.alignment = ALIGN_C if c > 1 else ALIGN_L
            cell.border    = border

    ws.freeze_panes = "A3"
    wb.save(OUTPUT)
    print(f"Excel gespeichert: {OUTPUT}")


# ── Main ─────────────────────────────────────────────────────────────────────

print("=== Schritt 1: CSV laden ===")
csv_summen = load_csv_tagessummen()

print("\n=== Schritt 2: OrcaScan laden (NAS) ===")
nas_ex = lade_nas(NAS_EXPRESS)
nas_no = lade_nas(NAS_NORMAL)
print(f"NAS: {len(nas_ex)} Express + {len(nas_no)} Normal = {len(nas_ex)+len(nas_no)} Eintraege")

print("\n=== Schritt 3: OrcaScan laden (Drive) ===")
drive_df = lade_drive()
print(f"Drive: {len(drive_df)} Eintraege")

print("\n=== Schritt 4: Zusammenfuehren ===")
alle = pd.concat([nas_ex, nas_no, drive_df], ignore_index=True).drop_duplicates(subset=['Barcode'])
print(f"Gesamt OrcaScan (dedupliziert): {len(alle)} Eintraege")

orca_pro_tag = alle.groupby('Datum').size()
orca_pro_tag.index = pd.to_datetime(orca_pro_tag.index)

print("\n=== Schritt 5: Vergleich erstellen ===")
alle_tage = sorted(set(csv_summen.index) | set(orca_pro_tag.index))
rows = []
for tag in alle_tage:
    csv_val  = int(csv_summen.get(tag, 0))
    orca_val = int(orca_pro_tag.get(tag, 0))
    if csv_val == 0 and orca_val == 0:
        continue
    rows.append({'Datum': tag.date(), 'CSV': csv_val, 'OrcaScan': orca_val})

vergleich_df = pd.DataFrame(rows)
schreibe_excel(vergleich_df)

print("\n=== Letzte 15 Tage ===")
for _, r in vergleich_df.tail(15).iterrows():
    diff = int(r['OrcaScan']) - int(r['CSV'])
    pct  = round((int(r['CSV']) - int(r['OrcaScan'])) / int(r['CSV']) * 100, 1) if r['CSV'] > 0 else 0
    print(f"  {str(r['Datum'])[:10]}  CSV={int(r['CSV']):5}  Orca={int(r['OrcaScan']):5}  Diff={diff:+5}  ({pct}% fehlen)")
