# -*- coding: utf-8 -*-
"""
Standalone-Skript: Exportiert alle Pakete OHNE Tour-Zuweisung als Excel.

Nutzt Bombadil's eigene Funktionen (fetch_*, compute_pickup_heute) – lädt das
Modul aber so, dass weder check_for_update() noch die UI startet.

Speichert die Excel im selben Format wie der Tour-1/Tour-2-Export.

Ausführen:
    python export_ohne_tour.py
"""
import sys
import json
import tempfile
import importlib.util
from pathlib import Path
from datetime import datetime

BASE_DIR     = Path(r"C:\Users\Abfuellung 15\Documents\Marcels Skripts\Bombadil")
BOMBADIL_PY  = BASE_DIR / "Bombadil.py"

# ─── Bombadil-Modul laden ohne check_for_update / UI ────────────────────────
print("Lade Bombadil-Modul ...")
src = BOMBADIL_PY.read_text(encoding="utf-8")
# check_for_update() ist ein Top-Level-Call ~Zeile 138 – entschärfen
src = src.replace("\ncheck_for_update()\n", "\npass  # disabled by export script\n", 1)
# BASE_DIR auf echten Bombadil-Ordner pinnen (sonst zeigt __file__ auf tmp)
src = src.replace(
    "BASE_DIR      = Path(__file__).resolve().parent",
    f'BASE_DIR      = Path(r"{BASE_DIR}")', 1)
# tmp-Datei in einem isolierten Ordner (Bombadil-relative imports sind keine,
# daher reicht das tmp-Verzeichnis):
tmp_path = Path(tempfile.gettempdir()) / "bombadil_no_update.py"
tmp_path.write_text(src, encoding="utf-8")

# Wichtig: BASE_DIR auf sys.path setzen damit .env etc. relativ gefunden wird
import os
os.chdir(BASE_DIR)
sys.path.insert(0, str(BASE_DIR))

spec = importlib.util.spec_from_file_location("bombadil_lib", tmp_path)
B    = importlib.util.module_from_spec(spec)
spec.loader.exec_module(B)
print("  OK")

# ─── Daten holen ────────────────────────────────────────────────────────────
print("Hole Abholer_DB aus OrcaScan ...")
abholer_df = B.fetch_abholer_orca()
print(f"  {len(abholer_df)} Zeilen")

print("Hole Tagesbote-Sheet aus OrcaScan ...")
tagesbote_df = B.fetch_sheet_orca(B.ORCA_TAGESBOTE_SHEET_ID)
print(f"  {len(tagesbote_df)} Zeilen")

# ─── compute_pickup_heute (Bombadil-Logik) ──────────────────────────────────
print("Berechne PU heute ...")
rows, diag = B.compute_pickup_heute(abholer_df, tagesbote_df, t2_cutoff=None)
print(f"  {len(rows)} Pakete im Tagesbote (Diag: {diag})")

# ─── Tour-Zuweisung (zeit-basiert wie in Bombadil 1.0.26) ──────────────────
heute   = datetime.now().strftime("%Y-%m-%d")
tz_path = BASE_DIR / f"tour_zeiten_{heute}.json"
tz      = json.loads(tz_path.read_text(encoding="utf-8")) if tz_path.exists() else {}
t1_bc   = set(tz.get("t1_barcodes") or [])
t2_bc   = set(tz.get("t2_barcodes") or [])
t1_time = tz.get("t1") or ""
t2_time = tz.get("t2") or ""
print(f"Tour-Zeiten: T1={t1_time or '-'}  T2={t2_time or '-'}  "
      f"(T1-Liste={len(t1_bc)}, T2-Liste={len(t2_bc)})")

def _hhmm(s):
    if not s:
        return ""
    parts = str(s).split()
    return parts[-1] if parts else ""

for r in rows:
    bc = r["barcode"]
    if r.get("tb_status", "").lower() == "offen":
        r["tour"] = ""
        continue
    if bc in t1_bc:
        r["tour"] = "T1"; continue
    if bc in t2_bc:
        r["tour"] = "T2"; continue
    vp = _hhmm(r.get("verpackt_at", ""))
    if vp and t1_time and vp <= t1_time:
        r["tour"] = "T1"
    elif vp and t2_time and (not t1_time or vp > t1_time) and vp <= t2_time:
        r["tour"] = "T2"
    else:
        r["tour"] = ""

# ─── Pakete OHNE Tour filtern ───────────────────────────────────────────────
ohne_tour = [r for r in rows
             if r.get("tour", "") == ""
             and r.get("tb_status", "").lower() == "verpackt"]
print(f"Pakete OHNE Tour-Zuweisung (Verpackt-Status): {len(ohne_tour)}")

if not ohne_tour:
    print("Keine Pakete zum Exportieren. Skript beendet.")
    sys.exit(0)

# ─── Excel im Bombadil-Format schreiben ────────────────────────────────────
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils  import get_column_letter

HEADERS = [
    "Paket-Barcode", "Reservierungsnr.", "Best.-Nr.", "Bestellnummer",
    "Datum", "Name", "Ankunftsort", "Ziel-Kiosk", "Lieferung-Name",
    "Standort", "Paketstatus", "Lieferung-Adresse", "Abgeholt_At",
    "Abholbereit_At", "Lieferung-Zusatz", "Verpackt_At", "Ziel Kiosk",
    "Unterschrift", "Scan-Datum", "Status", "Bestellwert", "Versicher.",
    "Lieferung", "Zahlung", "Rezept", "Notizen", "Paketfoto",
]
COL_WIDTHS = [30, 18, 12, 16, 14, 30, 16, 20, 20,
              14, 16, 20, 20, 20, 18, 20, 20,
              14, 18, 14, 14, 12, 14, 12, 10, 14, 14]

def _row(r):
    return [
        r["barcode"], None, None, None,
        r.get("scan_datum"), r["name"], None, r.get("zielkiosk"), None,
        None, r.get("db_status"), None, None,
        r.get("abholbereit_at"), None, r.get("verpackt_at"), r.get("zielkiosk"),
        None, r.get("scan_datum"), r.get("tb_status"), None, None,
        None, None, None, None, None,
    ]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ohne Tour"
ws.append(HEADERS)
for c in ws[1]:
    c.font = Font(bold=True)
for r in ohne_tour:
    ws.append(_row(r))
for i, w in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

out_path = BASE_DIR / f"PU_OhneTour_{heute}.xlsx"
wb.save(str(out_path))
print(f"\n>> Gespeichert: {out_path}")
print(f"  {len(ohne_tour)} Pakete")
