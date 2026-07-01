"""
Vergleicht DHL-Sendungen aus PDF-Versandlabeln (Backup-Ordner)
mit den NAS-Archiv-Daten (OrcaScan-Exporte) – pro Tag und Typ.

Ausgabe: Excel-Datei mit farbiger Markierung der Abweichungen.
"""

import os
import re
import pdfplumber
import pandas as pd
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PDF_FOLDER    = r"W:\Dokumentenaustausch\Tagesskripte\Datensicherung"
NAS_EXPRESS   = r"W:\Dokumentenaustausch\Tagesskripte\Bombadil\Archiv\DHL Express"
NAS_NORMAL    = r"W:\Dokumentenaustausch\Tagesskripte\Bombadil\Archiv\DHL Normal"
OUTPUT        = r"C:\Users\Abfuellung 15\Downloads\vergleich_dhl_pdfs_vs_orca.xlsx"
BOMBADIL_DIR  = r"C:\Users\Abfuellung 15\Documents\Marcels Skripts\Bombadil"

# Google Drive Ordner-IDs (aus Bombadil)
GDRIVE_FOLDER_DHL_NORMAL  = "1G5_i9zUvjqVBCnE_ME-fPkhK4j5SPWpe"
GDRIVE_FOLDER_DHL_EXPRESS = "1J_oZcBmPECYfL7xj5U0oxgao--sdzS_c"

PAT_EXPRESS = re.compile(r'\(J\)\s+(JD\d{2}\s+[\d\s]{15,})')
PAT_NORMAL  = re.compile(r'(?:Sendungsnr\.|Sendungsnummer)\s*[:\n]?\s*\(00\)\s*(\d{18,20})')


# ── 1. PDF-Extraktion ────────────────────────────────────────────────────────

def process_pdf(path):
    mtime    = os.path.getmtime(path)
    scan_dt  = datetime.fromtimestamp(mtime).date()
    try:
        with pdfplumber.open(path) as pdf:
            text = "".join(p.extract_text() or "" for p in pdf.pages)
        m_ex = PAT_EXPRESS.search(text)
        if m_ex:
            bc = "J" + m_ex.group(1).replace(" ", "")
            return ("express", bc, scan_dt)
        m_no = PAT_NORMAL.search(text)
        if m_no:
            bc = "00" + m_no.group(1)
            return ("normal", bc, scan_dt)
    except Exception:
        pass
    return (None, None, None)


def load_pdfs():
    files = [os.path.join(PDF_FOLDER, f)
             for f in os.listdir(PDF_FOLDER) if f.lower().endswith(".pdf")]
    print(f"{len(files)} PDFs gefunden, lese aus …")

    express_rows, normal_rows = [], []
    done = 0
    with ProcessPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(process_pdf, p): p for p in files}
        for fut in as_completed(futures):
            typ, bc, dt = fut.result()
            done += 1
            if typ == "express":
                express_rows.append({"Barcode": bc, "Datum": dt})
            elif typ == "normal":
                normal_rows.append({"Barcode": bc, "Datum": dt})
            if done % 1000 == 0:
                print(f"  {done}/{len(files)} PDFs …")

    df_ex = pd.DataFrame(express_rows)
    df_no = pd.DataFrame(normal_rows)
    print(f"PDFs: {len(df_ex)} Express, {len(df_no)} Normal gefunden.")
    return df_ex, df_no


# ── 2. NAS-Archiv laden ──────────────────────────────────────────────────────

def load_nas(folder):
    rows = []
    if not os.path.isdir(folder):
        return pd.DataFrame()
    for f in os.listdir(folder):
        if not f.lower().endswith(".xlsx"):
            continue
        path = os.path.join(folder, f)
        try:
            xls = pd.read_excel(path, sheet_name=None)
            for sheet in xls.values():
                sheet.columns = [str(c).strip() for c in sheet.columns]
                bc_col   = next((c for c in sheet.columns if "barcode" in c.lower()), None)
                scan_col = next((c for c in sheet.columns if "scan" in c.lower() or "date" in c.lower()), None)
                if bc_col and scan_col:
                    sub = sheet[[bc_col, scan_col]].copy()
                    sub.columns = ["Barcode", "Datum"]
                    sub["Barcode"] = sub["Barcode"].astype(str).str.strip()
                    sub["Datum"]   = pd.to_datetime(sub["Datum"], errors="coerce").dt.date
                    rows.append(sub.dropna())
        except Exception:
            pass
    if not rows:
        return pd.DataFrame()
    df = pd.concat(rows, ignore_index=True)
    df = df.drop_duplicates(subset=["Barcode"])
    return df


# ── 2b. Google Drive Backups laden ──────────────────────────────────────────

def _get_drive_service():
    """Erstellt Google Drive Service mit bestehendem Bombadil-Token (Pickle-Format)."""
    import pickle
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    token_path = os.path.join(BOMBADIL_DIR, "token.json")

    with open(token_path, "rb") as fh:
        creds = pickle.load(fh)

    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        with open(token_path, "wb") as fh:
            pickle.dump(creds, fh)

    return build("drive", "v3", credentials=creds, cache_discovery=False)


def load_drive_backups(folder_id, prefix):
    """Lädt alle DHL-Backup/Archiv-Dateien vom Drive (sucht nach Dateiname)."""
    try:
        from googleapiclient.http import MediaIoBaseDownload
        import io
        service = _get_drive_service()

        # Suche nach Dateinamen (drive.file-Scope erlaubt keinen Ordner-Filter für fremde Ordner)
        name_filter = f"DHL_{prefix}"
        query = (f"trashed=false and name contains '{name_filter}' "
                 f"and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'")
        result = service.files().list(q=query, fields="files(id,name)", pageSize=200).execute()
        files  = result.get("files", [])
        print(f"  Drive {prefix}: {len(files)} Dateien gefunden")

        rows = []
        for f in files:
            try:
                req  = service.files().get_media(fileId=f["id"])
                buf  = io.BytesIO()
                dl   = MediaIoBaseDownload(buf, req)
                done = False
                while not done:
                    _, done = dl.next_chunk()
                buf.seek(0)
                xls = pd.read_excel(buf, sheet_name=None)
                for sheet in xls.values():
                    sheet.columns = [str(c).strip() for c in sheet.columns]
                    bc_col   = next((c for c in sheet.columns if "barcode" in c.lower()), None)
                    scan_col = next((c for c in sheet.columns if "scan" in c.lower() or "date" in c.lower()), None)
                    if bc_col and scan_col:
                        sub = sheet[[bc_col, scan_col]].copy()
                        sub.columns = ["Barcode", "Datum"]
                        sub["Barcode"] = sub["Barcode"].astype(str).str.strip()
                        sub["Datum"]   = pd.to_datetime(sub["Datum"], errors="coerce").dt.date
                        rows.append(sub.dropna())
            except Exception as e:
                print(f"    Fehler bei {f['name']}: {e}")

        if not rows:
            return pd.DataFrame()
        df = pd.concat(rows, ignore_index=True)
        df = df.drop_duplicates(subset=["Barcode"])
        return df

    except Exception as e:
        print(f"  Drive {prefix} nicht geladen: {e}")
        return pd.DataFrame()


def merge_nas_drive(nas_df, drive_df):
    """Kombiniert NAS + Drive, dedupliziert per Barcode (Drive gewinnt bei Konflikten)."""
    if nas_df.empty and drive_df.empty:
        return pd.DataFrame()
    if nas_df.empty:
        return drive_df
    if drive_df.empty:
        return nas_df
    # Drive-Barcodes haben Vorrang
    drive_barcodes = set(drive_df["Barcode"])
    nas_only = nas_df[~nas_df["Barcode"].isin(drive_barcodes)]
    combined = pd.concat([drive_df, nas_only], ignore_index=True)
    return combined


# ── 3. Vergleich erstellen ───────────────────────────────────────────────────

def make_comparison(pdf_df, nas_df, label):
    """Erstellt Tages-Vergleich zwischen PDF und NAS für einen Typ."""
    if pdf_df.empty and nas_df.empty:
        return pd.DataFrame()

    def counts(df):
        if df.empty:
            return pd.Series(dtype=int)
        return df.groupby("Datum").size()

    pdf_counts = counts(pdf_df)
    nas_counts = counts(nas_df)

    alle_tage = sorted(set(pdf_counts.index) | set(nas_counts.index))
    rows = []
    for tag in alle_tage:
        p = int(pdf_counts.get(tag, 0))
        n = int(nas_counts.get(tag, 0))
        diff = p - n
        rows.append({
            "Datum":      tag,
            "PDF-Anzahl": p,
            "OrcaScan":   n,
            "Differenz":  diff,
            "Status":     "✅ gleich" if diff == 0
                          else (f"📄 +{diff} nur in PDF" if diff > 0
                                else f"🔍 {diff} nur in Orca"),
        })
    df = pd.DataFrame(rows)
    df["Typ"] = label
    return df


# ── 4. Excel-Ausgabe ─────────────────────────────────────────────────────────

def write_excel(df_ex_cmp, df_no_cmp):
    wb = Workbook()
    wb.remove(wb.active)

    FILL_OK   = PatternFill("solid", start_color="C6EFCE")  # grün
    FILL_WARN = PatternFill("solid", start_color="FFEB9C")  # gelb
    FILL_ERR  = PatternFill("solid", start_color="FFC7CE")  # rot
    FILL_HEAD = PatternFill("solid", start_color="2C3E50")
    FONT_HEAD = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    FONT_NORM = Font(name="Arial", size=10)
    FONT_BOLD = Font(name="Arial", size=10, bold=True)
    ALIGN_C   = Alignment(horizontal="center", vertical="center")
    ALIGN_L   = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Datum", "PDF-Anzahl", "OrcaScan", "Differenz", "Status"]
    col_w   = [14,       13,           12,          12,          26]

    for sheet_name, df_cmp in [("DHL Express", df_ex_cmp), ("DHL Normal", df_no_cmp)]:
        ws = wb.create_sheet(sheet_name)

        # Überschrift
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Vergleich {sheet_name}: PDF-Versandlabel vs. OrcaScan-Archiv"
        ws["A1"].font    = Font(name="Arial", size=13, bold=True, color="2C3E50")
        ws["A1"].alignment = ALIGN_L
        ws.row_dimensions[1].height = 22

        # Header-Zeile
        for col, (h, w) in enumerate(zip(headers, col_w), start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill      = FILL_HEAD
            cell.font      = FONT_HEAD
            cell.alignment = ALIGN_C
            cell.border    = border
            ws.column_dimensions[get_column_letter(col)].width = w

        if df_cmp.empty:
            ws.cell(row=3, column=1, value="Keine Daten verfügbar.")
            continue

        for r_idx, row in enumerate(df_cmp.itertuples(index=False), start=3):
            diff = row.Differenz
            fill = FILL_OK if diff == 0 else (FILL_WARN if abs(diff) <= 5 else FILL_ERR)

            vals = [str(row.Datum), row._1, row.OrcaScan, diff, row.Status]  # _1 = "PDF-Anzahl" (Bindestrich → positional)
            for c_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill      = fill
                cell.font      = FONT_BOLD if diff != 0 else FONT_NORM
                cell.alignment = ALIGN_C if c_idx > 1 else ALIGN_L
                cell.border    = border

        ws.row_dimensions[2].height = 18
        ws.freeze_panes = "A3"

    # Zusammenfassung
    ws_sum = wb.create_sheet("Zusammenfassung", 0)
    ws_sum["A1"] = "Zusammenfassung DHL-Vergleich"
    ws_sum["A1"].font = Font(name="Arial", size=14, bold=True, color="2C3E50")
    ws_sum.merge_cells("A1:D1")

    sum_rows = [("", "Gesamt Tage", "Tage gleich", "Tage mit Abweichung")]
    for label, df_c in [("DHL Express", df_ex_cmp), ("DHL Normal", df_no_cmp)]:
        if df_c.empty:
            sum_rows.append((label, 0, 0, 0))
        else:
            gesamt = len(df_c)
            gleich = int((df_c["Differenz"] == 0).sum())
            abweich = gesamt - gleich
            sum_rows.append((label, gesamt, gleich, abweich))

    for r, row in enumerate(sum_rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws_sum.cell(row=r, column=c, value=val)
            if r == 2:
                cell.fill = FILL_HEAD
                cell.font = FONT_HEAD
            else:
                cell.font = FONT_NORM
            cell.alignment = ALIGN_C if c > 1 else ALIGN_L
            cell.border = border

    for col, w in zip("ABCD", [20, 14, 14, 20]):
        ws_sum.column_dimensions[col].width = w

    wb.save(OUTPUT)
    print(f"\nExcel gespeichert: {OUTPUT}")


# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=== Schritt 1: PDFs auslesen ===")
    pdf_ex, pdf_no = load_pdfs()

    print("\n=== Schritt 2a: NAS-Archiv laden ===")
    nas_ex = load_nas(NAS_EXPRESS)
    nas_no = load_nas(NAS_NORMAL)
    print(f"NAS: {len(nas_ex)} Express-Eintraege, {len(nas_no)} Normal-Eintraege")

    print("\n=== Schritt 2b: Google Drive Backups laden ===")
    drive_ex = load_drive_backups(GDRIVE_FOLDER_DHL_EXPRESS, "Express")
    drive_no = load_drive_backups(GDRIVE_FOLDER_DHL_NORMAL,  "Normal")
    print(f"Drive: {len(drive_ex)} Express-Eintraege, {len(drive_no)} Normal-Eintraege")

    print("\n=== Schritt 2c: NAS + Drive zusammenfuehren ===")
    orca_ex = merge_nas_drive(nas_ex, drive_ex)
    orca_no = merge_nas_drive(nas_no, drive_no)
    print(f"Gesamt: {len(orca_ex)} Express-Eintraege, {len(orca_no)} Normal-Eintraege")

    print("\n=== Schritt 3: Vergleich erstellen ===")
    cmp_ex = make_comparison(pdf_ex, orca_ex, "Express")
    cmp_no = make_comparison(pdf_no, orca_no, "Normal")

    print("\n=== Schritt 4: Excel schreiben ===")
    write_excel(cmp_ex, cmp_no)

    # Kurze Vorschau (ohne Emojis fuer Windows-Konsole)
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    for label, df in [("EXPRESS", cmp_ex), ("NORMAL", cmp_no)]:
        if df.empty:
            continue
        abw = df[df["Differenz"] != 0]
        print(f"\n{label}: {len(abw)} Tage mit Abweichung (von {len(df)} Tagen gesamt)")
        for _, r in df.tail(14).iterrows():
            print(f"  {str(r['Datum'])[:10]}  PDF={int(r['PDF-Anzahl']):4}  Orca={int(r['OrcaScan']):4}  Diff={int(r['Differenz']):+4}")
