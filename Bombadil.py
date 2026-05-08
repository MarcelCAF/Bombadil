# ============================================================
# Bombadil – robuste Start-sichere Version (KOMPLETT)
# + DHL_Express (heute)
# + DHL (heute): DHL_Normal Sheet (nur heute) -> Export YYMMDD.xlsx (Downloads, überschreiben)
# Fix: entfernt führendes Apostroph in "Package Barcode" + Excel-Export Spalte als TEXT (@)
# v2: einheitlicher TableTab, Sortierung, Alternating-Rows, Statusleiste, Pickups/Tag-Tab
# ============================================================

# Konsolenfenster unter Windows ausblenden
import sys as _sys
if _sys.platform == "win32":
    import ctypes as _ctypes
    _ctypes.windll.user32.ShowWindow(
        _ctypes.windll.kernel32.GetConsoleWindow(), 0)

import re
import time
import threading
from pathlib import Path
from datetime import datetime, date, timedelta

# -------------------------
# Robuste Imports
# -------------------------
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    from tkinter import ttk
except Exception as e:
    raise RuntimeError("tkinter/ttk nicht verfügbar") from e

try:
    import pandas as pd
except Exception as e:
    raise RuntimeError("pandas fehlt (pip install pandas openpyxl)") from e

# openpyxl (für Excel-Formatierung @)
try:
    from openpyxl import load_workbook
except Exception as e:
    raise RuntimeError("openpyxl fehlt (pip install openpyxl)") from e

# TZ robust (Windows ggf. tzdata)
try:
    from zoneinfo import ZoneInfo
    TZ = ZoneInfo("Europe/Berlin")
except Exception:
    TZ = None

# Google Drive API (optional – pip install google-api-python-client google-auth google-auth-oauthlib)
try:
    import io as _io
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

try:
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request as _GAuthRequest
    import pickle as _pickle
    OAUTH_AVAILABLE = True
except ImportError:
    OAUTH_AVAILABLE = False


# ============================================================
# Pfade
# ============================================================
BASE_DIR      = Path(__file__).resolve().parent
SETTINGS_FILE = BASE_DIR / "settings.json"
LOGO_PATH = BASE_DIR / "logo.png"   # optional

# ============================================================
# Version & Auto-Updater
# ============================================================
VERSION = "1.0.20"

GITHUB_RAW = "https://raw.githubusercontent.com/MarcelCAF/Bombadil/master"

def check_for_update():
    """Prüft beim Start ob eine neuere Version auf GitHub verfügbar ist.
    Funktioniert auch auf Netzlaufwerken (NAS) – die Datei wird direkt
    am Speicherort (NAS-Pfad) aktualisiert."""
    import urllib.request
    import shutil
    import subprocess

    eigene_datei = Path(__file__).resolve()

    try:
        with urllib.request.urlopen(f"{GITHUB_RAW}/version.txt", timeout=5) as r:
            remote_version = r.read().decode().strip()
    except Exception:
        return  # Kein Internet oder GitHub nicht erreichbar → einfach starten

    def _version_tuple(v):
        return tuple(int(x) for x in v.split("."))

    if _version_tuple(remote_version) <= _version_tuple(VERSION):
        return  # Bereits aktuell

    # Update-Dialog
    import tkinter as tk
    from tkinter import messagebox
    _root = tk.Tk()
    _root.withdraw()
    antwort = messagebox.askyesno(
        "Update verfügbar",
        f"Neue Version {remote_version} verfügbar (aktuell: {VERSION}).\n\nJetzt aktualisieren?"
    )
    _root.destroy()

    if not antwort:
        return

    # Neue Datei herunterladen
    eigene_datei = Path(__file__).resolve()
    backup_datei = eigene_datei.with_suffix(f".backup_{VERSION}.py")
    try:
        shutil.copy2(eigene_datei, backup_datei)
        with urllib.request.urlopen(f"{GITHUB_RAW}/Bombadil.py", timeout=15) as r:
            neue_version = r.read()
        eigene_datei.write_bytes(neue_version)
        # Bombadil neu starten
        subprocess.Popen([_sys.executable, str(eigene_datei)])
        _sys.exit()
    except Exception as e:
        import tkinter as tk
        from tkinter import messagebox
        _root = tk.Tk()
        _root.withdraw()
        messagebox.showerror("Update fehlgeschlagen", f"Fehler beim Update:\n{e}")
        _root.destroy()

check_for_update()

POLL_MS = 3000
ADMIN_MODE = True   # False = Lite-Version (kein Cleanup, kein Backup)

# -------------------------
# Spalten (robust)
# -------------------------
COL_STATUS      = ["Paketstatus", "Paketstatus "]
COL_ABHOLBEREIT = ["Abholbereit_At", "Abholbereit_AT", "Abholbereit_at"]
COL_ABGEHOLT    = ["Abgeholt_At", "Abgeholt_AT", "Abgeholt_at"]
COL_BARCODE     = ["Paket-Barcode", "Paket Barcode", "Paketbarcode", "Barcode"]
COL_NAME        = ["Name", "Kunde", "Customer", "Empfänger", "Empfaenger"]
COL_ZAHLUNG     = ["Zahlung", "Zahlung ", "Payment"]
COL_ZIELKIOSK   = ["Ziel-Kiosk", "Ziel Kiosk", "Ziel_Kiosk", "Zielkiosk"]
COL_BESTELLWERT = ["Bestellwert", "bestellwert", "Bestell-Wert", "Bestellwert "]

# Verpackt-Timestamp
COL_VERPACKT_AT = ["Verpackt_At", "Verpackt_AT", "verpackt_at", "VerpacktAt"]

# Scan-Datum (Abholer_DB)
COL_SCAN_DATE   = ["Scan-Datum", "Scan Datum", "ScanDatum", "Scan_Datum", "Scan-Date"]

# DHL_Express
DHL_COL_BARCODE = ["Package Barcode", "Paket-Barcode", "Paket Barcode", "Paketbarcode", "Barcode"]
DHL_COL_SCAN    = ["Date Of Scan", "Date of Scan", "DateOfScan", "Scan Date", "ScanDate"]

# DHL (heute) Merge / Orca-Export
ORCA_COL_BARCODE = ["Package Barcode", "Paket-Barcode", "Paket Barcode", "Paketbarcode", "Barcode", "barcode"]
ORCA_COL_SCAN    = ["Date of Scan", "Date Of Scan", "DateOfScan", "Scan Date", "ScanDate", "date"]

# Tagesbote Spalten
COL_TB_LIEFERUNG      = ["Lieferung", "Lieferung "]
COL_TB_KONTROLLSTATUS = ["Kontrollstatus", "Kontrollstatus "]
COL_TB_BARCODE        = ["Paket-Barcode", "Paket Barcode", "Paketbarcode", "Barcode"]
COL_TB_NAME           = ["Name", "Kunde"]
COL_TB_DATUM          = ["Datum", "Date", "Lieferdatum", "Scan-Datum", "Scan Date"]
COL_TB_STATUS         = ["Status"]
COL_TB_ZIELKIOSK      = ["Ziel-Kiosk", "Ziel Kiosk", "Ziel_Kiosk", "Zielkiosk"]
COL_TB_ZAHLUNG        = ["Zahlung"]

# OrcaScan API Konfiguration (aus .env geladen)
import os as _os
try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv(dotenv_path=BASE_DIR / ".env")
except ImportError:
    # python-dotenv nicht installiert → automatisch nachinstallieren
    import subprocess as _sp
    _sp.run([_os.sys.executable if hasattr(_os, "sys") else "python", "-m", "pip", "install", "python-dotenv"], check=False)
    try:
        from dotenv import load_dotenv as _load_dotenv
        _load_dotenv(dotenv_path=BASE_DIR / ".env")
    except Exception:
        pass  # Notfall: Bombadil startet trotzdem, API-Calls schlagen fehl

_env_missing = not (BASE_DIR / ".env").exists()
ORCA_API_KEY             = _os.getenv("ORCA_API_KEY")
ORCA_BASE_URL            = _os.getenv("ORCA_BASE_URL")
ORCA_ABHOLER_SHEET_ID    = _os.getenv("ORCA_ABHOLER_SHEET_ID")    # Abholer_DB
ORCA_DHL_NORMAL_SHEET_ID = _os.getenv("ORCA_DHL_NORMAL_SHEET_ID") # DHL_Normal
ORCA_DHL_EX_SHEET_ID     = _os.getenv("ORCA_DHL_EX_SHEET_ID")     # DHL_Express
ORCA_TAGESBOTE_SHEET_ID  = _os.getenv("ORCA_TAGESBOTE_SHEET_ID")  # Tagesbote

if _env_missing:
    import tkinter as _tk
    from tkinter import messagebox as _mb
    _r = _tk.Tk(); _r.withdraw()
    _mb.showwarning("Konfiguration fehlt",
        "Die Datei '.env' wurde nicht gefunden!\n\n"
        "Bitte die .env-Datei in den Bombadil-Ordner legen.\n"
        "(Bei Marcel oder IT erfragen)")
    _r.destroy()

# Feste Abfahrtszeiten der Tagesboten-Touren (Lokalzeit, UTC+1)
TOUR_1_ABFAHRT = (11, 14)   # nur informativ
TOUR_2_ABFAHRT = (13, 45)   # nur informativ

# ── Tour-Zeiten Persistenz (tagesaktuell) ────────────────────────────────────
import json as _json_mod

def _tour_zeiten_pfad():
    import datetime as _dt_tz
    heute = (_dt_tz.datetime.now()).strftime("%Y-%m-%d")
    return BASE_DIR / f"tour_zeiten_{heute}.json"

def _load_tour_zeiten() -> dict:
    """Gibt {"t1": "HH:MM"|None, "t2": "HH:MM"|None} für den heutigen Tag zurück."""
    try:
        p = _tour_zeiten_pfad()
        if p.exists():
            return _json_mod.loads(p.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {"t1": None, "t2": None}

def _save_tour_zeiten(t1, t2, t1_barcodes=None, t2_barcodes=None):
    try:
        tz = _load_tour_zeiten()
        _tour_zeiten_pfad().write_text(
            _json_mod.dumps({
                "t1": t1, "t2": t2,
                "t1_barcodes": t1_barcodes if t1_barcodes is not None else tz.get("t1_barcodes", []),
                "t2_barcodes": t2_barcodes if t2_barcodes is not None else tz.get("t2_barcodes", []),
            }), encoding="utf-8")
    except Exception:
        pass

# Google Drive Konfiguration
GDRIVE_FOLDER_ID         = "1a5Wg-fFhF11ux5d7Tl5oVqcq9fRkp5yX"   # Tagesbote Upload
TOURLISTEN_DIR           = Path(r"W:\Automatisierungen\16_EMMA\7_CAF\Buchen\Emma-3")
GDRIVE_ABHOLER_FOLDER_ID = "1OSnMmPf--uqt4ulDGy3ILT61pUuCPzpn"   # Abholer_DB Archiv
TAGESBOTE_SEARCH_NAME = "tagesbote"
SERVICE_ACCOUNT_FILE  = BASE_DIR / "service_account.json"
OAUTH_CREDENTIALS_FILE = BASE_DIR / "oauth_credentials.json"
OAUTH_TOKEN_FILE       = BASE_DIR / "token.json"
GDRIVE_SCOPES         = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.file",      # für Upload
]
GDRIVE_UPLOAD_SCOPES  = [
    "https://www.googleapis.com/auth/drive.file",      # nur Upload/eigene Dateien
]

DE_WOCHENTAGE = {0: "Mo", 1: "Di", 2: "Mi", 3: "Do", 4: "Fr", 5: "Sa", 6: "So"}


# -------------------------
# Helper
# -------------------------
def first_existing(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def today_date():
    if TZ:
        return datetime.now(TZ).date()
    return date.today()


def newest_excel(folder: str) -> str | None:
    """Gibt die neueste Abholer_DB*.xlsx/.xls im Ordner zurück (ignoriert alle anderen Dateien)."""
    try:
        p = Path(folder)
        if not p.exists():
            return None
        files = list(p.glob("Abholer_DB*.xlsx")) + list(p.glob("Abholer_DB*.xls"))
        if not files:
            return None
        return str(max(files, key=lambda f: f.stat().st_mtime))
    except Exception:
        return None


def fmt_dt(v) -> str:
    if pd.isna(v):
        return ""
    try:
        if hasattr(v, "to_pydatetime"):
            v = v.to_pydatetime()
        if isinstance(v, datetime):
            # UTC → lokale Zeit (Europe/Berlin) umrechnen
            if v.tzinfo is not None and TZ is not None:
                v = v.astimezone(TZ).replace(tzinfo=None)
            return v.strftime("%d.%m.%Y %H:%M")
        return str(v)
    except Exception:
        return str(v)


def norm_str(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip().lower()


def read_any_table(path: str) -> pd.DataFrame:
    """Excel oder CSV robust einlesen (CSV: mehrere Encodings und Separatoren)"""
    p   = Path(path)
    ext = p.suffix.lower()

    if ext in [".xlsx", ".xls"]:
        return pd.read_excel(path)

    if ext == ".csv":
        encodings = ["utf-8-sig", "utf-8", "cp1252"]
        seps      = [";", ",", "\t"]
        last_err  = None
        for enc in encodings:
            for sep in seps:
                try:
                    df = pd.read_csv(path, sep=sep, encoding=enc)
                    if df.shape[1] >= 2:
                        return df
                except Exception as e:
                    last_err = e
        raise ValueError(f"CSV konnte nicht gelesen werden: {path}\nLetzter Fehler: {last_err}")

    raise ValueError(f"Dateityp nicht unterstützt: {ext} ({path})")


# ============================================================
# Barcode-Cleaner
# ============================================================
def clean_barcode(value) -> str:
    if pd.isna(value):
        return ""
    s = str(value).strip()
    if s.startswith('="') and s.endswith('"') and len(s) >= 4:
        s = s[2:-1].strip()
    if s.startswith("'"):
        s = s[1:].lstrip()
    s = s.strip()
    # "158892672.0" → "158892672" (OrcaScan liefert Zahlen als float)
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    return s


# ============================================================
# Robuste Dateisuche: DHL_Normal
# ============================================================
def _filename_tokens(name: str) -> list[str]:
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.split(" ") if s else []


def find_latest_dhl_normal(folder: Path) -> Path | None:
    """Findet die neueste DHL_Normal-Datei im Ordner (enthält 'dhl' + 'normal' im Namen)."""
    exts = {".xlsx", ".xls", ".csv"}
    try:
        files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in exts]
    except Exception:
        return None

    hits = [p for p in files
            if "dhl" in _filename_tokens(p.name) and "normal" in _filename_tokens(p.name)]
    if not hits:
        return None
    return max(hits, key=lambda f: f.stat().st_mtime)


# ============================================================
# Excel-Export Helper: Spalte(n) als Text formatieren (@)
# ============================================================
def write_excel_text_cols(df: pd.DataFrame, out_path: Path, text_cols: list[str]):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(out_path, index=False)

    wb     = load_workbook(out_path)
    ws     = wb.active
    header = [cell.value for cell in ws[1]]
    col_idx = {name: i for i, name in enumerate(header, start=1) if name in text_cols}

    for name, idx in col_idx.items():
        ws.cell(row=1, column=idx).number_format = "@"
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if cell.value is not None:
                cell.value        = str(cell.value)
                cell.number_format = "@"

    wb.save(out_path)


def backup_abholer_db(export_folder: Path) -> Path:
    """Lädt die komplette Abholer_DB und speichert sie als Excel-Backup."""
    today = date.today().strftime("%Y-%m-%d")
    backup_dir = export_folder / "Backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    out_path = backup_dir / f"Abholer_DB_Backup_{today}.xlsx"
    df = fetch_abholer_orca()
    write_excel_text_cols(df, out_path, text_cols=list(df.columns))
    return out_path


# ============================================================
# Kernlogik – Hauptliste
# ============================================================
def compute_all_rows(source):
    df = source if isinstance(source, pd.DataFrame) else pd.read_excel(source)

    c_status      = first_existing(df, COL_STATUS)
    c_abholbereit = first_existing(df, COL_ABHOLBEREIT)
    c_abgeholt    = first_existing(df, COL_ABGEHOLT)
    c_barcode     = first_existing(df, COL_BARCODE)
    c_name        = first_existing(df, COL_NAME)
    c_zahlung     = first_existing(df, COL_ZAHLUNG)
    c_zielkiosk   = first_existing(df, COL_ZIELKIOSK)
    c_bestellwert = first_existing(df, COL_BESTELLWERT)
    c_verpackt_at = first_existing(df, COL_VERPACKT_AT)

    # Pflichtfelder – ohne diese kann gar nichts berechnet werden
    missing_required = [n for n, c in {
        "Paket-Barcode": c_barcode,
        "Paketstatus":   c_status,
    }.items() if c is None]
    if missing_required:
        raise ValueError(
            "Fehlende Pflicht-Spalten: " + ", ".join(missing_required) +
            "\nGefundene Spalten: " + ", ".join(map(str, df.columns))
        )

    # Optionale Spalten: fehlende als leere Spalte ergänzen statt Fehler
    for col_name, col_ref, col_list in [
        ("Abholbereit_At", c_abholbereit, COL_ABHOLBEREIT),
        ("Abgeholt_At",    c_abgeholt,    COL_ABGEHOLT),
        ("Zahlung",        c_zahlung,     COL_ZAHLUNG),
        ("Ziel-Kiosk",     c_zielkiosk,   COL_ZIELKIOSK),
        ("Name",           c_name,        COL_NAME),
    ]:
        if col_ref is None:
            df[col_name] = ""
            # Zeiger auf den neuen Spaltennamen setzen
            if col_list is COL_ABHOLBEREIT:
                c_abholbereit = col_name
            elif col_list is COL_ABGEHOLT:
                c_abgeholt = col_name
            elif col_list is COL_ZAHLUNG:
                c_zahlung = col_name
            elif col_list is COL_ZIELKIOSK:
                c_zielkiosk = col_name
            elif col_list is COL_NAME:
                c_name = col_name

    df[c_abholbereit] = pd.to_datetime(df[c_abholbereit], errors="coerce")
    df[c_abgeholt]    = pd.to_datetime(df[c_abgeholt],    errors="coerce")
    if c_verpackt_at:
        df[c_verpackt_at] = pd.to_datetime(df[c_verpackt_at], errors="coerce")

    today           = today_date()
    yesterday       = today - timedelta(days=1)
    seven_days_ago  = today - timedelta(days=7)
    three_weeks_ago = today - timedelta(days=21)

    status_norm = df[c_status].astype(str).str.strip().str.lower()

    # ── Tabs: volle Daten (kein Datumsfilter) ───────────────────────
    abholbereit_df             = df[status_norm.eq("abholbereit") & df[c_abgeholt].isna()].copy()
    older7_df                  = abholbereit_df[abholbereit_df[c_abholbereit].dt.date < seven_days_ago].copy()
    yesterday_df               = df[df[c_abgeholt].dt.date == yesterday].copy()
    verpackt_no_abholbereit_df = df[status_norm.eq("verpackt") & df[c_abholbereit].isna()].copy()

    # ── Report-Zählungen: nur letzte 7 Tage ─────────────────────────
    df7        = df[df[c_abholbereit].dt.date >= seven_days_ago].copy()
    status7    = df7[c_status].astype(str).str.strip().str.lower()
    n_abholbereit = len(df7[status7.eq("abholbereit")])
    n_older7      = len(older7_df)
    n_yesterday        = len(yesterday_df)
    day_before_yest    = yesterday - timedelta(days=1)
    day_before_yest_df = df[df[c_abgeholt].dt.date == day_before_yest].copy()
    n_day_before_yest  = len(day_before_yest_df)

    # Verpackt-Kachel: zeigt aktuelle Pakete mit Status "Verpackt" ohne Abholbereit_At
    # (= dieselbe Menge wie die Verpackt-Tabelle)
    n_verpackt = len(verpackt_no_abholbereit_df)

    def rows_abhol(dfx):
        """Abholbereit-Zeilen: Barcode, Abholbereit_At, Name, Ziel-Kiosk, Wartezeit – älteste zuerst."""
        dfx = dfx.sort_values(by=c_abholbereit, ascending=True, na_position="last")
        out = []
        for _, r in dfx.iterrows():
            bc = "" if pd.isna(r[c_barcode]) else str(r[c_barcode]).strip()
            if not bc:
                continue
            nm = "" if pd.isna(r[c_name]) else str(r[c_name]).strip()
            zk = "" if pd.isna(r[c_zielkiosk]) else str(r[c_zielkiosk]).strip()
            dt_raw = r[c_abholbereit]
            dt_str = fmt_dt(dt_raw)
            if pd.notna(dt_raw):
                try:
                    days = (today - pd.Timestamp(dt_raw).normalize().date()).days
                    wt = f"{days} Tage" if days != 1 else "1 Tag"
                except Exception:
                    wt = ""
            else:
                wt = "?"
            oid = "" if "_id" not in r.index or pd.isna(r["_id"]) else str(r["_id"]).strip()
            out.append((bc, dt_str, nm, zk, wt, oid))   # oid an Index 5 (versteckt)
        return out

    def rows_older7(dfx):
        """älter als 7 Tage: Barcode, Abholbereit_At, Name, Ziel-Kiosk, Wartezeit – älteste zuerst."""
        dfx = dfx.sort_values(by=c_abholbereit, ascending=True, na_position="last")
        out = []
        for _, r in dfx.iterrows():
            bc = "" if pd.isna(r[c_barcode]) else str(r[c_barcode]).strip()
            if not bc:
                continue
            nm = "" if pd.isna(r[c_name])      else str(r[c_name]).strip()
            zk = "" if pd.isna(r[c_zielkiosk]) else str(r[c_zielkiosk]).strip()
            dt_raw = r[c_abholbereit]
            dt_str = fmt_dt(dt_raw)
            if pd.notna(dt_raw):
                try:
                    days = (today - pd.Timestamp(dt_raw).normalize().date()).days
                    wt = f"{days} Tage" if days != 1 else "1 Tag"
                except Exception:
                    wt = ""
            else:
                wt = "?"
            out.append((bc, nm, dt_str, zk, wt))
        return out

    # Zahlung offen
    zahlung_norm     = df[c_zahlung].apply(norm_str)
    keywords_pay     = ["unbezahlt", "offen", "vor ort", "vor-ort", "vorort"]

    def has_keyword(s: str) -> bool:
        return bool(s) and any(k in s for k in keywords_pay)

    zahlung_mask      = zahlung_norm.apply(has_keyword)
    zahlung_offen_df  = df[zahlung_mask].copy()

    def _status_sort(s):
        sl = str(s).lower()
        if sl == "abholbereit": return 0
        if "verpackt" in sl:    return 1
        if sl == "abgeholt":    return 3   # ganz nach unten
        return 2                            # offene/sonstige unter die Verpackten

    zahlung_offen_df["_sort"] = zahlung_offen_df[c_status].apply(_status_sort)
    zahlung_offen_df = zahlung_offen_df.sort_values(
        ["_sort", c_abholbereit], ascending=[True, True], na_position="last")

    rows_pay = []
    for _, r in zahlung_offen_df.iterrows():
        bc = "" if pd.isna(r[c_barcode]) else str(r[c_barcode]).strip()
        if not bc:
            continue
        nm         = "" if pd.isna(r[c_name])      else str(r[c_name]).strip()
        z_raw      = "" if pd.isna(r[c_zahlung])   else str(r[c_zahlung]).strip()
        status_raw = "" if pd.isna(r[c_status])    else str(r[c_status]).strip()
        bw         = ("" if c_bestellwert is None or pd.isna(r[c_bestellwert])
                      else str(r[c_bestellwert]).strip())
        oid        = "" if "_id" not in r.index or pd.isna(r["_id"]) else str(r["_id"]).strip()
        dt_raw     = r[c_abholbereit]
        if pd.notna(dt_raw):
            try:
                days = (today - pd.Timestamp(dt_raw).normalize().date()).days
                wt = f"{days} Tage" if days != 1 else "1 Tag"
            except Exception:
                wt = ""
        else:
            wt = "?"
        # Reihenfolge: bc, nm, bw, status, dt, wt, oid
        # Indizes:       0   1   2    3     4   5   6
        _ = z_raw  # nicht mehr in UI, bleibt aber in DB-Spalte 'Zahlung'
        rows_pay.append((bc, nm, bw, status_raw, fmt_dt(dt_raw), wt, oid))

    def rows_yesterday(dfx):
        """Gestern abgeholt: Barcode, Name, Abgeholt_At, Ziel-Kiosk – neueste zuerst."""
        dfx = dfx.sort_values(by=c_abgeholt, ascending=False, na_position="last")
        out = []
        for _, r in dfx.iterrows():
            bc = "" if pd.isna(r[c_barcode])    else str(r[c_barcode]).strip()
            if not bc:
                continue
            nm = "" if pd.isna(r[c_name])       else str(r[c_name]).strip()
            zk = "" if pd.isna(r[c_zielkiosk])  else str(r[c_zielkiosk]).strip()
            out.append((bc, nm, fmt_dt(r[c_abgeholt]), zk))
        return out

    # Kissel > 3 Wochen
    ziel_norm   = df[c_zielkiosk].apply(norm_str)
    kissel_mask = ziel_norm.str.contains("kissel", na=False)
    kissel_df   = df[
        status_norm.eq("abholbereit")
        & kissel_mask
        & df[c_abholbereit].notna()
        & (df[c_abholbereit].dt.date < three_weeks_ago)
    ].copy()

    kissel_df = kissel_df.sort_values(by=c_abholbereit, ascending=True, na_position="last")
    rows_kissel = []
    for _, r in kissel_df.iterrows():
        bc = "" if pd.isna(r[c_barcode])    else str(r[c_barcode]).strip()
        if not bc:
            continue
        nm  = "" if pd.isna(r[c_name])      else str(r[c_name]).strip()
        zk  = "" if pd.isna(r[c_zielkiosk]) else str(r[c_zielkiosk]).strip()
        dt_raw = r[c_abholbereit]
        dt_str = fmt_dt(dt_raw)
        if pd.notna(dt_raw):
            try:
                days = (today - pd.Timestamp(dt_raw).normalize().date()).days
                wt = f"{days} Tage" if days != 1 else "1 Tag"
            except Exception:
                wt = ""
        else:
            wt = "?"
        rows_kissel.append((bc, nm, dt_str, zk, wt))

    # Verpackt ohne Abholbereit_At – sortiert älteste zuerst
    def rows_verpackt_fn(dfx):
        if c_verpackt_at:
            dfx = dfx.sort_values(by=c_verpackt_at, ascending=True, na_position="last")
        out = []
        for _, r in dfx.iterrows():
            bc = "" if pd.isna(r[c_barcode]) else str(r[c_barcode]).strip()
            if not bc:
                continue
            nm = "" if pd.isna(r[c_name])      else str(r[c_name]).strip()
            zk = "" if pd.isna(r[c_zielkiosk]) else str(r[c_zielkiosk]).strip()
            dt_raw = r[c_verpackt_at] if c_verpackt_at else None
            dt_str = fmt_dt(dt_raw) if dt_raw is not None else ""
            if dt_raw is not None and pd.notna(dt_raw):
                try:
                    days = (today - pd.Timestamp(dt_raw).normalize().date()).days
                    wt = f"{days} Tage" if days != 1 else "1 Tag"
                except Exception:
                    wt = ""
            else:
                wt = "?"
            out.append((bc, dt_str, nm, zk, wt))
        return out
    rows_verpackt = rows_verpackt_fn(verpackt_no_abholbereit_df)

    # Abholungen letzte 7 Tage – für Balkendiagramm
    # Abgeholt_At ist oft leer → Paketstatus="abgeholt" mit Fallback auf Abholbereit_At
    _abgeholt_df = df[status_norm == "abgeholt"].copy()
    _eff_date = _abgeholt_df[c_abgeholt].fillna(_abgeholt_df[c_abholbereit])
    _abgeholt_df["_eff_date"] = _eff_date
    _abgeholt_df = _abgeholt_df.dropna(subset=["_eff_date"])
    _abgeholt_df["_date"] = _abgeholt_df["_eff_date"].dt.date
    abgeholt_by_day = _abgeholt_df.groupby("_date").size()
    daily_7 = [
        (today - timedelta(days=i),
         int(abgeholt_by_day.get(today - timedelta(days=i), 0)))
        for i in range(6, -1, -1)   # ältester Tag zuerst
        if (today - timedelta(days=i)).weekday() != 6   # kein Sonntag
    ]

    report_data = {
        "abholbereit": len(abholbereit_df),
        "verpackt":    n_verpackt,
        "pay":         len(zahlung_offen_df),
        "older7":      len(older7_df),
        "kissel":      len(kissel_df),
        "yesterday":          n_yesterday,
        "day_before_yest":    n_day_before_yest,
        "daily7":      daily_7,
        "today":       today,
    }

    return (
        report_data,
        rows_abhol(abholbereit_df),
        rows_older7(older7_df),
        rows_yesterday(yesterday_df),
        rows_pay,
        rows_kissel,
        rows_verpackt,
    )


# ============================================================
# Kernlogik – DHL_Express (heute)
# ============================================================
def compute_dhl_today_rows(source):
    df = source if isinstance(source, pd.DataFrame) else pd.read_excel(source)

    c_barcode = first_existing(df, DHL_COL_BARCODE)
    c_scan    = first_existing(df, DHL_COL_SCAN)

    missing = [n for n, c in {"Package Barcode": c_barcode, "Date Of Scan": c_scan}.items() if c is None]
    if missing:
        raise ValueError(
            "DHL_Express – fehlende Spalten: " + ", ".join(missing) +
            "\nGefundene Spalten: " + ", ".join(map(str, df.columns))
        )

    df[c_scan] = pd.to_datetime(df[c_scan], errors="coerce")

    today    = today_date()
    today_df = df[df[c_scan].dt.date == today].copy()

    rows = []
    for _, r in today_df.iterrows():
        bc = clean_barcode(r[c_barcode])
        if not bc:
            continue
        rows.append((bc, fmt_dt(r[c_scan])))

    return rows


# ============================================================
# Kernlogik – DHL heute: DHL_Normal Sheet, wochentag-abhängiges Zeitfenster
# ============================================================
def _dhl_zeitfenster(heute=None):
    """
    Liefert (start_dt, end_dt) für den DHL-Export-Zeitraum, abhängig vom Wochentag.

    Cutoff-Zeiten:
      Mo–Fr:   16:15 (DHL-Abholung)
      Sa:      14:00 (DHL-Abholung)
      So:      kein Cutoff (wird als Mo behandelt – fallback)

    Fenster:
      Di–Fr:    gestern 16:15  → heute 16:15
      Sa:       Fr 16:15       → heute 14:00
      Mo:       Sa 14:00       → heute 16:15   (überspringt Sonntag + Sa nach 14:00)
      So:       Sa 14:00       → So 23:59     (selten – kein Arbeitstag)
    """
    import datetime as _dt
    if heute is None:
        heute = _dt.date.today()
    wd = heute.weekday()   # Mo=0 … So=6

    # End-Cutoff (heute)
    if wd == 5:                              # Samstag
        end_dt = _dt.datetime.combine(heute, _dt.time(14, 0))
    elif wd == 6:                            # Sonntag (Fallback)
        end_dt = _dt.datetime.combine(heute, _dt.time(23, 59))
    else:                                    # Mo–Fr
        end_dt = _dt.datetime.combine(heute, _dt.time(16, 15))

    # Start-Cutoff (Vortag oder weiter zurück)
    if wd == 0:                              # Montag → Samstag 14:00
        sa = heute - _dt.timedelta(days=2)
        start_dt = _dt.datetime.combine(sa, _dt.time(14, 0))
    elif wd == 6:                            # Sonntag → Samstag 14:00
        sa = heute - _dt.timedelta(days=1)
        start_dt = _dt.datetime.combine(sa, _dt.time(14, 0))
    elif wd == 5:                            # Samstag → Freitag 16:15
        fr = heute - _dt.timedelta(days=1)
        start_dt = _dt.datetime.combine(fr, _dt.time(16, 15))
    else:                                    # Di–Fr → gestern 16:15
        gestern = heute - _dt.timedelta(days=1)
        start_dt = _dt.datetime.combine(gestern, _dt.time(16, 15))

    return start_dt, end_dt


def compute_dhl_normal_today(source):
    df = source if isinstance(source, pd.DataFrame) else read_any_table(source)

    c_barcode = first_existing(df, ORCA_COL_BARCODE)
    c_scan    = first_existing(df, ORCA_COL_SCAN)

    missing = [n for n, c in {"Package Barcode": c_barcode, "Date of Scan": c_scan}.items() if c is None]
    if missing:
        raise ValueError(
            "DHL_Normal – fehlende Spalten: " + ", ".join(missing) +
            "\nGefundene Spalten: " + ", ".join(map(str, df.columns))
        )

    x = df[[c_barcode, c_scan]].copy()
    x.columns = ["Package Barcode", "Date of Scan"]
    x["Date of Scan"]    = pd.to_datetime(x["Date of Scan"], errors="coerce")
    if x["Date of Scan"].dt.tz is not None:
        x["Date of Scan"] = x["Date of Scan"].dt.tz_convert(None)
    x["Package Barcode"] = x["Package Barcode"].apply(clean_barcode)
    x = x[x["Package Barcode"].astype(str).str.len() > 0]

    # DHL-Zeitfenster je nach Wochentag (lokale Systemzeit)
    start_dt, end_dt = _dhl_zeitfenster()
    x_today = x[
        (x["Date of Scan"] >= start_dt) &
        (x["Date of Scan"] <= end_dt)
    ].copy()
    x_today.drop_duplicates(subset=["Package Barcode", "Date of Scan"], inplace=True)
    x_today.sort_values(by=["Package Barcode", "Date of Scan"], inplace=True)

    rows = [(str(r["Package Barcode"]), fmt_dt(r["Date of Scan"])) for _, r in x_today.iterrows()]
    return x_today, rows


# Alias für Rückwärtskompatibilität (manuelle Dateiauswahl)
def compute_orca_merge_today(source1, source2=None):
    return compute_dhl_normal_today(source1)


def add_working_days(d: date, n: int) -> date:
    """Addiert n Werktage (Mo–Fr) zu d. Negative n subtrahieren."""
    step      = 1 if n >= 0 else -1
    remaining = abs(n)
    current   = d
    while remaining > 0:
        current += timedelta(days=step)
        if current.weekday() < 5:   # 0=Mo … 4=Fr
            remaining -= 1
    return current


def _get_oauth_drive_service():
    """
    Gibt einen authentifizierten Google Drive Service zurück (OAuth2).
    Beim ersten Aufruf öffnet sich ein Browser-Fenster zur Anmeldung.
    Das Token wird in token.json gespeichert und danach automatisch erneuert.
    """
    if not OAUTH_AVAILABLE:
        raise RuntimeError(
            "OAuth2-Bibliothek fehlt.\n"
            "Bitte ausführen: pip install google-auth-oauthlib"
        )
    if not OAUTH_CREDENTIALS_FILE.exists():
        raise FileNotFoundError(
            f"oauth_credentials.json nicht gefunden: {OAUTH_CREDENTIALS_FILE}\n\n"
            "Schritte:\n"
            "1. Google Cloud Console → APIs & Dienste → Anmeldedaten\n"
            "2. OAuth 2.0 Client-ID erstellen (Desktop-App)\n"
            "3. JSON herunterladen → als 'oauth_credentials.json' im Bombadil-Ordner speichern"
        )

    import io as _io
    creds = None

    # Gespeichertes Token laden (falls vorhanden)
    if OAUTH_TOKEN_FILE.exists():
        try:
            with open(OAUTH_TOKEN_FILE, "rb") as fh:
                creds = _pickle.load(fh)
        except Exception:
            creds = None

    # Token abgelaufen → automatisch erneuern
    if creds and creds.expired and creds.refresh_token:
        try:
            creds.refresh(_GAuthRequest())
        except Exception as _ref_err:
            # Token widerrufen oder abgelaufen (invalid_grant) → löschen, neu einloggen
            try:
                OAUTH_TOKEN_FILE.unlink(missing_ok=True)
            except Exception:
                pass
            creds = None

    if not creds or not creds.valid:
        # Kein gültiges Token → Browser öffnet sich für neuen Login
        flow  = InstalledAppFlow.from_client_secrets_file(
            str(OAUTH_CREDENTIALS_FILE), GDRIVE_UPLOAD_SCOPES
        )
        creds = flow.run_local_server(port=0)

    # Token speichern für nächsten Start
    with open(OAUTH_TOKEN_FILE, "wb") as fh:
        _pickle.dump(creds, fh)

    return build("drive", "v3", credentials=creds, cache_discovery=False)


def upload_excel_to_gdrive(df: pd.DataFrame, folder_id: str, filename: str):
    """Lädt einen DataFrame als .xlsx in den angegebenen Google Drive Ordner hoch (OAuth2)."""
    if not GDRIVE_AVAILABLE:
        raise RuntimeError("Google API nicht verfügbar (pip install google-api-python-client google-auth).")

    import io as _io
    service = _get_oauth_drive_service()

    # Timezone-aware Spalten für Excel-Export entfernen (openpyxl unterstützt keine tz-aware Timestamps)
    df_export = df.copy()
    for col in df_export.select_dtypes(include=["datetimetz"]).columns:
        df_export[col] = df_export[col].dt.tz_localize(None)

    buf = _io.BytesIO()
    df_export.to_excel(buf, index=False)
    buf.seek(0)

    metadata = {"name": filename, "parents": [folder_id]}
    media    = MediaIoBaseUpload(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    service.files().create(
        body=metadata, media_body=media,
        fields="id",
    ).execute()


def upload_json_to_gdrive(data: dict, folder_id: str, filename: str):
    """Lädt ein dict als JSON-Datei nach Google Drive hoch (überschreibt gleichnamige Datei)."""
    if not GDRIVE_AVAILABLE:
        return
    import io as _io
    service = _get_oauth_drive_service()
    content = _json_mod.dumps(data, ensure_ascii=False).encode("utf-8")
    # Alte Datei gleichen Namens löschen (Drive erlaubt Duplikate → wir wollen nur eine)
    q = f"name = '{filename}' and '{folder_id}' in parents and trashed = false"
    existing = service.files().list(q=q, fields="files(id)", pageSize=5).execute()
    for f in existing.get("files", []):
        try:
            service.files().delete(fileId=f["id"]).execute()
        except Exception:
            pass
    buf = _io.BytesIO(content)
    metadata = {"name": filename, "parents": [folder_id]}
    media = MediaIoBaseUpload(buf, mimetype="application/json", resumable=False)
    service.files().create(body=metadata, media_body=media, fields="id").execute()


def download_json_from_gdrive(folder_id: str, filename: str) -> "dict | None":
    """Lädt eine JSON-Datei aus Google Drive herunter. Gibt None zurück wenn nicht gefunden."""
    if not GDRIVE_AVAILABLE:
        return None
    try:
        service = _get_oauth_drive_service()
        q = f"name = '{filename}' and '{folder_id}' in parents and trashed = false"
        res = service.files().list(
            q=q, fields="files(id)", pageSize=5,
            orderBy="modifiedTime desc",
        ).execute()
        files = res.get("files", [])
        if not files:
            return None
        content = service.files().get_media(fileId=files[0]["id"]).execute()
        return _json_mod.loads(content.decode("utf-8"))
    except Exception:
        return None


# ============================================================
# Google Drive – Tagesbote laden
# ============================================================
def fetch_tagesbote_gdrive() -> "pd.DataFrame":
    """
    Lädt tagesbote.xlsx aus dem Google Drive Ordner 'Tagesbote Upload'.
    Benötigt service_account.json im Bombadil-Ordner.
    Unterstützt echte .xlsx-Dateien UND Google Sheets (werden als xlsx exportiert).
    """
    if not GDRIVE_AVAILABLE:
        raise RuntimeError(
            "Google API nicht verfügbar.\n\n"
            "Bitte installieren:\n"
            "  pip install google-api-python-client google-auth"
        )

    if not SERVICE_ACCOUNT_FILE.exists():
        raise FileNotFoundError(
            f"Service Account Datei nicht gefunden:\n{SERVICE_ACCOUNT_FILE}\n\n"
            "Schritte:\n"
            "1. Google Cloud Console → Service Account erstellen\n"
            "2. JSON-Schlüssel herunterladen\n"
            "3. Datei umbenennen in 'service_account.json'\n"
            "4. In Bombadil-Ordner legen\n"
            "5. Google Drive Ordner mit Service-Account-Email teilen"
        )

    creds   = service_account.Credentials.from_service_account_file(
        str(SERVICE_ACCOUNT_FILE), scopes=GDRIVE_SCOPES
    )
    service = build("drive", "v3", credentials=creds, cache_discovery=False)

    # Suche nach tagesbote-Datei im Ordner (neueste zuerst)
    # supportsAllDrives + includeItemsFromAllDrives = nötig für geteilte Laufwerke (Shared Drive)
    results = service.files().list(
        q=(
            f"'{GDRIVE_FOLDER_ID}' in parents "
            f"and name contains '{TAGESBOTE_SEARCH_NAME}' "
            f"and trashed = false"
        ),
        fields="files(id, name, mimeType)",
        orderBy="modifiedTime desc",
        pageSize=5,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()

    files = results.get("files", [])
    if not files:
        raise FileNotFoundError(
            "Keine Tagesbote-Datei im Google Drive Ordner gefunden.\n\n"
            f"Ordner-ID: {GDRIVE_FOLDER_ID}\n"
            f"Gesucht nach: '{TAGESBOTE_SEARCH_NAME}'\n\n"
            "Bitte prüfen:\n"
            "• Ist die Datei im richtigen Ordner?\n"
            "• Ist der Ordner mit dem Service Account geteilt?"
        )

    # Alle gefundenen Tagesbote-Dateien laden und zusammenführen
    all_frames = []
    for file in files:
        file_id   = file["id"]
        mime_type = file["mimeType"]

        # Google Sheet → als xlsx exportieren; echte xlsx → direkt herunterladen
        if mime_type == "application/vnd.google-apps.spreadsheet":
            request = service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            request = service.files().get_media(fileId=file_id)

        buf = _io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()

        buf.seek(0)
        try:
            df = pd.read_excel(buf)
            if not df.empty:
                all_frames.append(df)
        except Exception:
            pass  # Fehlerhafte Datei überspringen

    if not all_frames:
        raise FileNotFoundError("Keine lesbare Tagesbote-Datei im Google Drive Ordner gefunden.")

    return pd.concat(all_frames, ignore_index=True)


# ============================================================
# ============================================================
# OrcaScan API – Abholer_DB laden
# ============================================================
def fetch_abholer_orca() -> "pd.DataFrame":
    """Lädt die Abholer_DB direkt aus OrcaScan per API (alle Seiten).
    Bei HTTP 429 (Too Many Requests) wird bis zu 3× mit Wartezeit wiederholt."""
    import urllib.request as _urllib
    import urllib.error   as _urllib_err
    import json as _json
    import time as _time

    DROP_COLS = {"Unterschrift", "Paketfoto"}
    all_rows  = []
    seen_ids  = set()
    page      = 1
    MAX_PAGES = 50

    while page <= MAX_PAGES:
        url = f"{ORCA_BASE_URL}/sheets/{ORCA_ABHOLER_SHEET_ID}/rows?withTitles=true&page={page}&limit=500"
        req = _urllib.Request(url, headers={"Authorization": f"Bearer {ORCA_API_KEY}"})

        for attempt in range(4):
            try:
                with _urllib.urlopen(req, timeout=60) as resp:
                    data = _json.loads(resp.read())
                break
            except _urllib_err.HTTPError as e:
                if e.code == 429 and attempt < 3:
                    _time.sleep((attempt + 1) * 3)
                    continue
                raise RuntimeError(
                    f"OrcaScan API Fehler (Seite {page}): HTTP Error {e.code}: {e.reason}"
                ) from e
            except Exception as e:
                raise RuntimeError(f"OrcaScan API Fehler (Seite {page}): {e}") from e

        rows = data.get("data", [])
        if not rows:
            break

        # Duplikat-Erkennung: wenn alle IDs schon bekannt → API ignoriert page-Parameter
        new_ids = {r.get("_id") for r in rows if r.get("_id")}
        if new_ids and new_ids.issubset(seen_ids):
            break
        seen_ids.update(new_ids)

        for r in rows:
            for col in DROP_COLS:
                r.pop(col, None)
        all_rows.extend(rows)

        if len(rows) < 500:
            break

        page += 1
        _time.sleep(0.4)   # kurze Pause zwischen Seiten

    if not all_rows:
        raise RuntimeError("Abholer_DB via OrcaScan API: Keine Daten erhalten.")

    return pd.DataFrame(all_rows)


def fetch_sheet_orca(sheet_id: str, drop_cols: set = None) -> "pd.DataFrame":
    """Generisch: lädt beliebiges OrcaScan-Sheet per API (alle Seiten).
    Bei HTTP 429 (Too Many Requests) wird bis zu 3× mit Wartezeit wiederholt."""
    import urllib.request as _urllib
    import urllib.error   as _urllib_err
    import json as _json
    import time as _time

    url = f"{ORCA_BASE_URL}/sheets/{sheet_id}/rows?withTitles=true"
    req = _urllib.Request(url, headers={"Authorization": f"Bearer {ORCA_API_KEY}"})

    for attempt in range(4):
        try:
            with _urllib.urlopen(req, timeout=60) as resp:
                data = _json.loads(resp.read())
            break
        except _urllib_err.HTTPError as e:
            if e.code == 429 and attempt < 3:
                _time.sleep((attempt + 1) * 3)
                continue
            raise RuntimeError(
                f"OrcaScan API Fehler (Sheet {sheet_id}): HTTP {e.code}: {e.reason}"
            ) from e
        except Exception as e:
            raise RuntimeError(f"OrcaScan API Fehler (Sheet {sheet_id}): {e}") from e

    all_rows = data.get("data", [])

    if drop_cols:
        for r in all_rows:
            for col in drop_cols:
                r.pop(col, None)

    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()


# ============================================================
# Kernlogik – Tagesboten Abgleich
# ============================================================
def compute_tagesboten_abgleich(abholer_source, tagesbote_df: "pd.DataFrame"):
    """
    Vergleicht Tagesboten-Liste (nur Verpackt) mit Abholer_DB.

    abholer_source  – Dateipfad (str) ODER fertiger DataFrame

    Fehlerliste  – Barcode in Abholer_DB gefunden, aber Abholbereit_At ist leer
    Errorliste   – Barcode gar nicht in Abholer_DB vorhanden

    Rückgabe:
        rows_fehler  – Tupel-Liste für Fehlerliste-Tab
        rows_error   – Tupel-Liste für Error-Tab
        fehler_df    – DataFrame für CSV-Export (Fehlerliste)
        error_df     – DataFrame für CSV-Export (Errorliste)
    """
    if isinstance(abholer_source, pd.DataFrame):
        db = abholer_source.copy()
    else:
        db = pd.read_excel(abholer_source)
    tb = tagesbote_df.copy()

    # Spaltennamen normalisieren (GROSSBUCHSTABEN → Title Case)
    tb.columns = [str(c).strip().title() for c in tb.columns]
    # Doppelte Spaltennamen entfernen (können nach .title()-Normalisierung entstehen)
    tb = tb.loc[:, ~tb.columns.duplicated(keep="first")]

    # ── Abholer_DB Spalten ──────────────────────────────────────────
    c_barcode     = first_existing(db, COL_BARCODE)
    c_abholbereit = first_existing(db, COL_ABHOLBEREIT)
    c_status      = first_existing(db, COL_STATUS)
    c_abgeholt    = first_existing(db, COL_ABGEHOLT)
    c_name        = first_existing(db, COL_NAME)
    c_zahlung_db  = first_existing(db, COL_ZAHLUNG)
    c_zielkiosk_db= first_existing(db, COL_ZIELKIOSK)
    c_scan_db     = first_existing(db, COL_SCAN_DATE)

    missing_db = [n for n, c in {
        "Paket-Barcode":  c_barcode,
        "Abholbereit_At": c_abholbereit,
    }.items() if c is None]
    if missing_db:
        raise ValueError("Abholer_DB – fehlende Spalten: " + ", ".join(missing_db))

    # ── Tagesbote Spalten ───────────────────────────────────────────
    tb_barcode = first_existing(tb, COL_TB_BARCODE)
    tb_lief    = first_existing(tb, COL_TB_LIEFERUNG)
    tb_kont    = first_existing(tb, COL_TB_KONTROLLSTATUS)
    tb_name    = first_existing(tb, COL_TB_NAME)
    tb_datum   = first_existing(tb, COL_TB_DATUM)
    tb_status  = first_existing(tb, COL_TB_STATUS)
    tb_zk      = first_existing(tb, COL_TB_ZIELKIOSK)
    tb_zahlung = first_existing(tb, COL_TB_ZAHLUNG)

    missing_tb = [n for n, c in {
        "Paket-Barcode":  tb_barcode,
        "Lieferung":      tb_lief,
        "Kontrollstatus": tb_kont,
    }.items() if c is None]
    if missing_tb:
        raise ValueError("Tagesbote – fehlende Spalten: " + ", ".join(missing_tb))

    # ── Filter: Kontrollstatus = Verpackt (kein Datumsfilter) ──────
    kont_norm = tb[tb_kont].astype(str).str.strip().str.lower()
    pickup_verpackt = tb[kont_norm.eq("verpackt")].copy()
    # Duplikate entfernen (gleicher Barcode in mehreren Tagesboten-Dateien)
    pickup_verpackt = pickup_verpackt.drop_duplicates(subset=[tb_barcode], keep="first")

    # ── Normalisierung ──────────────────────────────────────────────
    db[c_barcode]             = db[c_barcode].apply(clean_barcode)
    db[c_abholbereit]         = pd.to_datetime(db[c_abholbereit], errors="coerce")
    pickup_verpackt[tb_barcode] = pickup_verpackt[tb_barcode].apply(clean_barcode)

    # Doppelte Barcodes: Eintrag MIT Abholbereit_At bevorzugen
    db_sorted  = db.sort_values(c_abholbereit, ascending=False, na_position="last")
    db_deduped = db_sorted.drop_duplicates(subset=[c_barcode], keep="first")
    db_lookup  = db_deduped.set_index(c_barcode)

    rows_fehler = []
    rows_error  = []
    fehler_rows = []
    error_rows  = []

    def _get(r, col):
        if not col:
            return ""
        v = r.get(col, "")
        return "" if pd.isna(v) else str(v).strip()

    def _db_get(bc, col):
        """Wert aus Abholer_DB für einen Barcode holen (Fallback wenn Tagesbote leer)."""
        if not col or bc not in db_lookup.index:
            return ""
        v = db_lookup.loc[bc, col]
        if isinstance(v, pd.Series):
            v = v.iloc[0]
        return "" if pd.isna(v) else str(v).strip()

    for _, r in pickup_verpackt.iterrows():
        bc = str(r[tb_barcode]).strip()
        if not bc or bc.lower() == "nan":
            continue

        nm = _get(r, tb_name)  or _db_get(bc, c_name)
        dt = _get(r, tb_datum) or _db_get(bc, c_scan_db)
        zk = _get(r, tb_zk)    or _db_get(bc, c_zielkiosk_db)
        st = _get(r, tb_status)or _db_get(bc, c_status)
        za = _get(r, tb_zahlung)or _db_get(bc, c_zahlung_db)
        row_tuple = (bc, nm, dt, zk, st, za)

        if bc in db_lookup.index:
            try:
                abholbereit = db_lookup.loc[bc, c_abholbereit]
                # Kann Series sein wenn Barcode mehrfach in DB → erstes Element nehmen
                if isinstance(abholbereit, pd.Series):
                    abholbereit = abholbereit.iloc[0]

                # Paketstatus auslesen (falls Spalte vorhanden)
                paketstatus = ""
                if c_status and c_status in db_lookup.columns:
                    ps = db_lookup.loc[bc, c_status]
                    if isinstance(ps, pd.Series):
                        ps = ps.iloc[0]
                    paketstatus = "" if pd.isna(ps) else str(ps).strip().lower()

                # Fehlerliste: Abholbereit_At fehlt UND Paketstatus nicht ok
                ok_status = {"abholbereit", "abgeholt"}
                abholbereit_fehlt = pd.isna(abholbereit)
                status_ok = paketstatus in ok_status
                if abholbereit_fehlt and not status_ok:
                    rows_fehler.append(row_tuple)
                    row_dict = r.to_dict()
                    # Barcode + Name direkt speichern (für OrcaScan-Update)
                    row_dict["_barcode"] = bc
                    row_dict["_name"]    = nm
                    # _id aus Abholer_DB mitspeichern für OrcaScan-Update
                    if "_id" in db_lookup.columns:
                        raw = db_lookup.loc[bc, "_id"]
                        if isinstance(raw, pd.Series):
                            raw = raw.iloc[0]
                        row_dict["_db_id"] = "" if pd.isna(raw) else str(raw).strip()
                    fehler_rows.append(row_dict)
            except Exception:
                # Im Zweifel: Fehlerliste (konservativer Ansatz)
                rows_fehler.append(row_tuple)
                fehler_rows.append(r.to_dict())
        else:
            rows_error.append(row_tuple)
            error_rows.append(r.to_dict())

    cols_fallback = list(pickup_verpackt.columns)
    fehler_df = (pd.DataFrame(fehler_rows) if fehler_rows
                 else pd.DataFrame(columns=cols_fallback))
    error_df  = (pd.DataFrame(error_rows)  if error_rows
                 else pd.DataFrame(columns=cols_fallback))

    return rows_fehler, rows_error, fehler_df, error_df


# ============================================================
# Kernlogik – PU heute
# ============================================================
def compute_pickup_heute(abholer_df: "pd.DataFrame", tagesbote_df: "pd.DataFrame",
                         t2_cutoff: "str | None" = None):
    """
    Zeigt alle heutigen PUs aus dem Tagesbote-Sheet, auch wenn sie nicht in der
    Abholer_DB vorhanden sind.

    Prozess:
      1. Gandalf importiert PUs ins Tagesbote-Sheet (Status: offen)
      2. Mitarbeiter-Scan → Status wechselt zu 'verpackt' (Tagesbote)
      3. Episodisch werden PUs in Abholer_DB kopiert → Scan-Datum aktualisiert sich
      4. Kurier-Scan bei Abholung → Verpackt_At in Abholer_DB
      5. Scan am Zielstandort → Abholbereit_At in Abholer_DB

    Rückgabe: Liste von Dicts:
        barcode, name, tb_status, in_db,
        scan_datum, verpackt_at, abholbereit_at, zielkiosk
    """
    import datetime as _dt

    tb = tagesbote_df.copy()
    db = abholer_df.copy()

    # Leeres Tagesbote-Sheet → keine Einträge, kein Fehler
    if tb.empty:
        return [], {"n_raw": 0, "n_dedup_drop": 0, "dup_barcodes": [], "n_empty_bc": 0}

    # Spaltennamen normalisieren
    tb.columns = [str(c).strip().title() for c in tb.columns]

    # _id vor Normalisierung sichern (wird "_Id" nach title())
    tb_id_col = "_Id" if "_Id" in tb.columns else None

    tb_barcode = first_existing(tb, COL_TB_BARCODE)
    tb_name    = first_existing(tb, COL_TB_NAME)
    tb_datum   = first_existing(tb, COL_TB_DATUM)
    tb_zk      = first_existing(tb, COL_TB_ZIELKIOSK)
    tb_kont    = first_existing(tb, COL_TB_KONTROLLSTATUS)  # offen / verpackt

    if not tb_barcode:
        gefundene = ", ".join(str(c) for c in tb.columns[:10])
        raise ValueError(
            f"Tagesbote – Spalte 'Paket-Barcode' nicht gefunden.\n"
            f"Gefundene Spalten: {gefundene}"
        )

    # Abholer_DB Spalten
    c_barcode     = first_existing(db, COL_BARCODE)
    c_name_db     = first_existing(db, COL_NAME)
    c_scan_date   = first_existing(db, COL_SCAN_DATE)
    c_verpackt    = first_existing(db, COL_VERPACKT_AT)
    c_abholbereit = first_existing(db, COL_ABHOLBEREIT)
    c_abgeholt    = first_existing(db, COL_ABGEHOLT)
    c_status      = first_existing(db, COL_STATUS)        # Paketstatus-Feld (location)

    # Barcodes normalisieren & Duplikate entfernen
    tb[tb_barcode] = tb[tb_barcode].apply(clean_barcode)
    _n_raw = len(tb)
    _dups  = tb[tb.duplicated(subset=[tb_barcode], keep=False) & tb[tb_barcode].notna()
                & (tb[tb_barcode] != "")]
    _dup_barcodes = list(_dups[tb_barcode].unique())
    tb = tb.drop_duplicates(subset=[tb_barcode], keep="first")
    _n_after_dedup = len(tb)

    # Abholer_DB Lookup aufbauen
    if c_barcode:
        db[c_barcode] = db[c_barcode].apply(clean_barcode)
        for col in [c_scan_date, c_verpackt, c_abholbereit, c_abgeholt]:
            if col:
                db[col] = pd.to_datetime(db[col], errors="coerce")
        db_deduped = db.drop_duplicates(subset=[c_barcode], keep="first")
        db_lookup  = db_deduped.set_index(c_barcode)
    else:
        db_lookup = None

    def _fmt_ts(val):
        """Timestamp → 'DD.MM. HH:MM' (UTC+1) oder leer."""
        if val is None or (hasattr(val, '__class__') and val.__class__.__name__ == 'NaTType'):
            return ""
        try:
            if pd.isna(val):
                return ""
        except Exception:
            pass
        try:
            if hasattr(val, "tzinfo") and val.tzinfo is not None:
                val = val + _dt.timedelta(hours=2)
            return val.strftime("%d.%m. %H:%M")
        except Exception:
            return str(val)

    rows = []
    _n_empty_bc = 0
    for _, r in tb.iterrows():
        bc = str(r[tb_barcode]).strip()
        if not bc or bc.lower() == "nan":
            _n_empty_bc += 1
            continue

        def _tb(col):
            if not col:
                return ""
            v = r.get(col, "")
            return "" if pd.isna(v) else str(v).strip()

        name      = _tb(tb_name)
        zk        = _tb(tb_zk)
        tb_status = "Verpackt" if _tb(tb_kont).lower() == "verpackt" else "Offen"
        tb_row_id = str(r[tb_id_col]).strip() if tb_id_col else ""

        in_db             = False
        scan_datum        = ""
        verpackt_at       = ""
        abholbereit_at    = ""
        _abholbereit_bool = False
        db_status         = ""   # normalisierter Paketstatus aus Abholer_DB
        tour              = ""   # "T1" | "T2" | "" (wird im in_db-Block gesetzt)
        db_id             = ""   # OrcaScan _id aus Abholer_DB (für manuelle Updates)

        if db_lookup is not None and bc in db_lookup.index:
            in_db  = True
            row_db = db_lookup.loc[bc]
            if isinstance(row_db, pd.DataFrame):
                row_db = row_db.iloc[0]
            # _id für spätere API-Updates mitspeichern
            if "_id" in db_lookup.columns:
                _raw_db_id = db_lookup.loc[bc, "_id"]
                if isinstance(_raw_db_id, pd.Series):
                    _raw_db_id = _raw_db_id.iloc[0]
                db_id = "" if pd.isna(_raw_db_id) else str(_raw_db_id).strip()

            def _get_db(col):
                if not col or col not in db_lookup.columns:
                    return None
                v = row_db.get(col) if isinstance(row_db, pd.Series) else row_db[col]
                try:
                    return None if pd.isna(v) else v
                except Exception:
                    return v

            scan_datum     = _fmt_ts(_get_db(c_scan_date))
            _raw_vp        = _get_db(c_verpackt)
            verpackt_at    = _fmt_ts(_raw_vp)
            abholbereit_at = _fmt_ts(_get_db(c_abholbereit))
            # db_status: normalisierter Paketstatus für Farblogik (abgeholt, retoure, abholbereit …)
            if c_status:
                st_raw = _get_db(c_status)
                if st_raw:
                    db_status = str(st_raw).strip().lower()
            # Abgeholt_At Timestamp → db_status = "abgeholt" (auch wenn location noch nicht gesetzt)
            if db_status != "abgeholt" and c_abgeholt:
                _raw_ag = _get_db(c_abgeholt)
                if _raw_ag is not None:
                    db_status = "abgeholt"
            # _abholbereit_bool: True wenn Timestamp gesetzt ODER Paketstatus = "Abholbereit"
            _abholbereit_bool = bool(abholbereit_at) or (db_status == "abholbereit")

            tour = ""  # Tour wird ausschließlich per Button gesetzt
            # Name-Fallback: wenn Tagesbote keinen Namen hat, aus Abholer_DB nehmen
            if not name and c_name_db:
                v = _get_db(c_name_db)
                if v is not None:
                    name = str(v).strip()

        rows.append({
            "barcode":           bc,
            "name":              name,
            "tb_status":         tb_status,
            "in_db":             "✓" if in_db else "✗",
            "_in_db_bool":       in_db,
            "scan_datum":        scan_datum,
            "verpackt_at":       verpackt_at,
            "abholbereit_at":    abholbereit_at,
            "_abholbereit_bool": _abholbereit_bool,
            "db_status":         db_status,
            "zielkiosk":         zk,
            "tour":              tour,
            "_tb_row_id":        tb_row_id,
            "_db_id":            db_id,
            "_raw_vp":           _raw_vp,
        })

    diag = {
        "n_raw":        _n_raw,
        "n_dedup_drop": _n_raw - _n_after_dedup,
        "dup_barcodes": _dup_barcodes,
        "n_empty_bc":   _n_empty_bc,
    }
    return rows, diag


# ============================================================
# GUI – Tooltip
# ============================================================
class ToolTip:
    """Kleines Popup das erscheint wenn die Maus über einem Widget bleibt."""

    def __init__(self, widget, text: str = "", delay: int = 600):
        self.widget   = widget
        self.text     = text
        self.delay    = delay
        self._job     = None
        self._tip_wnd = None
        if text:
            widget.bind("<Enter>",       self._schedule, add="+")
            widget.bind("<Leave>",       self._cancel,   add="+")
            widget.bind("<ButtonPress>", self._cancel,   add="+")

    def _schedule(self, _=None):
        self._cancel()
        self._job = self.widget.after(self.delay, self._show)

    def _cancel(self, _=None):
        if self._job:
            self.widget.after_cancel(self._job)
            self._job = None
        self._hide()

    def _show(self):
        if self._tip_wnd:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.configure(bg="#fffbe6")
        tk.Label(tw, text=self.text, justify="left",
                 background="#fffbe6", relief="solid", borderwidth=1,
                 font=("Segoe UI", 9), padx=8, pady=5,
                 wraplength=340).pack()
        self._tip_wnd = tw

    def _hide(self):
        if self._tip_wnd:
            try:
                self._tip_wnd.destroy()
            except Exception:
                pass
            self._tip_wnd = None

    # ── öffentliche API für manuelle Nutzung (z.B. Notebook-Tabs) ──
    def showtip(self, text: str, x: int, y: int):
        """Zeigt Tooltip sofort an der angegebenen Bildschirmposition."""
        if self._tip_wnd:
            # Text aktualisieren falls anderer Tab
            try:
                lbl = self._tip_wnd.winfo_children()[0]
                if lbl.cget("text") == text:
                    return          # gleicher Tooltip – nichts tun
                lbl.config(text=text)
                self._tip_wnd.wm_geometry(f"+{x + 14}+{y + 18}")
                return
            except Exception:
                self._hide()
        tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x + 14}+{y + 18}")
        tw.configure(bg="#fffbe6")
        tk.Label(tw, text=text, justify="left",
                 background="#fffbe6", relief="solid", borderwidth=1,
                 font=("Segoe UI", 9), padx=8, pady=5,
                 wraplength=340).pack()
        self._tip_wnd = tw

    def hidetip(self):
        self._hide()


def add_tooltip(widget, text: str, delay: int = 600) -> ToolTip:
    """Fügt einem Widget einen Tooltip hinzu und gibt das ToolTip-Objekt zurück."""
    return ToolTip(widget, text, delay)


# ============================================================
# Kernlogik – Aufräumen (Cleanup)
# ============================================================
def compute_cleanup_candidates(df: pd.DataFrame):
    """
    Gibt zwei DataFrames zurück, die aus OrcaScan gelöscht werden sollen:

    1. df_abgeholt  – Abgeholt_At gesetzt + älter als 3 Werktage
    2. df_scan_alt  – Abholbereit_At leer + Status 'Abgeholt' + Scan-Datum älter als 7 Tage
    """
    c_abholbereit = first_existing(df, COL_ABHOLBEREIT)
    c_abgeholt    = first_existing(df, COL_ABGEHOLT)
    c_status      = first_existing(df, COL_STATUS)
    c_scan        = first_existing(df, COL_SCAN_DATE)

    today          = today_date()
    cutoff_abgeholt = add_working_days(today, -3)   # älter als 3 Werktage
    cutoff_scan     = today - timedelta(days=7)

    df = df.copy()
    if c_abgeholt:
        df[c_abgeholt] = pd.to_datetime(df[c_abgeholt], errors="coerce")
    if c_abholbereit:
        df[c_abholbereit] = pd.to_datetime(df[c_abholbereit], errors="coerce")
    if c_scan:
        df[c_scan] = pd.to_datetime(df[c_scan], errors="coerce")

    # Kriterium 1: Abgeholt_At gesetzt + älter als 3 Werktage
    if c_abgeholt:
        mask1 = df[c_abgeholt].notna() & (df[c_abgeholt].dt.date < cutoff_abgeholt)
        df_abgeholt = df[mask1].copy()
    else:
        df_abgeholt = pd.DataFrame()

    # Kriterium 2: Abholbereit_At leer + Status "Abgeholt" + Scan-Datum älter als 7 Tage
    if c_abholbereit and c_status and c_scan:
        status_norm = df[c_status].astype(str).str.strip().str.lower()
        mask2 = (
            df[c_abholbereit].isna()
            & status_norm.eq("abgeholt")
            & df[c_scan].notna()
            & (df[c_scan].dt.date < cutoff_scan)
        )
        df_scan_alt = df[mask2].copy()
    else:
        df_scan_alt = pd.DataFrame()

    return df_abgeholt, df_scan_alt


def delete_rows_orca_bulk(row_ids: list, sheet_id: str = None, workers: int = 20) -> dict:
    """
    Löscht OrcaScan-Zeilen parallel (20 gleichzeitig).
    Gibt {'deleted': N, 'failed': N, 'errors': [str, ...]} zurück.
    errors enthält bis zu 10 eindeutige Fehlermeldungen.
    """
    import urllib.request as _urllib
    from concurrent.futures import ThreadPoolExecutor, as_completed

    sid = sheet_id or ORCA_ABHOLER_SHEET_ID

    def delete_one(row_id: str):
        url = f"{ORCA_BASE_URL}/sheets/{sid}/rows/{row_id}"
        req = _urllib.Request(
            url, method="DELETE",
            headers={"Authorization": f"Bearer {ORCA_API_KEY}"}
        )
        try:
            with _urllib.urlopen(req, timeout=30):
                return True, None
        except Exception as e:
            return False, str(e)

    deleted = failed = 0
    error_set = []
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(delete_one, rid) for rid in row_ids]
        for f in as_completed(futures):
            ok, err = f.result()
            if ok:
                deleted += 1
            else:
                failed += 1
                if err and err not in error_set and len(error_set) < 10:
                    error_set.append(err)

    return {"deleted": deleted, "failed": failed, "errors": error_set}


def update_rows_orca_bulk(row_updates: list, extra_fields: dict,
                          sheet_id: str = None, workers: int = 10) -> dict:
    """
    Aktualisiert mehrere OrcaScan-Zeilen per PUT.

    row_updates  – Liste von (row_id, barcode, name) Tripeln
                   row_id  = OrcaScan _id der Zeile
                   barcode = Paket-Barcode Wert
                   name    = Empfänger-Name (OrcaScan intern: receipiantName)
    extra_fields – zusätzliche Felder (z.B. location="Abholbereit", abholbereitu005fat=ts)
    Rückgabe: {"ok": [...], "failed": [...]}
    """
    import json as _json
    import time as _time
    import urllib.error   as _urllib_err
    import urllib.request as _urllib
    from concurrent.futures import ThreadPoolExecutor as _TPE

    sid = sheet_id or ORCA_ABHOLER_SHEET_ID

    def _update_one(triple):
        row_id, bc_val, name_val = triple

        if not bc_val:
            return row_id, "Barcode leer – Zeile übersprungen"
        if not name_val:
            return row_id, "Name (receipiantName) leer – Zeile übersprungen (Fallback fehlgeschlagen)"

        # Minimaler Payload: OrcaScan-Pflichtfelder + zu ändernde Felder
        # Hinweis: OrcaScan interne Feldnamen weichen von den Anzeigenamen ab!
        # receipiantName (Tippfehler in OrcaScan!) = Anzeige "Name"
        # location                                  = Anzeige "Paketstatus"
        # abholbereitu005fat                        = Anzeige "Abholbereit_At"
        payload = {
            "barcode":          bc_val,
            "receipiantName":   name_val,   # OrcaScan-Tippfehler: receipIANT (nicht receipTANT)
        }
        payload.update(extra_fields)   # location + abholbereitu005fat

        body = _json.dumps(payload, default=str).encode("utf-8")
        url  = f"{ORCA_BASE_URL}/sheets/{sid}/rows/{row_id}?partial=true"
        req  = _urllib.Request(url, data=body, method="PUT",
                               headers={"Authorization": f"Bearer {ORCA_API_KEY}",
                                        "Content-Type": "application/json"})
        for attempt in range(3):
            try:
                with _urllib.urlopen(req, timeout=15) as resp:
                    resp.read()
                return row_id, None
            except _urllib_err.HTTPError as e:
                resp_body = ""
                try:
                    resp_body = e.read().decode("utf-8", errors="replace")[:300]
                except Exception:
                    pass
                if e.code == 429 and attempt < 2:
                    _time.sleep((attempt + 1) * 2)
                    continue
                if e.code >= 500 and attempt < 2:
                    _time.sleep((attempt + 1) * 1)
                    continue
                body_preview = body.decode("utf-8", errors="replace")[:250]
                return row_id, (f"HTTP {e.code}: {resp_body}\n"
                                f"  → barcode='{bc_val}'\n"
                                f"  → receipiantName='{name_val}'\n"
                                f"  → body: {body_preview}")
            except Exception as ex:
                if attempt < 2:
                    _time.sleep((attempt + 1) * 1)
                    continue
                return row_id, str(ex)
        return row_id, "Max retries"

    ok, failed = [], []
    with _TPE(max_workers=workers) as pool:
        for row_id, err in pool.map(_update_one, row_updates):
            if err:
                failed.append(f"{row_id}: {err}")
            else:
                ok.append(row_id)

    return {"ok": ok, "failed": failed}


def create_rows_orca_bulk(row_data_list: list,
                          sheet_id: str = None, workers: int = 10) -> dict:
    """
    Erstellt mehrere neue OrcaScan-Zeilen per POST.

    row_data_list – Liste von dicts mit den Feldern (interne OrcaScan-Namen!)
                    Pflicht: "barcode" + "receipiantName"
    Rückgabe: {"ok": [barcodes], "failed": ["barcode: Fehler"]}
    """
    import json as _json
    import time as _time
    import urllib.error   as _urllib_err
    import urllib.request as _urllib
    from concurrent.futures import ThreadPoolExecutor as _TPE

    sid = sheet_id or ORCA_ABHOLER_SHEET_ID

    def _create_one(row_dict):
        bc_val = row_dict.get("barcode", "?")
        # Leere Felder weglassen (OrcaScan mag keine leeren Strings für manche Felder)
        payload = {k: v for k, v in row_dict.items() if v not in ("", None)}
        body = _json.dumps(payload, default=str).encode("utf-8")
        url  = f"{ORCA_BASE_URL}/sheets/{sid}/rows"
        req  = _urllib.Request(url, data=body, method="POST",
                               headers={"Authorization": f"Bearer {ORCA_API_KEY}",
                                        "Content-Type": "application/json"})
        for attempt in range(3):
            try:
                with _urllib.urlopen(req, timeout=15) as resp:
                    resp.read()
                return bc_val, None
            except _urllib_err.HTTPError as e:
                resp_body = ""
                try:
                    resp_body = e.read().decode("utf-8", errors="replace")[:200]
                except Exception:
                    pass
                if e.code == 429 and attempt < 2:
                    _time.sleep((attempt + 1) * 2)
                    continue
                return bc_val, f"HTTP {e.code}: {resp_body}"
            except Exception as ex:
                return bc_val, str(ex)
        return bc_val, "Max retries"

    ok, failed = [], []
    with _TPE(max_workers=workers) as pool:
        for bc, err in pool.map(_create_one, row_data_list):
            if err:
                failed.append(f"{bc}: {err}")
            else:
                ok.append(bc)

    return {"ok": ok, "failed": failed}


def run_cleanup(on_status=None, on_done=None, on_preview=None, dry_run=False):
    """
    Vollständiger Cleanup-Ablauf (läuft im Hintergrund-Thread).

    dry_run=True  → kein Löschen aus OrcaScan; on_preview(df) wird mit den Treffern aufgerufen.

    on_status(text)   – Statuszeilen-Update
    on_done(text)     – Abschluss oder Fehler
    on_preview(df)    – nur bei dry_run: DataFrame mit Treffern für Anzeige
    """
    def _notify(fn, *args):
        if fn:
            fn(*args)

    try:
        _notify(on_status, "Cleanup: Lade Abholer_DB aus OrcaScan …")
        df = fetch_abholer_orca()

        _notify(on_status, "Cleanup: Filtere Kandidaten …")
        df1, df2 = compute_cleanup_candidates(df)
        combined = pd.concat([df1, df2], ignore_index=True).drop_duplicates(subset=["_id"])

        if combined.empty:
            _notify(on_done, "Cleanup: Keine Einträge zum Löschen gefunden.")
            return

        # Excel-Archiv nach Google Drive
        ts       = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f"Abholer_DB_Archiv_{ts}.xlsx"
        _notify(on_status, f"Cleanup: Exportiere {len(combined)} Zeilen nach Google Drive …")
        upload_excel_to_gdrive(combined, GDRIVE_ABHOLER_FOLDER_ID, filename)

        if dry_run:
            _notify(on_preview, combined)
            _notify(on_done,
                f"DRY RUN – {len(combined)} Treffer  |  Archiv gespeichert: {filename}\n"
                f"  Kriterium 1 (Abgeholt >3 Werktage): {len(df1)}\n"
                f"  Kriterium 2 (Scan-Datum >7 Tage):   {len(df2)}\n"
                f"Nichts wurde gelöscht."
            )
            return

        # Echtes Löschen
        row_ids = combined["_id"].dropna().astype(str).tolist()
        _notify(on_status, f"Cleanup: Lösche {len(row_ids)} Zeilen aus OrcaScan …")
        result  = delete_rows_orca_bulk(row_ids)

        status_text = (
            f"Cleanup abgeschlossen: {result['deleted']} gelöscht, "
            f"{result['failed']} fehlgeschlagen  |  Archiv: {filename}"
        )
        _notify(on_done, status_text, result.get("errors", []))

    except Exception as e:
        _notify(on_done, f"Cleanup Fehler: {e}")


# ============================================================
# GUI – Generischer Tabellen-Tab
# ============================================================
class FlowFrame(tk.Frame):
    """
    Frame das seine Kind-Widgets automatisch in die nächste Zeile umbricht,
    wenn die verfügbare Breite nicht ausreicht (ähnlich wie CSS flex-wrap).
    """

    def __init__(self, master, bg, hgap=12, vgap=4, **kw):
        super().__init__(master, bg=bg, **kw)
        self._hgap  = hgap
        self._vgap  = vgap
        self._items: list = []
        self.bind("<Configure>", lambda e: self.after_idle(self._relayout))

    def add_item(self, widget):
        self._items.append(widget)
        self.after_idle(self._relayout)
        return widget

    def _relayout(self):
        w_avail = self.winfo_width()
        if w_avail < 10 or not self._items:
            return
        for w in self._items:
            w.update_idletasks()

        x, y, row_h = self._hgap, self._vgap, 0
        for w in self._items:
            ww = w.winfo_reqwidth()
            wh = w.winfo_reqheight()
            if x + ww + self._hgap > w_avail and x > self._hgap:
                x      = self._hgap
                y     += row_h + self._vgap
                row_h  = 0
            w.place(x=x, y=y)
            x     += ww + self._hgap
            row_h  = max(row_h, wh)

        self.config(height=max(y + row_h + self._vgap, 1))


class TableTab:
    """
    Einheitlicher Tabellen-Tab mit Suche und Excel-artiger Zellen-Auswahl (tksheet).

    columns      – Liste von (col_id, header_label, pixel_width)
    today_header – wenn True zeigt der Zähler das heutige Datum (DHL-Style)
    """

    def __init__(self, parent, title: str, columns: list[tuple[str, str, int]],
                 today_header: bool = False, row_color_fn=None,
                 editable_col_map=None, orca_sheet_id=None, orca_id_idx=None,
                 legend_items=None):
        try:
            from tksheet import Sheet as _Sheet
        except ImportError:
            _Sheet = None

        self.columns             = columns
        self.today_header        = today_header
        self._Sheet              = _Sheet
        self._editable_col_map   = editable_col_map or {}
        self._orca_sheet_id      = orca_sheet_id
        self._orca_id_idx        = orca_id_idx

        self.frame = tk.Frame(parent)

        if title:
            tk.Label(self.frame, text=title, anchor="w",
                     font=("Segoe UI", 9, "bold")).pack(fill="x", pady=(4, 2))

        # -- Suchzeile
        sr = tk.Frame(self.frame)
        sr.pack(fill="x", pady=(2, 6))
        tk.Label(sr, text="Suche:", width=8, anchor="w").pack(side="left")
        self.search = tk.StringVar()
        tk.Entry(sr, textvariable=self.search).pack(side="left", fill="x", expand=True)
        self.count_lbl = tk.Label(sr, text=self._count_text(0),
                                  anchor="e", width=(28 if today_header else 16))
        self.count_lbl.pack(side="left", padx=(8, 0))

        # -- Tabelle (tksheet) + optionale Seitenlegende
        tr = tk.Frame(self.frame)
        tr.pack(fill="both", expand=True)

        if legend_items:
            _legend = tk.Frame(tr, bg="#f5f5f5", bd=1, relief="groove", width=185)
            _legend.pack(side="right", fill="y", padx=(4, 0), pady=(0, 4))
            _legend.pack_propagate(False)
            tk.Label(_legend, text="Legende", font=("Segoe UI", 9, "bold"),
                     bg="#f5f5f5", anchor="w").pack(fill="x", padx=8, pady=(8, 4))
            for _color, _text in legend_items:
                _row_l = tk.Frame(_legend, bg="#f5f5f5")
                _row_l.pack(fill="x", padx=8, pady=2)
                tk.Frame(_row_l, bg=_color, width=16, height=16,
                         relief="solid", bd=1).pack(side="left")
                tk.Label(_row_l, text=f"  {_text}", font=("Segoe UI", 8),
                         bg="#f5f5f5", anchor="w", wraplength=145,
                         justify="left").pack(side="left", fill="x")
            _sheet_host = tk.Frame(tr)
            _sheet_host.pack(side="left", fill="both", expand=True)
        else:
            _sheet_host = tr

        headers = [c[1] for c in columns]

        if _Sheet:
            try:
                self.sheet = _Sheet(
                    _sheet_host,
                    headers=headers,
                    show_row_index=False,
                    header_font=("Segoe UI", 10, "bold"),
                    font=("Segoe UI", 9),
                    outline_color="#1a3a5c",
                )
            except Exception:
                # Minimalversion falls Parameter nicht passen
                self.sheet = _Sheet(_sheet_host, headers=headers)
            self.sheet.pack(fill="both", expand=True)
            try:
                self.sheet.set_column_widths([c[2] for c in columns])
            except Exception:
                pass
            try:
                self.sheet.enable_bindings()
            except Exception:
                pass
            try:
                self.sheet.extra_bindings([
                    ("cell_select",   self._on_cell_click),
                    ("end_edit_cell", self._on_cell_edited),
                ])
            except Exception:
                pass
        else:
            # Fallback: einfaches Label wenn tksheet nicht installiert
            tk.Label(_sheet_host, text="tksheet nicht installiert.\nBitte 'pip install tksheet' ausführen.",
                     fg="red", font=("Segoe UI", 10)).pack(expand=True)
            self.sheet = None

        # -- Buttons
        br = tk.Frame(self.frame)
        br.pack(fill="x", pady=(6, 0))
        b_copy = tk.Button(br, text="Gefilterte Liste kopieren", command=self.copy_filtered)
        b_copy.pack(side="left")
        add_tooltip(b_copy, "Kopiert alle aktuell angezeigten Zeilen in die Zwischenablage.")
        b_clear = tk.Button(br, text="Suche leeren", command=lambda: self.search.set(""))
        b_clear.pack(side="left", padx=6)
        add_tooltip(b_clear, "Suchfeld leeren und alle Einträge wieder anzeigen.")
        self.btn_frame = br  # öffentlich für externe Buttons

        self.rows:              list[tuple] = []
        self.filtered:          list[tuple] = []
        self._total:            int         = 0
        self._on_rows_changed               = None  # optionaler Callback nach Zeilen-Änderung
        self._row_color_fn                  = row_color_fn
        self.search.trace_add("write", lambda *_: self.refresh())

        # Mausrad-Scroll: vom ganzen Frame an tksheet weiterleiten
        if self.sheet is not None:
            def _mw(e):
                self.sheet.yview_scroll(int(-1 * (e.delta / 120)), "units")
            def _bind_mw(w):
                w.bind("<MouseWheel>", _mw, add="+")
                for ch in w.winfo_children():
                    _bind_mw(ch)
            self.frame.bind("<Map>", lambda e: _bind_mw(self.frame), add="+")

    # ------------------------------------------------------------------ helpers

    def _count_text(self, n: int, total: int | None = None) -> str:
        t = total if total is not None else n
        if self.today_header:
            d = today_date().strftime("%d.%m.%Y")
            return f"Heute ({d}): {t}"
        return f"{n} / {t} Einträge" if t != n else f"{n} Einträge"
    def _on_header_click(self, event):
        """Spaltenheader klicken: 1x aufsteigend, 2x absteigend, 3x Standard."""
        try:
            col = event.selected.column
        except Exception:
            return
        if col is None:
            return
        if col is None:
            return
        _sortable = {0, 1, 2, 3, 5, 6, 7}
        if col not in _sortable:
            return
        if self._sort_col == col:
            self._sort_dir = (self._sort_dir + 1) % 3
            if self._sort_dir == 0:
                self._sort_col = None
        else:
            self._sort_col = col
            self._sort_dir = 1
        self._refresh_ui()

    def _on_cell_click(self, event):
        """Barcode der angeklickten Zeile in die Zwischenablage kopieren."""
        try:
            row = event.row
            if 0 <= row < len(self.filtered) and self.filtered[row]:
                barcode = str(self.filtered[row][0])
                if barcode:
                    self.frame.clipboard_clear()
                    self.frame.clipboard_append(barcode)
                    # Kurze Bestätigung im count_lbl
                    orig = self.count_lbl.cget("text")
                    self.count_lbl.config(text=f"📋 Kopiert: {barcode[:28]}", fg="#27ae60")
                    self.frame.after(1800, lambda: self.count_lbl.config(
                        text=orig, fg="black"))
        except Exception:
            pass

    def _on_cell_edited(self, event):
        """Zell-Bearbeitung → OrcaScan aktualisieren (nur für konfigurierte Spalten)."""
        import threading as _thr
        try:
            col_idx = event.column
            row_idx = event.row
            if (col_idx not in self._editable_col_map
                    or not self._orca_sheet_id
                    or self._orca_id_idx is None):
                return
            if row_idx >= len(self.filtered):
                return
            raw_row = self.filtered[row_idx]
            row_id  = raw_row[self._orca_id_idx]
            barcode = raw_row[0]
            name    = raw_row[2]
            if not row_id or not barcode:
                return
            try:
                new_val = str(event.value).strip()
            except Exception:
                new_val = str(self.sheet.get_cell_data(row_idx, col_idx) or "").strip()
            orca_field = self._editable_col_map[col_idx]
            # Lokal sofort aktualisieren
            lst = list(raw_row)
            lst[col_idx] = new_val
            new_row = tuple(lst)
            self.filtered[row_idx] = new_row
            try:
                idx = self.rows.index(raw_row)
                self.rows[idx] = new_row
            except ValueError:
                pass
            # OrcaScan im Hintergrund aktualisieren
            def _send():
                update_rows_orca_bulk(
                    [(row_id, barcode, name)],
                    {orca_field: new_val},
                    sheet_id=self._orca_sheet_id
                )
            _thr.Thread(target=_send, daemon=True).start()
            # Kurze Bestätigung im count_lbl
            orig = self.count_lbl.cget("text")
            self.count_lbl.config(
                text=f"✓ Gespeichert: {new_val[:25]}", fg="#27ae60")
            self.frame.after(2500, lambda: self.count_lbl.config(text=orig, fg="black"))
        except Exception:
            pass

    # ------------------------------------------------------------------ public

    def set_rows(self, rows):
        self.rows   = list(rows or [])
        self._total = len(self.rows)
        self.refresh()

    def refresh(self):
        q = (self.search.get() or "").strip().lower()
        self.filtered = (
            self.rows[:]
            if not q
            else [r for r in self.rows if any(q in str(v).lower() for v in r)]
        )
        self._redraw()

    def _redraw(self):
        if self.sheet is None:
            return
        data = [[str(v) for v in r] for r in self.filtered]
        try:
            self.sheet.set_sheet_data(data, reset_col_positions=False)
        except Exception:
            self.sheet.set_sheet_data(data)
        if self._row_color_fn is not None:
            try:
                self.sheet.dehighlight_all()
                for i, row in enumerate(self.filtered):
                    color = self._row_color_fn(row)
                    if color:
                        self.sheet.highlight_rows(rows=[i], bg=color, redraw=False)
                self.sheet.refresh()
            except Exception:
                pass
        self.count_lbl.config(text=self._count_text(len(self.filtered), self._total))

    def delete_selected_rows(self):
        """Löscht alle aktuell markierten Zeilen aus der Liste (nur Anzeige, nicht OrcaScan)."""
        if self.sheet is None:
            return
        selected_indices = set()
        try:
            # Markierte Zeilen aus tksheet holen
            for r in (self.sheet.get_selected_rows() or []):
                selected_indices.add(r)
            # Auch einzelne Zellen berücksichtigen
            for r, _c in (self.sheet.get_selected_cells() or []):
                selected_indices.add(r)
        except Exception:
            pass
        if not selected_indices:
            return
        # Gefilterte Indizes → originale Tupel
        to_delete = {self.filtered[i] for i in selected_indices if i < len(self.filtered)}
        self.rows  = [row for row in self.rows if row not in to_delete]
        self._total = len(self.rows)
        self.refresh()
        if self._on_rows_changed:
            self._on_rows_changed()

    def get_selected_rows(self) -> list:
        """Gibt die aktuell markierten Zeilen als Liste von Tupeln zurück."""
        if self.sheet is None:
            return []
        selected_indices = set()
        try:
            for r in (self.sheet.get_selected_rows() or []):
                selected_indices.add(r)
            for r, _c in (self.sheet.get_selected_cells() or []):
                selected_indices.add(r)
        except Exception:
            pass
        return [self.filtered[i] for i in sorted(selected_indices) if i < len(self.filtered)]

    def copy_filtered(self):
        if not self.filtered:
            return
        text = "\n".join("\t".join(str(v) for v in r).rstrip() for r in self.filtered)
        root = self.frame.winfo_toplevel()
        root.clipboard_clear()
        root.clipboard_append(text)


# ============================================================
# GUI – DHL Merge Tab
# ============================================================
class DHLMergeTab:
    def __init__(self, parent, get_export_folder=None):
        self.frame = tk.Frame(parent)
        tk.Label(self.frame,
                 text="DHL (heute) – DHL_Normal (nur heute)",
                 anchor="w", font=("Segoe UI", 9, "bold")).pack(fill="x", pady=(4, 6))

        self.source_folder     = Path.home() / "Downloads"
        self.get_export_folder = get_export_folder or (lambda: Path.home() / "Downloads")
        self._orca_df: "pd.DataFrame | None" = None   # gespeicherter OrcaScan DataFrame

        row = tk.Frame(self.frame)
        row.pack(fill="x", pady=(0, 6))
        self.orca_export_btn = tk.Button(row, text="OrcaScan exportieren",
                  command=self.export_orca, state="disabled")
        self.orca_export_btn.pack(side="left", padx=(0, 6))
        add_tooltip(self.orca_export_btn, "Exportiert die via OrcaScan geladenen DHL-Daten\nals Excel-Datei in den Exportordner.")
        b_clear_dhl = tk.Button(row, text="Tab leeren", command=self.clear)
        b_clear_dhl.pack(side="left")
        add_tooltip(b_clear_dhl, "Tabelle leeren und Status zurücksetzen.")

        self.status_lbl = tk.Label(self.frame,
                                   text="Bereit. Erwartet DHL_Normal-Datei im Quellordner.",
                                   anchor="w")
        self.status_lbl.pack(fill="x", pady=(0, 4))

        self.path_lbl = tk.Label(self.frame, text="Exportpfad: —", anchor="w")
        self.path_lbl.pack(fill="x", pady=(0, 8))

        self.preview = TableTab(
            self.frame,
            title="Vorschau – heutige Zeilen (Merge)",
            columns=[("barcode", "Package Barcode", 420), ("dt", "Date of Scan", 260)],
            today_header=True,
        )
        self.preview.frame.pack(fill="both", expand=True)  # Bug-Fix: frame war vorher nicht gepackt
        self.preview.set_rows([])

    def set_rows_direct(self, rows: list, merged_df: "pd.DataFrame | None" = None):
        """Zeilen direkt setzen (z.B. aus OrcaScan API). DataFrame optional für Export."""
        self._orca_df = merged_df
        self.preview.set_rows(rows)
        self.status_lbl.config(text=f"OrcaScan geladen  –  {len(rows)} Zeilen heute")
        self.path_lbl.config(text="Exportpfad: —")
        self.orca_export_btn.config(state="normal" if merged_df is not None and not merged_df.empty else "disabled")

    def export_orca(self):
        """Exportiert den zuletzt via OrcaScan geladenen Merge-DataFrame."""
        if self._orca_df is None or self._orca_df.empty:
            messagebox.showinfo("Export", "Keine OrcaScan-Daten zum Exportieren vorhanden.")
            return
        try:
            out_path = self._save_to_downloads_xlsx_overwrite(self._orca_df)
            self.path_lbl.config(text=f"Exportpfad: {out_path}")
            self.status_lbl.config(
                text=f"✅  Exportiert: {len(self._orca_df)} Zeilen  →  {out_path.name}"
            )
        except Exception as e:
            messagebox.showerror("Export Fehler", str(e))

    def clear(self):
        self.status_lbl.config(text="Bereit. Erwartet DHL_Normal-Datei im Quellordner.")
        self.path_lbl.config(text="Exportpfad: —")
        self.preview.set_rows([])

    def choose_source_folder(self):
        p = filedialog.askdirectory()
        if p:
            self.source_folder = Path(p)
            self.src_lbl.config(text=f"Quellordner: {self.source_folder}")

    def _save_to_downloads_xlsx_overwrite(self, merged_df: pd.DataFrame) -> Path:
        out_path = self.get_export_folder() / (today_date().strftime("%y%m%d") + ".xlsx")
        # Zeitzone aus Datumsspalten entfernen (Excel unterstützt keine tz-aware Datetimes)
        df = merged_df.copy()
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)
        write_excel_text_cols(df, out_path, text_cols=["Package Barcode"])
        return out_path

    def auto_import_merge(self):
        try:
            src = self.source_folder
            f_n = find_latest_dhl_normal(src)

            if not f_n:
                raise ValueError(
                    "Auto-Import konnte DHL_Normal nicht finden.\n\n"
                    f"Quellordner:\n  {src}\n\n"
                    "Beispiel:\n"
                    "  DHL_Normal.xlsx, DHL-Normal - Kopie.xlsx"
                )

            merged_df, rows = compute_orca_merge_today(str(f_n))
            out_path = self._save_to_downloads_xlsx_overwrite(merged_df)

            self.status_lbl.config(
                text=f"OK | DHL_Normal: {f_n.name} | Heute: {len(rows)} Zeilen"
            )
            self.path_lbl.config(text=f"Exportpfad: {out_path}")
            self.preview.set_rows(rows)

        except Exception as e:
            messagebox.showerror("DHL (heute) – Auto-Import Fehler", str(e))

    def manual_select_merge(self):
        files = filedialog.askopenfilenames(
            title="Bitte DHL_Normal Datei wählen (Excel oder CSV)",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("Alle Dateien", "*.*")]
        )
        if not files:
            return
        if len(files) != 1:
            messagebox.showerror("DHL (heute)", "Bitte genau 1 Datei auswählen.")
            return

        f1 = files[0]
        try:
            merged_df, rows = compute_orca_merge_today(f1)
            out_path = self._save_to_downloads_xlsx_overwrite(merged_df)

            self.status_lbl.config(text=f"Manuell OK | Heute: {len(rows)} Zeilen")
            self.path_lbl.config(text=f"Exportpfad: {out_path}")
            self.preview.set_rows(rows)

        except Exception as e:
            messagebox.showerror("DHL (heute) – Fehler", str(e))


# ============================================================
# Hilfsfunktion – Archiv-Dateien aus Google Drive laden
# ============================================================
def fetch_archiv_gdrive() -> "pd.DataFrame":
    """
    Lädt alle Abholer_DB_Archiv_*.xlsx Dateien aus dem Archiv-Ordner (Google Drive).
    Benutzt OAuth2 (gleicher Mechanismus wie der Archiv-Upload beim Cleanup).
    Gibt einen kombinierten DataFrame zurück (leerer DF wenn keine Archive vorhanden).
    """
    if not GDRIVE_AVAILABLE:
        raise RuntimeError("Google API nicht verfügbar (pip install google-api-python-client google-auth).")

    service = _get_oauth_drive_service()

    results = service.files().list(
        q=(
            f"'{GDRIVE_ABHOLER_FOLDER_ID}' in parents "
            f"and name contains 'Abholer_DB_Archiv' "
            f"and trashed = false"
        ),
        fields="files(id, name, mimeType)",
        orderBy="modifiedTime desc",
        pageSize=500,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()

    files = results.get("files", [])
    if not files:
        return pd.DataFrame()

    # ── Nur Dateien der letzten 12 Monate laden (Datum aus Dateiname) ──────
    import re as _re2
    from datetime import date as _date
    _cutoff = _date.today().replace(year=_date.today().year - 1)
    filtered_files = []
    for f in files:
        m = _re2.search(r'(\d{4}-\d{2}-\d{2})', f.get("name", ""))
        if m:
            try:
                file_date = _date.fromisoformat(m.group(1))
                if file_date >= _cutoff:
                    filtered_files.append(f)
            except ValueError:
                pass   # Dateiname mit ungültigem Datum überspringen
        # Dateien ohne erkennbares Datum werden übersprungen
    files = filtered_files

    if not files:
        return pd.DataFrame()

    import io as _io2
    all_frames = []
    for file in files:
        try:
            request = service.files().get_media(fileId=file["id"])
            buf = _io2.BytesIO()
            downloader = MediaIoBaseDownload(buf, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            buf.seek(0)
            df = pd.read_excel(buf)
            if not df.empty:
                all_frames.append(df)
        except Exception:
            pass  # Fehlerhafte Datei überspringen

    if not all_frames:
        return pd.DataFrame()

    combined = pd.concat(all_frames, ignore_index=True)
    # Duplikate entfernen (gleicher Barcode kann in mehreren Archiv-Dateien vorkommen)
    c_bc = first_existing(combined, COL_BARCODE)
    if c_bc:
        combined = combined.drop_duplicates(subset=[c_bc], keep="last")
    return combined


# ============================================================
# GUI – Statistik Tab (PU + DHL kombiniert)
# ============================================================
class StatistikTab:
    """
    Kombinierte Statistikansicht:
      Sektion 1 – PU: Lieferungen (Verpackt_At) + Kundenabholungen (Abgeholt_At)
      Sektion 2 – DHL: DHL Normal + DHL Express
    """

    # Farben PU
    _COL_ANLIEF = "#16a085"   # teal  – Lieferungen
    _COL_ABHOL  = "#1a237e"   # indigo – Kundenabholungen
    # Farben DHL
    _COL_NORMAL   = "#f9a825"   # gelb  – DHL Normal
    _COL_EXPRESS  = "#b71c1c"   # rot   – DHL Express
    _COL_ABHOLUNG = "#2e7d32"   # grün  – Kundenabholungen
    _COL_BG       = "#f0f2f5"

    def __init__(self, parent, start_loading=None, stop_loading=None):
        self._start_loading_cb = start_loading or (lambda msg="Lade …": None)
        self._stop_loading_cb  = stop_loading  or (lambda msg="Bereit.": None)

        # ── PU Daten ─────────────────────────────────────────────────
        self._main_df:    "pd.DataFrame | None" = None
        self._archiv_df:  "pd.DataFrame | None" = None
        self._pu_weekly_data:  list = []
        self._pu_monthly_data: list = []
        self._pu_daily_data:   list = []
        self._pu_view_mode = tk.StringVar(value="weekly")
        self._pu_loading   = False

        # ── DHL Daten ─────────────────────────────────────────────────
        self._normal_df:    "pd.DataFrame | None" = None
        self._express_df:   "pd.DataFrame | None" = None
        self._dhl_weekly_data:  list = []
        self._dhl_monthly_data: list = []
        self._dhl_daily_data:   list = []
        self._dhl_view_mode = tk.StringVar(value="weekly")
        self._dhl_loading   = False

        # ── Scrollbarer Container (geteilt) ──────────────────────────
        self.frame = tk.Frame(parent, bg=self._COL_BG)
        _scrollbar = tk.Scrollbar(self.frame, orient="vertical")
        _scrollbar.pack(side="right", fill="y")
        _scroll_canvas = tk.Canvas(self.frame, bg=self._COL_BG,
                                   highlightthickness=0,
                                   yscrollcommand=_scrollbar.set)
        _scroll_canvas.pack(side="left", fill="both", expand=True)
        _scrollbar.config(command=_scroll_canvas.yview)
        self._inner = tk.Frame(_scroll_canvas, bg=self._COL_BG)
        _scroll_win = _scroll_canvas.create_window((0, 0), window=self._inner, anchor="nw")

        def _on_inner_configure(e):
            _scroll_canvas.configure(scrollregion=_scroll_canvas.bbox("all"))
        def _on_canvas_configure(e):
            _scroll_canvas.itemconfig(_scroll_win, width=e.width)
        def _on_mousewheel(e):
            _scroll_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        def _bind_mousewheel(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            for child in widget.winfo_children():
                _bind_mousewheel(child)

        self._inner.bind("<Configure>", _on_inner_configure)
        _scroll_canvas.bind("<Configure>", _on_canvas_configure)
        _scroll_canvas.bind("<MouseWheel>", _on_mousewheel)
        self._inner.bind("<Map>", lambda e: _bind_mousewheel(self._inner), add="+")

        # ══════════════════════════════════════════════════════════════
        # SEKTION 1 – PU Statistik
        # ══════════════════════════════════════════════════════════════

        hdr_pu = tk.Frame(self._inner, bg=self._COL_BG)
        hdr_pu.pack(fill="x", padx=16, pady=(10, 4))
        tk.Label(hdr_pu, text="📈  PU Statistik",
                 font=("Segoe UI", 12, "bold"),
                 bg=self._COL_BG, fg="#2c3e50").pack(side="left")
        self._pu_status_lbl = tk.Label(hdr_pu, text="⏳  Lade Archiv …",
                                       font=("Segoe UI", 9), bg=self._COL_BG, fg="#888")
        self._pu_status_lbl.pack(side="right", padx=6)

        leg_pu = tk.Frame(self._inner, bg=self._COL_BG)
        leg_pu.pack(fill="x", padx=16, pady=(0, 6))
        for color, text in [(self._COL_ANLIEF, "🚐 Lieferungen  (Verpackt_At)"),
                            (self._COL_ABHOL,  "📦 Kundenabholungen  (Abgeholt_At)")]:
            tk.Frame(leg_pu, bg=color, width=14, height=14, relief="flat").pack(side="left")
            tk.Label(leg_pu, text=f" {text}   ", font=("Segoe UI", 9),
                     bg=self._COL_BG, fg="#2c3e50").pack(side="left")

        cards_pu = tk.Frame(self._inner, bg=self._COL_BG)
        cards_pu.pack(fill="x", padx=16, pady=(4, 10))

        def _card_pu(parent, col, icon, label1, label2):
            outer = tk.Frame(parent, bg=col, padx=2, pady=2)
            outer.pack(side="left", padx=6, expand=True, fill="x")
            inner = tk.Frame(outer, bg=col, padx=14, pady=10)
            inner.pack(fill="both")
            cnt = tk.Label(inner, text="–", font=("Segoe UI", 28, "bold"),
                           bg=col, fg="white", anchor="center")
            cnt.pack(fill="x")
            tk.Label(inner, text=f"{icon}  {label1}",
                     font=("Segoe UI", 9, "bold"), bg=col, fg="white").pack()
            tk.Label(inner, text=label2,
                     font=("Segoe UI", 8), bg=col, fg="#d0f0e8").pack()
            return cnt

        self._lbl_anlief_woche = _card_pu(cards_pu, self._COL_ANLIEF, "🚐", "Lieferungen", "Diese Woche")
        self._lbl_anlief_monat = _card_pu(cards_pu, "#117a65",        "🚐", "Lieferungen", "Dieser Monat")
        self._lbl_abhol_woche  = _card_pu(cards_pu, "#1a237e",        "📦", "Abholungen",  "Diese Woche")
        self._lbl_abhol_monat  = _card_pu(cards_pu, "#283593",        "📦", "Abholungen",  "Dieser Monat")

        tk.Frame(self._inner, bg="#dde2e8", height=1).pack(fill="x", padx=16, pady=(0, 8))

        # ── Chart + Kiosk nebeneinander ────────────────────────────────
        pu_side = tk.Frame(self._inner, bg=self._COL_BG)
        pu_side.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        # Linke Spalte: Chart
        chart_col = tk.Frame(pu_side, bg=self._COL_BG)
        chart_col.pack(side="left", fill="both", expand=True)

        chart_hdr_pu = tk.Frame(chart_col, bg=self._COL_BG)
        chart_hdr_pu.pack(fill="x", pady=(0, 4))
        self._pu_chart_title_lbl = tk.Label(chart_hdr_pu, text="Letzte 4 Kalenderwochen",
                                            font=("Segoe UI", 10, "bold"),
                                            bg=self._COL_BG, fg="#2c3e50")
        self._pu_chart_title_lbl.pack(side="left")

        pu_toggle_frame = tk.Frame(chart_hdr_pu, bg=self._COL_BG)
        pu_toggle_frame.pack(side="right")
        self._pu_toggle_btns = {}

        def _pu_switch_view(mode):
            self._pu_view_mode.set(mode)
            titles = {"weekly": "Letzte 4 Kalenderwochen",
                      "daily":  "Letzte 4 Wochen – Tagesansicht",
                      "monthly":"Letzte 6 Monate"}
            self._pu_chart_title_lbl.config(text=titles.get(mode, ""))
            for m, b in self._pu_toggle_btns.items():
                b.config(bg="#2c3e50" if m == mode else "#dde2e8",
                         fg="white"   if m == mode else "#2c3e50",
                         relief="sunken" if m == mode else "flat")
            self._pu_redraw_chart()

        for _mode, _label, _active in [("daily", "Tagesansicht", False),
                                        ("weekly", "Wochenübersicht", True),
                                        ("monthly", "Monatsübersicht", False)]:
            _b = tk.Button(pu_toggle_frame, text=_label,
                           font=("Segoe UI", 8), padx=8, pady=2, bd=1,
                           bg="#2c3e50" if _active else "#dde2e8",
                           fg="white"   if _active else "#2c3e50",
                           relief="sunken" if _active else "flat",
                           command=lambda m=_mode: _pu_switch_view(m))
            _b.pack(side="left", padx=(0, 2))
            self._pu_toggle_btns[_mode] = _b

        self._pu_canvas = tk.Canvas(chart_col, bg=self._COL_BG, height=260, highlightthickness=0)
        self._pu_canvas.pack(fill="both", expand=True, pady=(4, 0))
        self._pu_canvas.bind("<Configure>", lambda e: self._pu_redraw_chart())

        # Vertikaler Trenner
        tk.Frame(pu_side, bg="#dde2e8", width=1).pack(side="left", fill="y", padx=(12, 12))

        # Rechte Spalte: Kiosk (feste Breite)
        kiosk_col = tk.Frame(pu_side, bg=self._COL_BG, width=360)
        kiosk_col.pack(side="left", fill="y")
        kiosk_col.pack_propagate(False)

        kiosk_hdr = tk.Frame(kiosk_col, bg=self._COL_BG)
        kiosk_hdr.pack(fill="x", pady=(0, 4))
        tk.Label(kiosk_hdr, text="Zielkiosk",
                 font=("Segoe UI", 10, "bold"),
                 bg=self._COL_BG, fg="#2c3e50").pack(side="left")
        self._kiosk_range = tk.StringVar(value="week")
        kiosk_toggle = tk.Frame(kiosk_hdr, bg=self._COL_BG)
        kiosk_toggle.pack(side="right")
        self._kiosk_btns = {}
        for mode, label in [("week", "Woche"), ("month", "Monat"), ("months6", "6 Monate")]:
            b = tk.Button(kiosk_toggle, text=label,
                          font=("Segoe UI", 8), padx=6, pady=2, bd=1,
                          bg="#2c3e50" if mode == "week" else "#dde2e8",
                          fg="white"   if mode == "week" else "#2c3e50",
                          relief="sunken" if mode == "week" else "flat",
                          command=lambda m=mode: self._pu_switch_kiosk_range(m))
            b.pack(side="left", padx=(0, 2))
            self._kiosk_btns[mode] = b
        self._kiosk_canvas = tk.Canvas(kiosk_col, bg=self._COL_BG, highlightthickness=0)
        self._kiosk_canvas.pack(fill="both", expand=True)
        self._kiosk_canvas.bind("<Configure>", lambda e: self._pu_redraw_kiosk_table())
        self._kiosk_data = {"week": [], "month": [], "months6": []}

        # ══════════════════════════════════════════════════════════════
        # Horizontaler Trenner
        # ══════════════════════════════════════════════════════════════
        tk.Frame(self._inner, bg="#aab4c4", height=3).pack(fill="x", padx=0, pady=(8, 0))

        # ══════════════════════════════════════════════════════════════
        # SEKTION 2 – DHL Statistik
        # ══════════════════════════════════════════════════════════════

        hdr_dhl = tk.Frame(self._inner, bg=self._COL_BG)
        hdr_dhl.pack(fill="x", padx=16, pady=(10, 4))
        tk.Label(hdr_dhl, text="📊  DHL Statistik",
                 font=("Segoe UI", 12, "bold"),
                 bg=self._COL_BG, fg="#2c3e50").pack(side="left")
        self._dhl_status_lbl = tk.Label(hdr_dhl, text="Noch nicht geladen.",
                                        font=("Segoe UI", 9), bg=self._COL_BG, fg="#888")
        self._dhl_status_lbl.pack(side="right", padx=6)

        leg_dhl = tk.Frame(self._inner, bg=self._COL_BG)
        leg_dhl.pack(fill="x", padx=16, pady=(0, 6))
        for color, text in [(self._COL_NORMAL,   "🚛 DHL Normal"),
                            (self._COL_EXPRESS,  "🚚 DHL Express"),
                            (self._COL_ABHOLUNG, "🏃 Abholung")]:
            tk.Frame(leg_dhl, bg=color, width=14, height=14, relief="flat").pack(side="left")
            tk.Label(leg_dhl, text=f" {text}   ", font=("Segoe UI", 9),
                     bg=self._COL_BG, fg="#2c3e50").pack(side="left")

        # ── Gesamt-Kacheln (Woche + Monat) ──────────────────────────────
        cards_gesamt = tk.Frame(self._inner, bg=self._COL_BG)
        cards_gesamt.pack(fill="x", padx=16, pady=(4, 4))

        def _card_big(parent, col, icon, label1, label2, sub_fg):
            outer = tk.Frame(parent, bg=col, padx=2, pady=2)
            outer.pack(side="left", padx=6, expand=True, fill="x")
            inner = tk.Frame(outer, bg=col, padx=14, pady=12)
            inner.pack(fill="both")
            cnt = tk.Label(inner, text="–", font=("Segoe UI", 34, "bold"),
                           bg=col, fg="white", anchor="center")
            cnt.pack(fill="x")
            tk.Label(inner, text=f"{icon}  {label1}",
                     font=("Segoe UI", 10, "bold"), bg=col, fg="white").pack()
            tk.Label(inner, text=label2,
                     font=("Segoe UI", 8), bg=col, fg=sub_fg).pack()
            return cnt

        self._lbl_gesamt_woche = _card_big(cards_gesamt, "#2c3e50", "📦", "Gesamt", "Diese Woche",   "#bbdefb")
        self._lbl_gesamt_monat = _card_big(cards_gesamt, "#263238", "📦", "Gesamt", "Dieser Monat",  "#bbdefb")

        # ── Split-Kacheln (Normal / Express / Abholung) ───────────────
        cards_dhl = tk.Frame(self._inner, bg=self._COL_BG)
        cards_dhl.pack(fill="x", padx=16, pady=(0, 10))

        def _card_dhl(parent, col, icon, label1, label2, sub_fg):
            outer = tk.Frame(parent, bg=col, padx=2, pady=2)
            outer.pack(side="left", padx=6, expand=True, fill="x")
            inner = tk.Frame(outer, bg=col, padx=14, pady=10)
            inner.pack(fill="both")
            cnt = tk.Label(inner, text="–", font=("Segoe UI", 28, "bold"),
                           bg=col, fg="white", anchor="center")
            cnt.pack(fill="x")
            tk.Label(inner, text=f"{icon}  {label1}",
                     font=("Segoe UI", 9, "bold"), bg=col, fg="white").pack()
            tk.Label(inner, text=label2,
                     font=("Segoe UI", 8), bg=col, fg=sub_fg).pack()
            return cnt

        self._lbl_normal_woche   = _card_dhl(cards_dhl, self._COL_NORMAL,   "🚛", "DHL Normal",  "Diese Woche", "#fff9c4")
        self._lbl_express_woche  = _card_dhl(cards_dhl, self._COL_EXPRESS,  "🚚", "DHL Express", "Diese Woche", "#ef9a9a")
        self._lbl_abholung_woche = _card_dhl(cards_dhl, self._COL_ABHOLUNG, "🏃", "Abholung",    "Diese Woche", "#a5d6a7")

        tk.Frame(self._inner, bg="#dde2e8", height=1).pack(fill="x", padx=16, pady=(0, 8))

        chart_hdr_dhl = tk.Frame(self._inner, bg=self._COL_BG)
        chart_hdr_dhl.pack(fill="x", padx=20, pady=(0, 4))
        self._dhl_chart_title_lbl = tk.Label(chart_hdr_dhl, text="Letzte 4 Kalenderwochen",
                                             font=("Segoe UI", 10, "bold"),
                                             bg=self._COL_BG, fg="#2c3e50")
        self._dhl_chart_title_lbl.pack(side="left")

        dhl_toggle_frame = tk.Frame(chart_hdr_dhl, bg=self._COL_BG)
        dhl_toggle_frame.pack(side="right")
        self._dhl_toggle_btns = {}

        def _dhl_switch_view(mode):
            self._dhl_view_mode.set(mode)
            titles = {"daily":   "Letzte 4 Wochen – Tagesansicht",
                      "weekly":  "Letzte 4 Kalenderwochen",
                      "monthly": "Letzte 6 Monate"}
            self._dhl_chart_title_lbl.config(text=titles.get(mode, ""))
            for m, b in self._dhl_toggle_btns.items():
                b.config(bg="#2c3e50" if m == mode else "#dde2e8",
                         fg="white"   if m == mode else "#2c3e50",
                         relief="sunken" if m == mode else "flat")
            self._dhl_redraw_chart()

        for _mode, _label, _active in [("daily", "Tagesansicht", False),
                                        ("weekly", "Wochenübersicht", True),
                                        ("monthly", "Monatsübersicht", False)]:
            _b = tk.Button(dhl_toggle_frame, text=_label,
                           font=("Segoe UI", 8), padx=8, pady=2, bd=1,
                           bg="#2c3e50" if _active else "#dde2e8",
                           fg="white"   if _active else "#2c3e50",
                           relief="sunken" if _active else "flat",
                           command=lambda m=_mode: _dhl_switch_view(m))
            _b.pack(side="left", padx=(0, 2))
            self._dhl_toggle_btns[_mode] = _b

        self._dhl_canvas = tk.Canvas(self._inner, bg=self._COL_BG, height=260, highlightthickness=0)
        self._dhl_canvas.pack(fill="x", padx=20, pady=(4, 24))
        self._dhl_canvas.bind("<Configure>", lambda e: self._dhl_redraw_chart())

        self._bind_mousewheel = _bind_mousewheel

    # ── Öffentliche Methoden ──────────────────────────────────────────

    def update_main(self, df: "pd.DataFrame"):
        """Wird aufgerufen wenn die Abholer_DB neu geladen wird."""
        self._main_df = df.copy() if df is not None else None
        self._pu_recalculate()
        # DHL-Statistik neu berechnen, falls DHL-Daten schon geladen sind
        # (kann vorkommen wenn Abholer_DB später fertig ist als DHL-Daten)
        if self._normal_df is not None or self._express_df is not None:
            self._dhl_recalculate()

    def load_archive_async(self):
        """Startet das Laden der Archiv-Dateien im Hintergrund (einmalig beim Start)."""
        if self._pu_loading:
            return
        self._pu_loading = True
        self._pu_status_lbl.config(text="⏳  Lade Archiv …")
        self._start_loading_cb("⏳  Lade PU-Archiv aus Google Drive …")

        def _worker():
            try:
                archiv_df = fetch_archiv_gdrive()
                self.frame.after(0, lambda d=archiv_df: self._pu_on_archive_loaded(d))
            except Exception as e:
                self.frame.after(0, lambda err=e: self._pu_on_archive_error(err))

        threading.Thread(target=_worker, daemon=True).start()

    def load_dhl_async(self):
        """Lädt DHL Normal + DHL Express aus OrcaScan im Hintergrund."""
        if self._dhl_loading:
            return
        self._dhl_loading = True
        self._dhl_status_lbl.config(text="⏳  Lade DHL-Daten …")
        self._start_loading_cb("⏳  Lade DHL-Daten aus OrcaScan …")

        def _worker():
            try:
                _drop      = {"signature", "packagePhoto"}
                normal_df  = fetch_sheet_orca(ORCA_DHL_NORMAL_SHEET_ID, drop_cols=_drop)
                express_df = fetch_sheet_orca(ORCA_DHL_EX_SHEET_ID,     drop_cols=_drop)
                self.frame.after(0, lambda n=normal_df, e=express_df: self._dhl_on_loaded(n, e))
            except Exception as ex:
                self.frame.after(0, lambda err=ex: self._dhl_on_error(err))

        threading.Thread(target=_worker, daemon=True).start()

    # ── PU interne Methoden ──────────────────────────────────────────

    def _pu_on_archive_loaded(self, archiv_df: "pd.DataFrame"):
        self._pu_loading = False
        n = len(archiv_df) if not archiv_df.empty else 0
        self._archiv_df = archiv_df if not archiv_df.empty else None
        self._pu_status_lbl.config(text=f"✅  {n} archivierte Einträge geladen")
        self._stop_loading_cb(f"PU-Archiv geladen: {n} Einträge")
        self._pu_recalculate()

    def _pu_on_archive_error(self, err: Exception):
        self._pu_loading = False
        self._stop_loading_cb("❌  PU-Archiv nicht geladen")
        err_str = str(err).lower()

        if "invalid_grant" in err_str or "token has been expired" in err_str:
            title = "PU Statistik – Google Login erneuern"
            msg   = (
                "Dein Google-Login ist abgelaufen.\n\n"
                "Das alte Token wurde automatisch gelöscht.\n"
                "Beim nächsten Programmstart öffnet sich einmalig\n"
                "ein Browser-Fenster für den neuen Login.\n\n"
                "Die Statistik zeigt bis dahin nur die aktuellen\n"
                "Daten aus OrcaScan (ohne Archiv)."
            )
        elif "oauth_credentials.json" in err_str or "nicht gefunden" in err_str:
            title = "PU Statistik – Google Login nicht eingerichtet"
            msg   = (
                "Die Datei 'oauth_credentials.json' fehlt im Bombadil-Ordner.\n\n"
                "Diese wird für den Zugriff auf das Google Drive Archiv benötigt.\n\n"
                "Schritte:\n"
                "1. Google Cloud Console → APIs & Dienste → Anmeldedaten\n"
                "2. OAuth 2.0 Client-ID erstellen (Desktop-App)\n"
                "3. JSON herunterladen → als 'oauth_credentials.json'\n"
                "   in den Bombadil-Ordner legen"
            )
        else:
            title = "PU Statistik – Archiv nicht geladen"
            msg   = (
                f"Die Archiv-Dateien konnten nicht geladen werden.\n\n"
                f"Fehler: {err}\n\n"
                f"Die Statistik zeigt nur die aktuellen Daten aus OrcaScan.\n"
                f"Archivierte (gecleante) Einträge fehlen in der Auswertung."
            )

        self._pu_status_lbl.config(text="❌  Archiv nicht geladen – siehe Popup")
        messagebox.showwarning(title, msg)

    def _pu_recalculate(self):
        try:
            self._pu_recalculate_inner()
        except Exception as _e:
            import traceback as _tb
            self._pu_status_lbl.config(text=f"⚠  Fehler: {_e}")
            print("[PU Statistik] Fehler in _recalculate:\n" + _tb.format_exc())

    def _pu_recalculate_inner(self):
        frames = []
        if self._main_df is not None and not self._main_df.empty:
            frames.append(self._main_df)
        if self._archiv_df is not None and not self._archiv_df.empty:
            frames.append(self._archiv_df)

        if not frames:
            return

        combined = pd.concat(frames, ignore_index=True)

        # Duplikate (gleicher Barcode) entfernen – aktuell gewinnt über Archiv
        c_bc = first_existing(combined, COL_BARCODE)
        if c_bc:
            combined = combined.drop_duplicates(subset=[c_bc], keep="first")

        c_verpackt_at = first_existing(combined, COL_VERPACKT_AT)
        c_abgeholt    = first_existing(combined, COL_ABGEHOLT)
        c_abholbereit = first_existing(combined, COL_ABHOLBEREIT)
        c_status      = first_existing(combined, COL_STATUS)
        c_scan_date   = first_existing(combined, COL_SCAN_DATE)

        if not c_verpackt_at:
            self._pu_status_lbl.config(text="⚠  Datumsfelder nicht gefunden.")
            return

        def _parse(series):
            """Parst Datum-Strings → tz-naive datetime64. Unterstützt UTC-aware und naive."""
            s = pd.to_datetime(series, errors="coerce", utc=True)
            return s.dt.tz_convert(None)   # → immer tz-naive datetime64[ns]

        # Alle Datumsspalten einheitlich parsen (tz-aware UND tz-naive werden normiert)
        for col in [c_verpackt_at, c_abgeholt, c_abholbereit, c_scan_date]:
            if col:
                combined[col] = _parse(combined[col])

        # ── Effektives Abholdatum ────────────────────────────────────
        # Ein Paket zählt als abgeholt wenn Paketstatus="abgeholt" (egal ob Zeitstempel).
        # Datums-Fallback: Abgeholt_At → Abholbereit_At → Verpackt_At → Scan-Datum

        if c_status:
            is_abgeholt = combined[c_status].astype(str).str.strip().str.lower() == "abgeholt"
        else:
            is_abgeholt = pd.Series(False, index=combined.index)

        # Start: Abgeholt_At-Spalte (oder komplett leer)
        if c_abgeholt:
            abhol_eff = combined[c_abgeholt].copy()
        else:
            abhol_eff = pd.Series(pd.NaT, index=combined.index, dtype="datetime64[ns]")

        # Fallbacks per direkter Zuweisung (sicherer als .where() bei dtype-Mischung)
        for fb_col in [c_abholbereit, c_verpackt_at, c_scan_date]:
            if fb_col is None:
                continue
            mask = is_abgeholt & abhol_eff.isna()
            if not mask.any():
                break
            abhol_eff = abhol_eff.copy()
            abhol_eff[mask] = combined.loc[mask, fb_col].values

        combined["_abhol_eff"] = abhol_eff

        today      = today_date()
        week_start = today - timedelta(days=today.weekday())   # Montag
        month_start = today.replace(day=1)

        def count_in_range(col, start, end=None):
            s = combined[col].dt.date
            mask = s >= start
            if end:
                mask = mask & (s < end)
            return int(mask.sum())

        # ── Kacheln ──────────────────────────────────────────────────
        n_aw = count_in_range(c_verpackt_at, week_start)
        n_am = count_in_range(c_verpackt_at, month_start)
        n_bw = count_in_range("_abhol_eff",  week_start)
        n_bm = count_in_range("_abhol_eff",  month_start)

        self._lbl_anlief_woche.config(text=str(n_aw))
        self._lbl_anlief_monat.config(text=str(n_am))
        self._lbl_abhol_woche.config(text=str(n_bw))
        self._lbl_abhol_monat.config(text=str(n_bm))

        # ── Wochendaten für Chart (letzte 4 Kalenderwochen) ──────────
        weekly = []
        for i in range(3, -1, -1):   # älteste zuerst
            ws = week_start - timedelta(weeks=i)
            we = ws + timedelta(days=7)
            kw = ws.isocalendar()[1]
            label = f"KW {kw:02d}\n{ws.strftime('%d.%m.')}–{(we - timedelta(days=1)).strftime('%d.%m.')}"
            na = count_in_range(c_verpackt_at, ws, we)
            nb = count_in_range("_abhol_eff",  ws, we)
            weekly.append((label, na, nb))

        self._pu_weekly_data = weekly

        # ── Tagesdaten für Chart (letzte 4 Wochen = 28 Tage) ─────────
        daily = []
        for i in range(27, -1, -1):   # ältester Tag zuerst
            d      = today - timedelta(days=i)
            d_next = d + timedelta(days=1)
            label  = d.strftime("%d.%m.")
            na = count_in_range(c_verpackt_at, d, d_next)
            nb = count_in_range("_abhol_eff",  d, d_next)
            daily.append((label, na, nb, d.weekday()))  # weekday: 0=Mo
        self._pu_daily_data = daily

        # ── Monatsdaten für Chart (letzte 6 Monate) ──────────────────
        monthly = []
        for i in range(5, -1, -1):   # ältester Monat zuerst
            # Ersten Tag des Monats berechnen (i Monate zurück)
            m_year  = today.year
            m_month = today.month - i
            while m_month <= 0:
                m_month += 12
                m_year  -= 1
            ms = today.replace(year=m_year, month=m_month, day=1)
            # Erster Tag des nächsten Monats
            if ms.month == 12:
                me = ms.replace(year=ms.year + 1, month=1, day=1)
            else:
                me = ms.replace(month=ms.month + 1, day=1)
            label_m = ms.strftime("%b %y")   # z.B. "Mär 26"
            na_m = count_in_range(c_verpackt_at, ms, me)
            nb_m = count_in_range("_abhol_eff",  ms, me)
            monthly.append((label_m, na_m, nb_m))

        self._pu_monthly_data = monthly

        # ── Kiosk-Aufschlüsselung ─────────────────────────────────────
        c_zielkiosk = first_existing(combined, COL_ZIELKIOSK)
        if c_zielkiosk:
            six_months_ago = today.replace(day=1)
            m6 = today.month - 5
            y6 = today.year
            while m6 <= 0:
                m6 += 12
                y6 -= 1
            six_months_start = today.replace(year=y6, month=m6, day=1)

            def kiosk_counts(date_col, start, end=None):
                s = combined[date_col].dt.date
                mask = s >= start
                if end:
                    mask = mask & (s < end)
                sub = combined[mask].copy()
                sub["_zk"] = sub[c_zielkiosk].astype(str).str.strip()
                sub["_zk"] = sub["_zk"].replace({"nan": "–", "": "–"})
                g = sub.groupby("_zk").size().reset_index(name="n")
                return dict(zip(g["_zk"], g["n"]))

            def kiosk_table(start, end=None):
                anlief = kiosk_counts(c_verpackt_at, start, end)
                abhol  = kiosk_counts("_abhol_eff",  start, end)
                kioski = sorted(set(anlief) | set(abhol))
                rows = []
                for k in kioski:
                    rows.append([k, str(anlief.get(k, 0)), str(abhol.get(k, 0))])
                # Summenzeile
                rows.append(["Gesamt",
                              str(sum(int(r[1]) for r in rows)),
                              str(sum(int(r[2]) for r in rows))])
                return rows

            self._kiosk_data = {
                "week":    kiosk_table(week_start),
                "month":   kiosk_table(month_start),
                "months6": kiosk_table(six_months_start),
            }
        else:
            self._kiosk_data = {"week": [], "month": [], "months6": []}

        self._pu_redraw_kiosk_table()
        self._bind_mousewheel(self._inner)

        ts = datetime.now().strftime("%d.%m.%Y %H:%M")
        n_total = len(combined)
        n_archiv = len(self._archiv_df) if self._archiv_df is not None else 0
        self._pu_status_lbl.config(
            text=f"Stand {ts}  –  {n_total} Einträge gesamt  ({len(frames[0])} aktuell + {n_archiv} Archiv)"
        )
        self._pu_redraw_chart()

    def _pu_switch_kiosk_range(self, mode: str):
        self._kiosk_range.set(mode)
        for m, b in self._kiosk_btns.items():
            if m == mode:
                b.config(bg="#2c3e50", fg="white", relief="sunken")
            else:
                b.config(bg="#dde2e8", fg="#2c3e50", relief="flat")
        self._pu_redraw_kiosk_table()

    def _pu_redraw_kiosk_table(self):
        c = self._kiosk_canvas
        c.delete("all")
        rows = self._kiosk_data.get(self._kiosk_range.get(), [])
        if not rows:
            c.create_text(10, 20, text="Bitte Abholer_DB laden …",
                          anchor="nw", font=("Segoe UI", 10), fill="#999")
            return

        # rows = list of [kiosk, anlief_str, abhol_str]
        # Summenzeile (letzte) separat behandeln
        data_rows = rows[:-1]
        sum_row   = rows[-1]

        PAD_LEFT  = 110
        PAD_RIGHT = 14
        BAR_H     = 18
        BAR_GAP   = 3
        GROUP_GAP = 14
        PAD_TOP   = 10

        canvas_w  = c.winfo_width() or 360
        bar_area_w = canvas_w - PAD_LEFT - PAD_RIGHT

        max_val = max(
            (max(int(r[1]), int(r[2])) for r in data_rows if r[1].isdigit() or r[2].isdigit()),
            default=1
        ) or 1

        for i, row in enumerate(data_rows):
            kiosk, n_anlief_s, n_abhol_s = row
            n_anlief = int(n_anlief_s) if n_anlief_s.isdigit() else 0
            n_abhol  = int(n_abhol_s)  if n_abhol_s.isdigit()  else 0
            y0   = PAD_TOP + i * (2 * BAR_H + BAR_GAP + GROUP_GAP)
            y_mid = y0 + BAR_H

            # Kiosk-Label links
            c.create_text(PAD_LEFT - 8, y_mid, text=kiosk,
                          anchor="e", font=("Segoe UI", 9, "bold"), fill="#2c3e50")

            # Bar 1: Lieferungen (teal)
            bw1 = max(int(n_anlief / max_val * bar_area_w), 4) if n_anlief else 0
            c.create_rectangle(PAD_LEFT, y0, PAD_LEFT + bar_area_w, y0 + BAR_H,
                                fill="#d5eae7", outline="")
            if bw1:
                c.create_rectangle(PAD_LEFT, y0, PAD_LEFT + bw1, y0 + BAR_H,
                                   fill=self._COL_ANLIEF, outline="")
            c.create_text(PAD_LEFT + bw1 + 6, y0 + BAR_H // 2,
                          text=f"🚐 {n_anlief}", anchor="w",
                          font=("Segoe UI", 9, "bold"), fill="#117a65")

            # Bar 2: Abholungen (indigo)
            y2  = y0 + BAR_H + BAR_GAP
            bw2 = max(int(n_abhol / max_val * bar_area_w), 4) if n_abhol else 0
            c.create_rectangle(PAD_LEFT, y2, PAD_LEFT + bar_area_w, y2 + BAR_H,
                                fill="#d0d5ec", outline="")
            if bw2:
                c.create_rectangle(PAD_LEFT, y2, PAD_LEFT + bw2, y2 + BAR_H,
                                   fill=self._COL_ABHOL, outline="")
            c.create_text(PAD_LEFT + bw2 + 6, y2 + BAR_H // 2,
                          text=f"📦 {n_abhol}", anchor="w",
                          font=("Segoe UI", 9, "bold"), fill="#283593")

        # Summenzeile
        y_sum = PAD_TOP + len(data_rows) * (2 * BAR_H + BAR_GAP + GROUP_GAP)
        c.create_line(PAD_LEFT - 8, y_sum, canvas_w - PAD_RIGHT, y_sum,
                      fill="#aaa", dash=(4, 2))
        y_sum += 6
        c.create_text(PAD_LEFT - 8, y_sum + BAR_H // 2,
                      text=f"Gesamt: 🚐 {sum_row[1]}  📦 {sum_row[2]}",
                      anchor="e", font=("Segoe UI", 9, "bold"), fill="#2c3e50")

        # Canvas-Höhe anpassen
        needed_h = y_sum + BAR_H + 10
        c.config(height=max(needed_h, 60))

    def _pu_redraw_chart(self):
        view = self._pu_view_mode.get()
        if view == "monthly":
            data = self._pu_monthly_data
        elif view == "daily":
            data = self._pu_daily_data
        else:
            data = self._pu_weekly_data
        c = self._pu_canvas
        c.delete("all")
        if not data:
            c.create_text(10, 20, text="Bitte Abholer_DB laden …",
                          anchor="nw", font=("Segoe UI", 10), fill="#999")
            return
        if view == "daily":
            self._pu_redraw_chart_daily(data)
            return

        PAD_LEFT   = 130
        PAD_RIGHT  = 60
        BAR_H      = 20
        BAR_GAP    = 3    # zwischen den zwei Balken einer Gruppe
        GROUP_GAP  = 18   # zwischen den Gruppen
        PAD_TOP    = 14

        canvas_w = c.winfo_width() or 820
        bar_area_w = canvas_w - PAD_LEFT - PAD_RIGHT
        max_cnt = max(max(na, nb) for _, na, nb in data) or 1

        group_h = 2 * BAR_H + BAR_GAP

        for i, (label, n_anlief, n_abhol) in enumerate(data):
            y0 = PAD_TOP + i * (group_h + GROUP_GAP)
            y_mid = y0 + group_h // 2

            # Wochen-Label links
            c.create_text(PAD_LEFT - 8, y_mid, text=label,
                          anchor="e", font=("Segoe UI", 9, "bold"), fill="#2c3e50")

            # ── Bar 1: Lieferungen (teal) ──
            y1 = y0
            bw1 = max(int(n_anlief / max_cnt * bar_area_w), 4) if n_anlief else 0
            c.create_rectangle(PAD_LEFT, y1, PAD_LEFT + bar_area_w, y1 + BAR_H,
                                fill="#d5eae7", outline="")
            if bw1:
                c.create_rectangle(PAD_LEFT, y1, PAD_LEFT + bw1, y1 + BAR_H,
                                   fill=self._COL_ANLIEF, outline="")
            c.create_text(PAD_LEFT + bw1 + 6, y1 + BAR_H // 2,
                          text=f"🚐 {n_anlief}", anchor="w",
                          font=("Segoe UI", 9, "bold"), fill="#117a65")

            # ── Bar 2: Abholungen (indigo) ──
            y2 = y0 + BAR_H + BAR_GAP
            bw2 = max(int(n_abhol / max_cnt * bar_area_w), 4) if n_abhol else 0
            c.create_rectangle(PAD_LEFT, y2, PAD_LEFT + bar_area_w, y2 + BAR_H,
                                fill="#d0d5ec", outline="")
            if bw2:
                c.create_rectangle(PAD_LEFT, y2, PAD_LEFT + bw2, y2 + BAR_H,
                                   fill=self._COL_ABHOL, outline="")
            c.create_text(PAD_LEFT + bw2 + 6, y2 + BAR_H // 2,
                          text=f"📦 {n_abhol}", anchor="w",
                          font=("Segoe UI", 9, "bold"), fill="#283593")

        # Canvas-Höhe anpassen
        needed_h = PAD_TOP + len(data) * (group_h + GROUP_GAP)
        c.config(height=max(needed_h + 10, 120))

    def _pu_redraw_chart_daily(self, data):
        _WOCHENTAGE = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

        c = self._pu_canvas
        canvas_w  = c.winfo_width() or 820
        canvas_h  = 230
        c.config(height=canvas_h)

        PAD_LEFT   = 12
        PAD_RIGHT  = 12
        PAD_TOP    = 18
        PAD_BOTTOM = 44   # Platz für Datum + Wochentag

        chart_w = canvas_w - PAD_LEFT - PAD_RIGHT
        chart_h = canvas_h - PAD_TOP - PAD_BOTTOM
        n_days  = len(data)
        group_w = chart_w / n_days
        bar_w   = max(int(group_w * 0.65), 3)
        base_y  = canvas_h - PAD_BOTTOM

        max_cnt = max((na for _, na, _, _ in data), default=1) or 1

        # Hilfslinie bei halber Höhe
        y_half = PAD_TOP + chart_h // 2
        c.create_line(PAD_LEFT, y_half, canvas_w - PAD_RIGHT, y_half,
                      fill="#eeeeee", dash=(2, 4))
        c.create_text(PAD_LEFT - 2, y_half, text=str(max_cnt // 2),
                      anchor="e", font=("Segoe UI", 7), fill="#bbb")

        for i, (label, n_anlief, _n_abhol, weekday) in enumerate(data):
            gx = PAD_LEFT + (i + 0.5) * group_w
            x_l = int(gx - bar_w / 2)
            x_r = int(gx + bar_w / 2)

            # Trennlinie bei jedem Montag
            if weekday == 0 and i > 0:
                x_sep = PAD_LEFT + i * group_w
                c.create_line(x_sep, PAD_TOP, x_sep, base_y,
                              fill="#dde2e8", dash=(2, 3))

            # Balken: Verpackt (teal)
            h = max(int(n_anlief / max_cnt * chart_h), 1) if n_anlief else 0
            c.create_rectangle(x_l, PAD_TOP, x_r, base_y, fill="#d5eae7", outline="")
            if h:
                c.create_rectangle(x_l, base_y - h, x_r, base_y,
                                   fill=self._COL_ANLIEF, outline="")
            if n_anlief and bar_w >= 5:
                c.create_text((x_l + x_r) // 2, min(base_y - h - 2, base_y - 4),
                              text=str(n_anlief), anchor="s",
                              font=("Segoe UI", 7), fill="#117a65")

            # X-Achsen-Labels: Wochentag (oben) + Datum (unten)
            wt_txt = _WOCHENTAGE[weekday]
            is_mo  = weekday == 0
            c.create_text(gx, base_y + 4, text=wt_txt, anchor="n",
                          font=("Segoe UI", 7, "bold" if is_mo else ""),
                          fill="#2c3e50" if is_mo else "#666")
            c.create_text(gx, base_y + 16, text=label, anchor="n",
                          font=("Segoe UI", 7), fill="#999")

        # X-Achsen-Linie
        c.create_line(PAD_LEFT, base_y, canvas_w - PAD_RIGHT, base_y, fill="#aaa")

    # ── DHL interne Methoden ─────────────────────────────────────────

    def _dhl_on_loaded(self, normal_df, express_df):
        self._dhl_loading = False
        self._normal_df  = normal_df  if not normal_df.empty  else None
        self._express_df = express_df if not express_df.empty else None
        n_n = len(normal_df)  if not normal_df.empty  else 0
        n_e = len(express_df) if not express_df.empty else 0
        self._dhl_status_lbl.config(text=f"✅  {n_n} DHL Normal + {n_e} DHL Express geladen")
        self._stop_loading_cb(f"DHL geladen: {n_n} Normal + {n_e} Express")
        self._dhl_recalculate()

    def _dhl_on_error(self, err):
        self._dhl_loading = False
        self._stop_loading_cb(f"❌  DHL-Fehler: {err}")
        self._dhl_status_lbl.config(text=f"❌  Fehler: {err}")

    def _dhl_recalculate(self):
        try:
            self._dhl_recalculate_inner()
        except Exception as _e:
            import traceback as _tb
            self._dhl_status_lbl.config(text=f"⚠  Fehler: {_e}")
            print("[DHL Statistik] Fehler:\n" + _tb.format_exc())

    def _dhl_recalculate_inner(self):
        def _to_dates(df):
            if df is None or df.empty:
                return pd.Series([], dtype="object")
            c_scan = first_existing(df, ORCA_COL_SCAN)
            if not c_scan:
                return pd.Series([], dtype="object")
            return (pd.to_datetime(df[c_scan], errors="coerce", utc=True)
                    .dt.tz_convert(None).dt.date.dropna())

        def _abholung_dates_from_main():
            df = self._main_df
            if df is None or df.empty:
                return pd.Series([], dtype="object")
            c_abgeholt    = first_existing(df, COL_ABGEHOLT)
            c_abholbereit = first_existing(df, COL_ABHOLBEREIT)
            c_status      = first_existing(df, COL_STATUS)
            if not c_status:
                return pd.Series([], dtype="object")
            status_norm = df[c_status].fillna("").astype(str).str.strip().str.lower()
            abgeholt_df = df[status_norm == "abgeholt"].copy()
            # Datum: Abgeholt_At, Fallback auf Abholbereit_At
            if c_abgeholt:
                eff = pd.to_datetime(abgeholt_df[c_abgeholt], errors="coerce", utc=True)
            else:
                eff = pd.Series(pd.NaT, index=abgeholt_df.index)
            if c_abholbereit:
                fb  = pd.to_datetime(abgeholt_df[c_abholbereit], errors="coerce", utc=True)
                eff = eff.fillna(fb)
            return eff.dt.tz_convert(None).dt.date.dropna()

        normal_dates   = _to_dates(self._normal_df)
        express_dates  = _to_dates(self._express_df)
        abholung_dates = _abholung_dates_from_main()

        today       = today_date()
        week_start  = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)

        def count(dates, start, end=None):
            if dates.empty:
                return 0
            m = dates >= start
            if end:
                m = m & (dates < end)
            return int(m.sum())

        # Kacheln
        n_woche  = count(normal_dates,   week_start)  + count(express_dates,  week_start)  + count(abholung_dates, week_start)
        n_monat  = count(normal_dates,   month_start) + count(express_dates,  month_start) + count(abholung_dates, month_start)
        self._lbl_gesamt_woche.config(   text=str(n_woche))
        self._lbl_gesamt_monat.config(   text=str(n_monat))
        self._lbl_normal_woche.config(   text=str(count(normal_dates,   week_start)))
        self._lbl_express_woche.config(  text=str(count(express_dates,  week_start)))
        self._lbl_abholung_woche.config( text=str(count(abholung_dates, week_start)))

        # Wochendaten (letzte 4 Kalenderwochen)
        weekly = []
        for i in range(3, -1, -1):
            ws = week_start - timedelta(weeks=i)
            we = ws + timedelta(days=7)
            kw = ws.isocalendar()[1]
            lbl = f"KW {kw:02d}\n{ws.strftime('%d.%m.')}–{(we - timedelta(days=1)).strftime('%d.%m.')}"
            weekly.append((lbl, count(normal_dates, ws, we),
                           count(express_dates, ws, we), count(abholung_dates, ws, we)))
        self._dhl_weekly_data = weekly

        # Tagesdaten (letzte 28 Tage)
        daily = []
        for i in range(27, -1, -1):
            d      = today - timedelta(days=i)
            d_next = d + timedelta(days=1)
            daily.append((d.strftime("%d.%m."),
                          count(normal_dates,   d, d_next),
                          count(express_dates,  d, d_next),
                          count(abholung_dates, d, d_next),
                          d.weekday()))
        self._dhl_daily_data = daily

        # Monatsdaten (letzte 6 Monate)
        monthly = []
        for i in range(5, -1, -1):
            m_year  = today.year
            m_month = today.month - i
            while m_month <= 0:
                m_month += 12
                m_year  -= 1
            ms = today.replace(year=m_year, month=m_month, day=1)
            me = ms.replace(month=ms.month + 1, day=1) if ms.month < 12 \
                 else ms.replace(year=ms.year + 1, month=1, day=1)
            monthly.append((ms.strftime("%b %y"),
                            count(normal_dates,   ms, me),
                            count(express_dates,  ms, me),
                            count(abholung_dates, ms, me)))
        self._dhl_monthly_data = monthly

        self._bind_mousewheel(self._inner)
        ts  = datetime.now().strftime("%d.%m.%Y %H:%M")
        n_n = len(self._normal_df)  if self._normal_df  is not None else 0
        n_e = len(self._express_df) if self._express_df is not None else 0
        n_a = int(abholung_dates.count()) if not abholung_dates.empty else 0
        self._dhl_status_lbl.config(
            text=f"Stand {ts}  –  {n_n} Normal · {n_e} Express · {n_a} Abholungen")
        self._dhl_redraw_chart()

    def _dhl_redraw_chart(self):
        view = self._dhl_view_mode.get()
        if view == "monthly":
            data = self._dhl_monthly_data
        elif view == "daily":
            data = self._dhl_daily_data
        else:
            data = self._dhl_weekly_data
        c = self._dhl_canvas
        c.delete("all")
        if not data:
            c.create_text(10, 20, text="Bitte Daten laden …",
                          anchor="nw", font=("Segoe UI", 10), fill="#999")
            return
        if view == "daily":
            self._dhl_redraw_chart_daily(data)
            return

        PAD_LEFT   = 130
        PAD_RIGHT  = 60
        BAR_H      = 20
        BAR_GAP    = 3
        GROUP_GAP  = 18
        PAD_TOP    = 14

        canvas_w   = c.winfo_width() or 820
        bar_area_w = canvas_w - PAD_LEFT - PAD_RIGHT
        max_cnt    = max(max(na, nb, nc) for _, na, nb, nc in data) or 1
        group_h    = 3 * BAR_H + 2 * BAR_GAP

        for i, (label, n_normal, n_express, n_abholung) in enumerate(data):
            y0    = PAD_TOP + i * (group_h + GROUP_GAP)
            y_mid = y0 + group_h // 2

            # Gesamt rechts oben anzeigen
            n_ges = n_normal + n_express + n_abholung
            c.create_text(PAD_LEFT - 8, y_mid - 6, text=label,
                          anchor="e", font=("Segoe UI", 9, "bold"), fill="#2c3e50")
            c.create_text(PAD_LEFT - 8, y_mid + 8, text=f"∑ {n_ges}",
                          anchor="e", font=("Segoe UI", 8), fill="#888")

            # Bar 1: DHL Normal
            y1  = y0
            bw1 = max(int(n_normal / max_cnt * bar_area_w), 4) if n_normal else 0
            c.create_rectangle(PAD_LEFT, y1, PAD_LEFT + bar_area_w, y1 + BAR_H,
                                fill="#fff9c4", outline="")
            if bw1:
                c.create_rectangle(PAD_LEFT, y1, PAD_LEFT + bw1, y1 + BAR_H,
                                   fill=self._COL_NORMAL, outline="")
            c.create_text(PAD_LEFT + bw1 + 6, y1 + BAR_H // 2,
                          text=f"🚛 {n_normal}", anchor="w",
                          font=("Segoe UI", 9, "bold"), fill="#f57f17")

            # Bar 2: DHL Express
            y2  = y0 + BAR_H + BAR_GAP
            bw2 = max(int(n_express / max_cnt * bar_area_w), 4) if n_express else 0
            c.create_rectangle(PAD_LEFT, y2, PAD_LEFT + bar_area_w, y2 + BAR_H,
                                fill="#ffcdd2", outline="")
            if bw2:
                c.create_rectangle(PAD_LEFT, y2, PAD_LEFT + bw2, y2 + BAR_H,
                                   fill=self._COL_EXPRESS, outline="")
            c.create_text(PAD_LEFT + bw2 + 6, y2 + BAR_H // 2,
                          text=f"🚚 {n_express}", anchor="w",
                          font=("Segoe UI", 9, "bold"), fill="#7f0000")

            # Bar 3: Abholung
            y3  = y0 + 2 * (BAR_H + BAR_GAP)
            bw3 = max(int(n_abholung / max_cnt * bar_area_w), 4) if n_abholung else 0
            c.create_rectangle(PAD_LEFT, y3, PAD_LEFT + bar_area_w, y3 + BAR_H,
                                fill="#c8e6c9", outline="")
            if bw3:
                c.create_rectangle(PAD_LEFT, y3, PAD_LEFT + bw3, y3 + BAR_H,
                                   fill=self._COL_ABHOLUNG, outline="")
            c.create_text(PAD_LEFT + bw3 + 6, y3 + BAR_H // 2,
                          text=f"🏃 {n_abholung}", anchor="w",
                          font=("Segoe UI", 9, "bold"), fill="#1b5e20")

        needed_h = PAD_TOP + len(data) * (group_h + GROUP_GAP)
        c.config(height=max(needed_h + 10, 120))

    def _dhl_redraw_chart_daily(self, data):
        _WOCHENTAGE = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]

        c = self._dhl_canvas
        canvas_w  = c.winfo_width() or 820
        canvas_h  = 250
        c.config(height=canvas_h)

        PAD_LEFT   = 12
        PAD_RIGHT  = 12
        PAD_TOP    = 18
        PAD_BOTTOM = 44

        chart_w = canvas_w - PAD_LEFT - PAD_RIGHT
        chart_h = canvas_h - PAD_TOP - PAD_BOTTOM
        n_days  = len(data)
        group_w = chart_w / n_days
        # 3 Balken pro Tag: etwas schmaler, mittig gruppiert
        bar_w   = max(int(group_w * 0.24), 2)
        bar_gap = max(int(group_w * 0.04), 1)
        base_y  = canvas_h - PAD_BOTTOM

        max_cnt = max((max(na, nb, nc) for _, na, nb, nc, _ in data), default=1) or 1

        y_half = PAD_TOP + chart_h // 2
        c.create_line(PAD_LEFT, y_half, canvas_w - PAD_RIGHT, y_half,
                      fill="#eeeeee", dash=(2, 4))
        c.create_text(PAD_LEFT - 2, y_half, text=str(max_cnt // 2),
                      anchor="e", font=("Segoe UI", 7), fill="#bbb")

        for i, (label, n_normal, n_express, n_abholung, weekday) in enumerate(data):
            gx = PAD_LEFT + (i + 0.5) * group_w

            if weekday == 0 and i > 0:
                x_sep = PAD_LEFT + i * group_w
                c.create_line(x_sep, PAD_TOP, x_sep, base_y,
                              fill="#dde2e8", dash=(2, 3))

            # 3 Balken nebeneinander, mittig zentriert
            total_w = 3 * bar_w + 2 * bar_gap
            x_start = int(gx - total_w / 2)

            # Bar 1: DHL Normal
            x1_l = x_start
            x1_r = x1_l + bar_w
            h1   = max(int(n_normal / max_cnt * chart_h), 1) if n_normal else 0
            c.create_rectangle(x1_l, PAD_TOP, x1_r, base_y, fill="#fff9c4", outline="")
            if h1:
                c.create_rectangle(x1_l, base_y - h1, x1_r, base_y,
                                   fill=self._COL_NORMAL, outline="")

            # Bar 2: DHL Express
            x2_l = x1_r + bar_gap
            x2_r = x2_l + bar_w
            h2   = max(int(n_express / max_cnt * chart_h), 1) if n_express else 0
            c.create_rectangle(x2_l, PAD_TOP, x2_r, base_y, fill="#ffcdd2", outline="")
            if h2:
                c.create_rectangle(x2_l, base_y - h2, x2_r, base_y,
                                   fill=self._COL_EXPRESS, outline="")

            # Bar 3: Abholung
            x3_l = x2_r + bar_gap
            x3_r = x3_l + bar_w
            h3   = max(int(n_abholung / max_cnt * chart_h), 1) if n_abholung else 0
            c.create_rectangle(x3_l, PAD_TOP, x3_r, base_y, fill="#c8e6c9", outline="")
            if h3:
                c.create_rectangle(x3_l, base_y - h3, x3_r, base_y,
                                   fill=self._COL_ABHOLUNG, outline="")

            # Tages-Gesamtzahl über der Gruppe
            n_total = n_normal + n_express + n_abholung
            if n_total:
                max_h = max(h1, h2, h3)
                x_mid = (x1_l + x3_r) // 2
                c.create_text(x_mid, min(base_y - max_h - 3, base_y - 4),
                              text=str(n_total), anchor="s",
                              font=("Segoe UI", 7), fill="#444")

            is_mo = weekday == 0
            c.create_text(gx, base_y + 4, text=_WOCHENTAGE[weekday], anchor="n",
                          font=("Segoe UI", 7, "bold" if is_mo else ""),
                          fill="#2c3e50" if is_mo else "#666")
            c.create_text(gx, base_y + 16, text=label, anchor="n",
                          font=("Segoe UI", 7), fill="#999")

        c.create_line(PAD_LEFT, base_y, canvas_w - PAD_RIGHT, base_y, fill="#aaa")


# ============================================================
# GUI – Tagesboten Abgleich Tab
# ============================================================
class TagesbotenAbgleichTab:
    """
    Vergleicht Tagesboten-Liste (Google Drive) mit Abholer_DB.

    Sub-Tab 1  Fehlerliste_dd/mm/yy  – in DB, aber Abholbereit_At leer
    Sub-Tab 2  Error_dd/mm/yy        – Barcode gar nicht in DB (Auto-Export + Banner)
    """

    _COLS = [
        ("barcode",  "Paket-Barcode", 220),
        ("name",     "Name",          240),
        ("datum",    "Datum",         160),
        ("zielkiosk","Ziel-Kiosk",    160),
        ("status",   "Status",        120),
        ("zahlung",  "Zahlung",       120),
    ]

    _AUTO_REFRESH_MS = 10 * 60 * 1000  # 10 Minuten

    def __init__(self, parent, get_abholer_df=None, get_abholer_path=None, get_export_folder=None):
        self.get_abholer_df    = get_abholer_df
        self.get_export_folder = get_export_folder or (lambda: Path.home() / "Downloads")
        self._last_fehler_df   = None
        self._last_error_df    = None
        self._auto_refresh_job = None

        self.frame = tk.Frame(parent)

        # ── Kopfzeile ───────────────────────────────────────────────
        tk.Label(self.frame, text="Tagesbote Abgleich",
                 anchor="w", font=("Segoe UI", 9, "bold")).pack(fill="x", pady=(4, 2))

        btn_row = tk.Frame(self.frame)
        btn_row.pack(fill="x", pady=(0, 4))
        b_tb_run = tk.Button(btn_row, text="Tagesbote laden & abgleichen", command=self._run)
        b_tb_run.pack(side="left")
        add_tooltip(b_tb_run, "Lädt die Tagesboten-Liste aus Google Drive\nund gleicht sie mit der Abholer_DB (OrcaScan) ab.\nNur Einträge mit Kontrollstatus 'Verpackt' werden geprüft.")
        b_tb_exp = tk.Button(btn_row, text="Fehlerliste exportieren (CSV)", command=self._export_fehlerliste)
        b_tb_exp.pack(side="left", padx=8)
        add_tooltip(b_tb_exp, "Exportiert die Fehlerliste als CSV in den Exportordner.\n(Pakete in Abholer_DB vorhanden, aber Abholbereit_At fehlt)")
        b_tb_clr = tk.Button(btn_row, text="Leeren", command=self._clear)
        b_tb_clr.pack(side="left")
        add_tooltip(b_tb_clr, "Alle Ergebnisse leeren und Tab zurücksetzen.")

        self.status_lbl = tk.Label(
            self.frame,
            text="'Tagesbote laden & abgleichen' klicken.",
            anchor="w",
        )
        self.status_lbl.pack(fill="x", pady=(0, 2))

        self._progress = ttk.Progressbar(self.frame, mode="indeterminate", length=300)
        self._progress.pack(anchor="w", pady=(0, 4))
        self._progress.pack_forget()   # anfangs versteckt

        # ── Rotes Banner – Fehlerliste (anfangs versteckt) ─────────
        self.notif_frame_fehler = tk.Frame(self.frame, bg="#c0392b")
        self.notif_lbl_fehler   = tk.Label(
            self.notif_frame_fehler, text="", bg="#c0392b", fg="white",
            font=("Segoe UI", 9, "bold"), anchor="w", padx=10, pady=5,
        )
        self.notif_lbl_fehler.pack(side="left", fill="x", expand=True)
        tk.Button(
            self.notif_frame_fehler, text="✕", bg="#a93226", fg="white",
            relief="flat", font=("Segoe UI", 9, "bold"), cursor="hand2",
            padx=8, pady=3,
            command=lambda: self.notif_frame_fehler.pack_forget(),
        ).pack(side="right", padx=4, pady=2)

        # ── Rotes Banner – Errorliste (anfangs versteckt) ───────────
        self.notif_frame = tk.Frame(self.frame, bg="#c0392b")
        self.notif_lbl   = tk.Label(
            self.notif_frame, text="", bg="#c0392b", fg="white",
            font=("Segoe UI", 9, "bold"), anchor="w", padx=10, pady=5,
        )
        self.notif_lbl.pack(side="left", fill="x", expand=True)
        tk.Button(
            self.notif_frame, text="✕", bg="#a93226", fg="white",
            relief="flat", font=("Segoe UI", 9, "bold"), cursor="hand2",
            padx=8, pady=3,
            command=lambda: self.notif_frame.pack_forget(),
        ).pack(side="right", padx=4, pady=2)
        # Banner werden erst bei Bedarf eingeblendet (pack_forget bis dahin)

        # ── Inneres Notebook mit zwei Sub-Tabs ─────────────────────
        self.inner_nb = ttk.Notebook(self.frame)
        self.inner_nb.pack(fill="both", expand=True)

        today_str = today_date().strftime("%d/%m/%y")

        self.sub_fehlerliste = TableTab(
            self.inner_nb,
            title=f"Fehlerliste {today_str}  –  in Abholer_DB vorhanden, aber Abholbereit_At fehlt",
            columns=self._COLS,
        )
        self.inner_nb.add(self.sub_fehlerliste.frame, text=f"Fehlerliste_{today_str}")

        # Abholbereit-Button in der Fehlerliste
        self._fehler_ids: list = []
        self._abholbereit_btn = tk.Button(
            self.sub_fehlerliste.btn_frame,
            text="✅  Pakete → Abholbereit setzen",
            command=self._set_abholbereit_orca,
            bg="#1a7a4a", fg="white", relief="flat",
            font=("Segoe UI", 9, "bold"), cursor="hand2",
            activebackground="#145e38", activeforeground="white",
            padx=10, pady=3, state="disabled"
        )
        self._abholbereit_btn.pack(side="left", padx=(16, 0))
        add_tooltip(self._abholbereit_btn,
                    "Setzt für alle Pakete in der Fehlerliste direkt in OrcaScan:\n"
                    "  • Abholbereit_At = aktuelle Uhrzeit\n"
                    "  • Paketstatus = Abholbereit")

        # Button: Auswahl → Abholbereit setzen
        self._abholbereit_sel_btn = tk.Button(
            self.sub_fehlerliste.btn_frame,
            text="✅  Auswahl → Abholbereit setzen",
            command=self._set_selected_abholbereit_orca,
            bg="#1e8449", fg="white", relief="flat",
            font=("Segoe UI", 9, "bold"), cursor="hand2",
            activebackground="#196f3d", activeforeground="white",
            padx=8, pady=3, state="disabled"
        )
        self._abholbereit_sel_btn.pack(side="left", padx=(8, 0))
        add_tooltip(self._abholbereit_sel_btn,
                    "Setzt nur die markierten Zeilen auf 'Abholbereit'.\n"
                    "Zeile(n) anklicken, dann diesen Button drücken.")

        self.sub_errorliste = TableTab(
            self.inner_nb,
            title=f"Error {today_str}  –  Barcode NICHT in Abholer_DB",
            columns=self._COLS,
        )
        self.inner_nb.add(self.sub_errorliste.frame, text=f"Error_{today_str}")

        # Löschen-Button im Error-Tab
        b_del = tk.Button(
            self.sub_errorliste.btn_frame,
            text="🗑  Auswahl löschen",
            command=self.sub_errorliste.delete_selected_rows,
            bg="#c0392b", fg="white", relief="flat",
            font=("Segoe UI", 9, "bold"), cursor="hand2",
            activebackground="#a93226", activeforeground="white",
            padx=8, pady=4,
        )
        b_del.pack(side="left", padx=(16, 0))
        add_tooltip(b_del,
                    "Markierte Zeile(n) aus der Errorliste entfernen.\n"
                    "Zeile anklicken, dann diesen Button drücken.\n"
                    "(Nur aus der Ansicht – OrcaScan wird nicht verändert)")

        # Button: Auswahl in Abholer_DB anlegen
        b_create_sel = tk.Button(
            self.sub_errorliste.btn_frame,
            text="📥  Auswahl → Abholer_DB anlegen",
            command=self._create_selected_in_abholer_orca,
            bg="#1f618d", fg="white", relief="flat",
            font=("Segoe UI", 9, "bold"), cursor="hand2",
            activebackground="#1a5276", activeforeground="white",
            padx=8, pady=4,
        )
        b_create_sel.pack(side="left", padx=(8, 0))
        add_tooltip(b_create_sel,
                    "Nur die markierten Zeilen in der Abholer_DB anlegen.\n"
                    "Zeile(n) anklicken, dann diesen Button drücken.")

        # Button: Fehler in Abholer_DB anlegen
        self._create_btn = tk.Button(
            self.sub_errorliste.btn_frame,
            text="📥  Fehler → Abholer_DB anlegen",
            command=self._create_in_abholer_orca,
            bg="#1a5276", fg="white", relief="flat",
            font=("Segoe UI", 9, "bold"), cursor="hand2",
            activebackground="#154360", activeforeground="white",
            padx=10, pady=3, state="disabled",
        )
        self._create_btn.pack(side="left", padx=(16, 0))
        add_tooltip(self._create_btn,
                    "Legt alle Barcodes aus der Errorliste direkt in OrcaScan\n"
                    "als neue Einträge in der Abholer_DB an.\n"
                    "Felder: Barcode, Name, Datum, Ziel-Kiosk, Zahlung\n"
                    "Paketstatus wird auf 'Verpackt' gesetzt.")

        # Callback: rotes Banner ausblenden wenn Errorliste leer
        self.sub_errorliste._on_rows_changed = self._check_error_banner

    # ── Aktionen ───────────────────────────────────────────────────
    def _safe_config(self, widget, **kwargs):
        """Widget-Update nur wenn Widget noch existiert (Thread-sicher)."""
        try:
            if widget.winfo_exists():
                widget.config(**kwargs)
        except Exception:
            pass

    def _start_progress(self, text: str):
        self._safe_config(self.status_lbl, text=text)
        try:
            self._progress.pack(anchor="w", pady=(0, 4))
            self._progress.start(12)
        except Exception:
            pass

    def _stop_progress(self, text: str):
        try:
            self._progress.stop()
            self._progress.pack_forget()
        except Exception:
            pass
        self._safe_config(self.status_lbl, text=text)

    def _finish_zahlung_update(self, result: dict):
        ok     = len(result.get("ok", []))
        failed = result.get("failed", [])
        if failed:
            msgs = "\n".join(f"  • {fid}: {err}" for fid, err in failed[:5])
            messagebox.showerror("Zahlung ändern – Fehler",
                f"{ok} OK, {len(failed)} Fehler:\n{msgs}")
        self._stop_progress(f"Zahlung geändert: {ok} OK  |  {len(failed)} Fehler")
        if ok:
            self._run()   # Daten neu laden damit Tab aktuell ist

    def _run(self, force_reload: bool = False):
        self._start_progress("⏳  Lade Abholer_DB aus OrcaScan …")
        self._hide_notification()

        def _on_error(err):
            messagebox.showerror("Tagesboten Abgleich – Fehler", str(err))
            self._stop_progress(f"❌  Fehler: {err}")

        def _worker():
            # Bei force_reload immer frisch aus OrcaScan laden (Cache überspringen)
            abholer_df = None if force_reload else \
                         (self.get_abholer_df() if self.get_abholer_df else None)
            if abholer_df is None:
                self.frame.after(0, lambda: self.status_lbl.config(
                    text="⏳  Lade Abholer_DB aus OrcaScan …"))
                try:
                    abholer_df = fetch_abholer_orca()
                except Exception as e:
                    self.frame.after(0, lambda err=e: _on_error(err))
                    return
            else:
                self.frame.after(0, lambda: self.status_lbl.config(
                    text="⏳  Abholer_DB bereits geladen – lade Tagesbote …"))

            self.frame.after(0, lambda: self.status_lbl.config(
                text="⏳  Lade Tagesbote aus Google Drive …"))

            try:
                tagesbote_df = fetch_tagesbote_gdrive()
            except Exception as e:
                self.frame.after(0, lambda err=e: _on_error(err))
                return

            self.frame.after(0, lambda: self.status_lbl.config(
                text="⏳  Vergleiche Daten …"))

            try:
                rows_fehler, rows_error, fehler_df, error_df = \
                    compute_tagesboten_abgleich(abholer_df, tagesbote_df)
                self.frame.after(0, lambda: self._apply_results(
                    rows_fehler, rows_error, fehler_df, error_df, abholer_df))
            except Exception as e:
                self.frame.after(0, lambda err=e: _on_error(err))

        threading.Thread(target=_worker, daemon=True).start()
        self._schedule_auto_refresh()

    def _schedule_auto_refresh(self):
        """Plant den nächsten Auto-Refresh in 10 Minuten."""
        if self._auto_refresh_job is not None:
            try:
                self.frame.after_cancel(self._auto_refresh_job)
            except Exception:
                pass
        self._auto_refresh_job = self.frame.after(
            self._AUTO_REFRESH_MS, self._auto_reload
        )

    def _auto_reload(self):
        """Stiller Auto-Reload alle 10 Minuten."""
        self._auto_refresh_job = None
        self._run(force_reload=False)

    def _apply_results(self, rows_fehler, rows_error, fehler_df, error_df,
                       abholer_df=None):
        self.sub_fehlerliste.set_rows(rows_fehler)
        self.sub_errorliste.set_rows(rows_error)
        self._last_fehler_df  = fehler_df
        self._last_error_df   = error_df
        self._last_abholer_df = abholer_df  # für OrcaScan-Update gespeichert

        # (row_id, barcode, name) speichern – direkt aus fehler_df lesen
        self._fehler_updates = []   # Liste von (row_id, barcode, name)
        if fehler_df is not None and "_db_id" in fehler_df.columns:
            # Abholer_DB-Index für Namens-Fallback: barcode → name
            abholer_name_map = {}
            if abholer_df is not None and "barcode" in abholer_df.columns:
                name_col = next((c for c in ("Name", "receipiantName")
                                 if c in abholer_df.columns), None)
                if name_col:
                    for _, arow in abholer_df.iterrows():
                        bc_a = str(arow.get("barcode", "")).strip()
                        nm_a = str(arow.get(name_col,  "")).strip()
                        if bc_a and nm_a and nm_a != "nan":
                            abholer_name_map[bc_a] = nm_a

            for _, row in fehler_df.iterrows():
                oid     = str(row.get("_db_id",  "")).strip()
                barcode = str(row.get("_barcode", "")).strip()
                name    = str(row.get("_name",    "")).strip()
                # Fallback: Name aus Abholer_DB wenn Tagesboten-Name leer
                if not name and barcode:
                    name = abholer_name_map.get(barcode, "")
                if oid and barcode:
                    self._fehler_updates.append((oid, barcode, name))
        n = len(rows_fehler)
        self._abholbereit_btn.config(
            text=f"✅  Alle {n} Pakete → Abholbereit setzen",
            state="normal" if n and self._fehler_updates else "disabled"
        )
        self._abholbereit_sel_btn.config(
            state="normal" if n and self._fehler_updates else "disabled"
        )

        n_err = len(rows_error)
        self._create_btn.config(
            text=f"📥  Alle {n_err} Fehler → Abholer_DB anlegen",
            state="normal" if n_err else "disabled",
        )

        self._stop_progress(
            f"✅  Abgleich abgeschlossen  |  "
            f"Fehlerliste: {len(rows_fehler)} Einträge  |  "
            f"Errorliste: {len(rows_error)} Einträge"
        )

        # Banner – Fehlerliste
        if rows_fehler:
            self._show_notification_fehler(
                f"⚠   {len(rows_fehler)} Paket(e) in Abholer_DB, aber Abholbereit_At fehlt!"
            )

        # Auto-Export + Banner wenn Errorliste nicht leer
        if rows_error:
            exported_path = self._auto_export_errorliste(error_df)
            self._show_notification(
                f"⚠   {len(rows_error)} Barcode(s) nicht in Abholer_DB!  "
                f"→  Errorliste exportiert:  {exported_path}"
            )
            self.inner_nb.select(self.sub_errorliste.frame)

    def _auto_export_errorliste(self, error_df: "pd.DataFrame") -> str:
        today_str = today_date().strftime("%d-%m-%y")
        out_path  = self.get_export_folder() / f"Errorliste_{today_str}.csv"
        try:
            error_df.to_csv(str(out_path), index=False, encoding="utf-8-sig", sep=";")
            return str(out_path)
        except Exception as e:
            messagebox.showerror("Errorliste Export – Fehler", str(e))
            return "Export fehlgeschlagen"

    def _export_fehlerliste(self):
        if self._last_fehler_df is None or self._last_fehler_df.empty:
            messagebox.showinfo("Fehlerliste", "Keine Fehlereinträge vorhanden.")
            return

        today_str = today_date().strftime("%d-%m-%y")
        out_path  = self.get_export_folder() / f"Fehlerliste_{today_str}.csv"
        try:
            self._last_fehler_df.to_csv(
                str(out_path), index=False, encoding="utf-8-sig", sep=";"
            )
            messagebox.showinfo(
                "Fehlerliste exportiert",
                f"CSV gespeichert:\n{out_path}\n\n({len(self._last_fehler_df)} Einträge)"
            )
        except Exception as e:
            messagebox.showerror("CSV Export – Fehler", str(e))

    def _set_selected_abholbereit_orca(self):
        """Setzt nur die markierten Zeilen der Fehlerliste auf Abholbereit."""
        selected = self.sub_fehlerliste.get_selected_rows()
        if not selected:
            messagebox.showinfo("Auswahl", "Keine Zeile markiert.\nBitte zuerst eine oder mehrere Zeilen anklicken.")
            return
        # Barcode → (oid, barcode, name) aus _fehler_updates mappen
        updates_map = {bc: (oid, bc, nm) for oid, bc, nm in getattr(self, "_fehler_updates", [])}
        updates = [updates_map[row[0]] for row in selected if row[0] in updates_map]
        if not updates:
            messagebox.showwarning("Fehler", "Für die Auswahl wurden keine OrcaScan-IDs gefunden.")
            return
        if not messagebox.askyesno(
            "Abholbereit setzen – Auswahl",
            f"{len(updates)} markierte(s) Paket(e) in OrcaScan auf 'Abholbereit' setzen\n"
            f"und Abholbereit_At auf die aktuelle Uhrzeit stempeln?\n\n"
            f"Diese Aktion kann nicht rückgängig gemacht werden."
        ):
            return

        from datetime import timezone as _tz
        ts = datetime.now(_tz.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        extra_fields = {
            "location":           "Abholbereit",
            "abholbereitu005fat": ts,
        }
        self._start_progress(f"Aktualisiere {len(updates)} Zeilen in OrcaScan …")
        self._abholbereit_sel_btn.config(state="disabled")

        def _worker():
            result = update_rows_orca_bulk(updates, extra_fields)
            self.frame.after(0, lambda: self._after_abholbereit(result))

        threading.Thread(target=_worker, daemon=True).start()

    def _set_abholbereit_orca(self):
        raw_updates = getattr(self, "_fehler_updates", [])   # (oid, barcode, name) Tripel
        updates = list(raw_updates)

        if not updates:
            messagebox.showinfo("Abgleich", "Keine Einträge gefunden – nichts zu aktualisieren.")
            return
        if not messagebox.askyesno(
            "Abholbereit setzen",
            f"{len(updates)} Paket(e) in OrcaScan auf 'Abholbereit' setzen\n"
            f"und Abholbereit_At auf die aktuelle Uhrzeit stempeln?\n\n"
            f"Diese Aktion kann nicht rückgängig gemacht werden."
        ):
            return

        from datetime import timezone as _tz
        ts           = datetime.now(_tz.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        # OrcaScan interne Feldnamen (nicht die Anzeigenamen!):
        # "location"          = Anzeige "Paketstatus"
        # "abholbereitu005fat" = Anzeige "Abholbereit_At"  (u005f = Underscore _ in OrcaScan-Encoding)
        extra_fields = {
            "location":           "Abholbereit",
            "abholbereitu005fat": ts,
        }

        self._start_progress(f"Aktualisiere {len(updates)} Zeilen in OrcaScan …")
        self._abholbereit_btn.config(state="disabled")

        def _worker():
            result = update_rows_orca_bulk(updates, extra_fields)
            self.frame.after(0, lambda: self._after_abholbereit(result))

        threading.Thread(target=_worker, daemon=True).start()

    def _after_abholbereit(self, result: dict):
        ok     = len(result["ok"])
        failed = result["failed"]
        msg    = f"{ok} Paket(e) erfolgreich aktualisiert."
        if failed:
            detail = "\n".join(str(f) for f in failed[:5])
            msg += f"\n\n{len(failed)} Fehler (erste 5):\n{detail}"
            messagebox.showwarning("Abgleich – teilweise Fehler", msg)
        else:
            messagebox.showinfo("Abgleich – Fertig", msg)
        self._stop_progress(f"Abholbereit gesetzt: {ok} OK  |  {len(failed)} Fehler")
        # Cache ungültig machen und Abgleich neu laden damit die aktualisierten
        # OrcaScan-Daten sofort sichtbar sind (nicht der veraltete Cache)
        if ok > 0:
            self._run(force_reload=True)

    def _create_selected_in_abholer_orca(self):
        """Legt nur die markierten Error-Zeilen als neue Einträge in der Abholer_DB an."""
        rows = self.sub_errorliste.get_selected_rows()
        if not rows:
            messagebox.showinfo("Auswahl", "Keine Zeile markiert.\nBitte zuerst eine oder mehrere Zeilen anklicken.")
            return
        if not messagebox.askyesno(
            "Abholer_DB – Auswahl anlegen",
            f"{len(rows)} markierte(n) Barcode(s) in der Abholer_DB anlegen?\n\n"
            f"Felder die übertragen werden:\n"
            f"  • Barcode, Name, Datum, Ziel-Kiosk, Zahlung\n"
            f"  • Paketstatus = Verpackt\n\n"
            f"Diese Aktion kann nicht rückgängig gemacht werden."
        ):
            return

        row_data = []
        for row in rows:
            bc, nm, dt, zk, _st, za = row
            if not bc:
                continue
            row_data.append({
                "barcode":        bc,
                "receipiantName": nm,
                "location":       "Verpackt",
                "datum":          dt,
                "zielu002dkiosk": zk,
                "zahlung":        za,
            })

        if not row_data:
            messagebox.showwarning("Fehler", "Keine gültigen Barcodes in der Auswahl.")
            return

        self._start_progress(f"Lege {len(row_data)} Einträge in Abholer_DB an …")

        def _worker():
            result = create_rows_orca_bulk(row_data)
            self.frame.after(0, lambda: self._after_create(result))

        threading.Thread(target=_worker, daemon=True).start()

    def _create_in_abholer_orca(self):
        """Legt alle Error-Zeilen als neue Einträge in der Abholer_DB an."""
        rows = list(self.sub_errorliste.rows)
        if not rows:
            messagebox.showinfo("Abholer_DB", "Keine Fehler-Einträge vorhanden.")
            return
        if not messagebox.askyesno(
            "Abholer_DB – Einträge anlegen",
            f"{len(rows)} Barcode(s) als neue Einträge in der Abholer_DB anlegen?\n\n"
            f"Felder die übertragen werden:\n"
            f"  • Barcode, Name, Datum, Ziel-Kiosk, Zahlung\n"
            f"  • Paketstatus = Verpackt\n\n"
            f"Diese Aktion kann nicht rückgängig gemacht werden."
        ):
            return

        # row = (barcode, name, datum, zielkiosk, status, zahlung)
        row_data = []
        for row in rows:
            bc, nm, dt, zk, _st, za = row
            if not bc:
                continue
            entry = {
                "barcode":          bc,
                "receipiantName":   nm,
                "location":         "Verpackt",
                "datum":            dt,
                "zielu002dkiosk":   zk,
                "zahlung":          za,
            }
            row_data.append(entry)

        if not row_data:
            messagebox.showwarning("Fehler", "Keine gültigen Barcodes gefunden.")
            return

        self._start_progress(f"Lege {len(row_data)} Einträge in Abholer_DB an …")
        self._create_btn.config(state="disabled")

        def _worker():
            result = create_rows_orca_bulk(row_data)
            self.frame.after(0, lambda: self._after_create(result))

        threading.Thread(target=_worker, daemon=True).start()

    def _after_create(self, result: dict):
        ok     = len(result["ok"])
        failed = result["failed"]
        ok_set = set(result["ok"])

        if failed:
            detail = "\n".join(str(f) for f in failed[:5])
            msg = (f"{ok} Eintrag/Einträge erfolgreich in Abholer_DB angelegt.\n\n"
                   f"{len(failed)} Fehler (erste 5):\n{detail}")
            messagebox.showwarning("Abholer_DB – teilweise Fehler", msg)
        else:
            messagebox.showinfo(
                "Abholer_DB – Fertig",
                f"✅ {ok} Eintrag/Einträge erfolgreich angelegt!\n\n"
                f"Beim nächsten Abgleich sollten diese Barcodes\n"
                f"nicht mehr in der Errorliste auftauchen."
            )

        # Erfolgreich angelegte Zeilen aus Errorliste entfernen
        if ok_set:
            self.sub_errorliste.rows = [r for r in self.sub_errorliste.rows
                                        if r[0] not in ok_set]
            self.sub_errorliste._total = len(self.sub_errorliste.rows)
            self.sub_errorliste.refresh()
            self._check_error_banner()

        self._stop_progress(f"Abholer_DB: {ok} angelegt | {len(failed)} Fehler")
        n = len(self.sub_errorliste.rows)
        self._create_btn.config(
            text=f"📥  Alle {n} Fehler → Abholer_DB anlegen" if n
                 else "📥  Fehler → Abholer_DB anlegen",
            state="normal" if n else "disabled",
        )

    def _show_notification(self, text: str):
        self.notif_lbl.config(text=text)
        self.notif_frame.pack(fill="x", before=self.inner_nb)

    def _show_notification_fehler(self, text: str):
        self.notif_lbl_fehler.config(text=text)
        self.notif_frame_fehler.pack(fill="x", before=self.inner_nb)

    def _hide_notification(self):
        self.notif_frame.pack_forget()
        self.notif_frame_fehler.pack_forget()

    def _check_error_banner(self):
        """Banner ausblenden wenn Errorliste nach Löschen leer ist."""
        if not self.sub_errorliste.rows and self.sub_errorliste._total == 0:
            self.notif_frame.pack_forget()

    def _clear(self):
        self.sub_fehlerliste.set_rows([])
        self.sub_errorliste.set_rows([])
        self._last_fehler_df = None
        self._last_error_df  = None
        self._hide_notification()
        self.status_lbl.config(
            text="Bitte zuerst Abholer_DB laden, dann 'Tagesbote laden & abgleichen' klicken."
        )


# ============================================================
# GUI – PU heute Tab
# ============================================================
class PickupHeuteTab:
    """
    Zeigt alle heutigen Pickups aus dem OrcaScan Tagesbote-Sheet
    mit aktuellem Status + Timestamps aus der Abholer_DB.
    Aktualisiert sich automatisch alle 2 Minuten.
    """

    # Spalten: Quelle Tagesbote + Abholer_DB
    _COL_HEADERS = ["Tour", "Paket-Barcode", "Name", "Packstatus", "In DB",
                    "Verpackt", "Abholbereit", "Ziel-Kiosk"]
    _COL_WIDTHS  = [48, 220, 240, 110, 60, 155, 155, 120]

    _COLOR_TOUR1 = "#ccdaf5"   # blau – Tour 1
    _COLOR_TOUR2 = "#c4dfd0"   # grün – Tour 2

    # Zeilenfarben – zeigen wo im Prozess das PU gerade steckt
    _COLOR_ABHOLBEREIT = "#bdd8c8"   # salbeigrün  – am Standort angekommen
    _COLOR_ABGEHOLT    = "#bdd0e8"   # stahlblau   – vom Kunden abgeholt
    _COLOR_RETOURE     = "#f0ccb0"   # apricot     – Retoure
    _COLOR_VERPACKT    = "#f0e4a0"   # buttergelb  – verpackt, wartet auf Kurier
    _COLOR_OFFEN       = "#f0b8b0"   # altrosa     – Packstatus offen

    PU_REFRESH_INTERVAL_MS = 2 * 60 * 1000   # 2 Minuten

    def __init__(self, parent, get_abholer_df=None, on_count_change=None,
                 get_export_folder=None, on_pu_loaded=None):
        self.get_abholer_df   = get_abholer_df
        self.get_export_folder = get_export_folder
        self._loading         = False
        self.on_count_change  = on_count_change   # Callback(n) → Kachel im Report aktualisieren
        self._on_pu_loaded    = on_pu_loaded       # Callback() → nach jedem vollständigem Load
        self._last_load_date  = None              # Datum des letzten erfolgreichen Loads
        self._refresh_job     = None              # after()-Job für Auto-Refresh
        self._all_rows        = []                # Alle geladenen Zeilen (leer bis erster Load)
        self._tb_panel_visible = False

        self.frame = tk.Frame(parent)

        # ── Kopfzeile ────────────────────────────────────────────────────
        head = tk.Frame(self.frame)
        head.pack(fill="x", pady=(4, 2))

        self.title_lbl = tk.Label(
            head, text="🚐  PU heute",
            anchor="w", font=("Segoe UI", 10, "bold")
        )
        self.title_lbl.pack(side="left")

        self.count_lbl = tk.Label(
            head, text="", anchor="e",
            font=("Segoe UI", 9), fg="#555"
        )
        self.count_lbl.pack(side="right", padx=8)

        # ── Button-Zeile ─────────────────────────────────────────────────
        btn_row = tk.Frame(self.frame)
        btn_row.pack(fill="x", pady=(0, 2))

        self.b_laden = tk.Button(
            btn_row, text="🔄  Jetzt laden", command=self._run
        )
        self.b_laden.pack(side="left")

        self.b_tagesbote = tk.Button(
            btn_row, text="📋  Tagesbote ▶",
            command=self._toggle_tagesbote_panel
        )
        self.b_tagesbote.pack(side="left", padx=(6, 0))

        self.status_lbl = tk.Label(
            btn_row, text="Noch nicht geladen.",
            anchor="w", font=("Segoe UI", 9), fg="#555"
        )
        self.status_lbl.pack(side="left", padx=10)

        # ── Tour-Abfahrt-Zeile ────────────────────────────────────────────
        tour_row = tk.Frame(self.frame)
        tour_row.pack(fill="x", pady=(0, 2))

        tk.Label(tour_row, text="Abfahrt:", font=("Segoe UI", 9)
                 ).pack(side="left", padx=(2, 6))

        self.b_t1_los = tk.Button(
            tour_row, text="🚐  Tour 1 abgefahren",
            command=self._set_t1_abfahrt,
            bg="#4a90d9", fg="white", activebackground="#3a7ac9",
            activeforeground="white", relief="flat",
            font=("Segoe UI", 9, "bold"), padx=8, pady=2, cursor="hand2"
        )
        self.b_t1_los.pack(side="left", padx=(0, 4))

        self.b_t2_los = tk.Button(
            tour_row, text="🚐  Tour 2 abgefahren",
            command=self._set_t2_abfahrt,
            bg="#4a90d9", fg="white", activebackground="#3a7ac9",
            activeforeground="white", relief="flat",
            font=("Segoe UI", 9, "bold"), padx=8, pady=2, cursor="hand2"
        )
        self.b_t2_los.pack(side="left", padx=(0, 8))

        self.tour_zeit_lbl = tk.Label(
            tour_row, text="", font=("Segoe UI", 9), fg="#555"
        )
        self.tour_zeit_lbl.pack(side="left")

        # gespeicherten Status vom heutigen Tag wiederherstellen
        self._restore_tour_buttons()

        # ── Such- und Filterzeile ─────────────────────────────────────────
        filter_row = tk.Frame(self.frame)
        filter_row.pack(fill="x", pady=(0, 4))

        tk.Label(filter_row, text="🔍", font=("Segoe UI", 11)
                 ).pack(side="left", padx=(2, 4))
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._refresh_ui())
        tk.Entry(filter_row, textvariable=self._search_var,
                 font=("Segoe UI", 10), width=24
                 ).pack(side="left", padx=(0, 14))

        tk.Label(filter_row, text="Status:", font=("Segoe UI", 9)
                 ).pack(side="left")
        self._filter_var = tk.StringVar(value="Alle")
        _status_opts = ["Alle", "Offen", "Verpackt", "Am Standort", "Abgeholt", "Tour 1", "Tour 2"]
        _cb = ttk.Combobox(filter_row, textvariable=self._filter_var,
                           values=_status_opts, state="readonly",
                           font=("Segoe UI", 9), width=18)
        _cb.pack(side="left", padx=4)
        _cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_ui())

        self._sort_col = None   # None = Standard, int = Spaltenindex
        self._sort_dir = 0      # 0 = Standard, 1 = aufsteigend, 2 = absteigend

        self._all_rows: list = []   # vollständige Basis für Suche / Filter

        # ── Hauptbereich: Tabelle + Legende nebeneinander ────────────────
        body = tk.Frame(self.frame)
        body.pack(fill="both", expand=True)
        self._body = body   # Referenz für Toggle-Panel

        # ── Tagesboten-Seitenpanel (rechts, zunächst versteckt) ──────────
        self._tb_panel = tk.Frame(body, bd=1, relief="groove", width=430)
        self._tb_panel.pack_propagate(False)
        self._tb_abgleich = TagesbotenAbgleichTab(
            self._tb_panel,
            get_abholer_df    = self.get_abholer_df,
            get_export_folder = self.get_export_folder,
        )
        self._tb_abgleich.frame.pack(fill="both", expand=True)
        # Panel bleibt zunächst versteckt (kein pack-Aufruf)

        # ── Legende (rechts, feste Breite) ───────────────────────────────
        legend = tk.Frame(body, bg="#f5f5f5", bd=1, relief="groove", width=170)
        legend.pack(side="right", fill="y", padx=(4, 6), pady=(0, 6))
        legend.pack_propagate(False)

        tk.Label(legend, text="Legende", font=("Segoe UI", 9, "bold"),
                 bg="#f5f5f5", anchor="w").pack(fill="x", padx=8, pady=(8, 2))

        tk.Label(legend, text="Touren (Spalte)", font=("Segoe UI", 8, "bold"),
                 bg="#f5f5f5", fg="#555", anchor="w").pack(fill="x", padx=8, pady=(4, 1))
        for color, text in [
            (self._COLOR_TOUR1, "T1 – Tour 1  (bis 11:15)"),
            (self._COLOR_TOUR2, "T2 – Tour 2  (ab 11:15)"),
        ]:
            row_l = tk.Frame(legend, bg="#f5f5f5")
            row_l.pack(fill="x", padx=8, pady=1)
            tk.Frame(row_l, bg=color, width=16, height=16, relief="solid", bd=1
                     ).pack(side="left")
            tk.Label(row_l, text=f"  {text}", font=("Segoe UI", 8),
                     bg="#f5f5f5", anchor="w").pack(side="left", fill="x")

        tk.Frame(legend, bg="#ccc", height=1).pack(fill="x", padx=8, pady=(6, 2))
        tk.Label(legend, text="Paketstatus (Zeile)", font=("Segoe UI", 8, "bold"),
                 bg="#f5f5f5", fg="#555", anchor="w").pack(fill="x", padx=8, pady=(2, 1))

        for color, text in [
            (self._COLOR_OFFEN,       "Offen – noch nicht verpackt"),
            (self._COLOR_VERPACKT,    "Verpackt – wartet auf Kurier"),
            (self._COLOR_ABHOLBEREIT, "Am Standort angekommen"),
            (self._COLOR_ABGEHOLT,    "Vom Kunden abgeholt"),
            (self._COLOR_RETOURE,     "Retoure"),
        ]:
            row_l = tk.Frame(legend, bg="#f5f5f5")
            row_l.pack(fill="x", padx=8, pady=2)
            tk.Frame(row_l, bg=color, width=16, height=16, relief="solid", bd=1
                     ).pack(side="left")
            tk.Label(row_l, text=f"  {text}", font=("Segoe UI", 8),
                     bg="#f5f5f5", anchor="w", wraplength=130, justify="left"
                     ).pack(side="left", fill="x")

        # ── Tabelle ──────────────────────────────────────────────────────
        try:
            import tksheet as tks
            self._sheet = tks.Sheet(
                body,
                headers=self._COL_HEADERS,
                theme="light",
                show_row_index=True,
                row_index_width=36,
                show_x_scrollbar=True,
                show_y_scrollbar=True,
            )
            self._sheet.enable_bindings("single_select", "row_select", "column_select", "copy", "column_width_resize")
            self._sheet.pack(fill="both", expand=True, padx=(4, 0), pady=(0, 6))


            for i, w in enumerate(self._COL_WIDTHS):
                self._sheet.column_width(column=i, width=w)

            try:
                self._sheet.extra_bindings([
                    ("cell_select", self._on_cell_click),
                    ("column_select", self._on_header_click),
                ])
            except Exception:
                pass
            self._sheet.bind("<Button-3>", self._on_right_click)
        except ImportError:
            self._sheet = None
            tk.Label(self.frame, text="tksheet nicht installiert").pack()

        self._last_data: list = []        # zuletzt gerenderte Zeilen für Barcode-Kopieren
        self._displayed_rows: list = []   # gefilterte/sortierte Zeilen die aktuell sichtbar sind

        # Automatischer erster Load 2 Sekunden nach Programmstart
        self.frame.after(2000, self._run)

    def _on_cell_click(self, event):
        """Barcode der angeklickten Zeile in die Zwischenablage kopieren."""
        try:
            row = event.row
            if 0 <= row < len(self._last_data) and self._last_data[row]:
                barcode = str(self._last_data[row][0])
                if barcode:
                    self.frame.clipboard_clear()
                    self.frame.clipboard_append(barcode)
                    orig = self.count_lbl.cget("text")
                    self.count_lbl.config(text=f"📋 Kopiert: {barcode[:28]}", fg="#27ae60")
                    self.frame.after(1800, lambda: self.count_lbl.config(
                        text=orig, fg="#555"))
        except Exception:
            pass

    def _on_right_click(self, event):
        """Rechtsklick auf eine Zeile → Kontextmenü für Packstatus-Korrektur."""
        if not self._sheet:
            return
        # Zeilenindex aus y-Koordinate ermitteln
        try:
            row_idx = self._sheet.identify_row(event, allow_end=False)
        except Exception:
            try:
                sel = list(self._sheet.get_selected_rows())
                row_idx = sel[0] if sel else None
            except Exception:
                row_idx = None
        if row_idx is None:
            return
        disp = self._displayed_rows
        if row_idx >= len(disp):
            return
        r = disp[row_idx]

        menu = tk.Menu(self.frame, tearoff=0)

        # ── Packstatus (Tagesboten-Sheet) ──────────────────────────────
        if r.get("_tb_row_id"):
            already_vp = r["tb_status"].lower() == "verpackt"
            menu.add_command(
                label="✓  Verpackt setzen",
                state="disabled" if already_vp else "normal",
                command=lambda: self._set_kontrollstatus(r, "Verpackt")
            )
            menu.add_command(
                label="○  Offen setzen",
                state="disabled" if not already_vp else "normal",
                command=lambda: self._set_kontrollstatus(r, "Offen")
            )

        # ── Abholbereit + Löschen (Abholer_DB) ─────────────────────────
        if r.get("_db_id"):
            if r.get("_tb_row_id"):
                menu.add_separator()
            already_ab = r["_abholbereit_bool"]
            menu.add_command(
                label="📦  Abholbereit setzen",
                state="disabled" if already_ab else "normal",
                command=lambda: self._set_abholbereit_single(r)
            )
            menu.add_separator()
            menu.add_command(
                label="🗑️  Aus Abholer_DB löschen",
                command=lambda: self._delete_from_db(r)
            )

        if menu.index("end") is None:
            return   # leeres Menü → nicht anzeigen

        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _set_kontrollstatus(self, r, new_status):
        """Setzt den Kontrollstatus einer Zeile in OrcaScan und aktualisiert die Anzeige."""
        import threading as _thr
        import json as _json
        import urllib.request as _urllib
        import urllib.error   as _urllib_err
        self.status_lbl.config(
            text=f"⏳  Setze Packstatus auf '{new_status}' …", fg="#e67e22")

        def _worker():
            # Tagesboten-Sheet nutzt interne Kleinschreibung:
            #   "name" statt "receipiantName", "kontrollstatus" statt "Kontrollstatus"
            payload = {
                "barcode":         r["barcode"],
                "name":            r["name"],
                "kontrollstatus":  new_status,
            }
            body = _json.dumps(payload).encode("utf-8")
            url  = (f"{ORCA_BASE_URL}/sheets/{ORCA_TAGESBOTE_SHEET_ID}"
                    f"/rows/{r['_tb_row_id']}?partial=true")
            req  = _urllib.Request(url, data=body, method="PUT",
                                   headers={"Authorization": f"Bearer {ORCA_API_KEY}",
                                            "Content-Type": "application/json"})
            err = None
            try:
                with _urllib.urlopen(req, timeout=15) as resp:
                    resp.read()
            except _urllib_err.HTTPError as e:
                err = f"HTTP {e.code}"
                try:
                    err += ": " + e.read().decode("utf-8", errors="replace")[:200]
                except Exception:
                    pass
            except Exception as ex:
                err = str(ex)
            self.frame.after(0, lambda: _done(err))

        def _done(err):
            if err:
                self.status_lbl.config(
                    text=f"❌  Fehler beim Setzen: {err[:80]}", fg="#c0392b")
            else:
                r["tb_status"] = new_status   # lokal sofort aktualisieren
                self._refresh_ui()
                self.status_lbl.config(
                    text=f"✓  Packstatus '{new_status}' gesetzt – {r['barcode']}", fg="#27ae60")

        _thr.Thread(target=_worker, daemon=True).start()

    def _toggle_tagesbote_panel(self):
        """Tagesboten-Seitenpanel ein-/ausklappen."""
        if self._tb_panel_visible:
            self._tb_panel.pack_forget()
            self._tb_panel_visible = False
            self.b_tagesbote.config(text="📋  Tagesbote ▶")
        else:
            # Sheet kurz entfernen, damit der Pack-Manager Platz neu verteilt.
            # (Das Sheet hat expand=True und hat sonst bereits den ganzen Platz
            #  beansprucht, bevor das Panel eingeblendet wird.)
            if self._sheet:
                self._sheet.pack_forget()
            self._tb_panel.pack(in_=self._body, side="right",
                                fill="y", padx=(4, 0), pady=(0, 6))
            if self._sheet:
                self._sheet.pack(fill="both", expand=True,
                                 padx=(4, 0), pady=(0, 6))
            self._tb_panel_visible = True
            self.b_tagesbote.config(text="◀  Tagesbote")

    def _set_abholbereit_single(self, r):
        """Einzelne Zeile per Rechtsklick auf Abholbereit setzen."""
        import threading as _thr
        from datetime import datetime, timezone as _tz

        self.status_lbl.config(
            text=f"⏳  Setze '{r['barcode']}' auf Abholbereit …", fg="#e67e22")

        def _worker():
            ts     = datetime.now(_tz.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            result = update_rows_orca_bulk(
                [(r["_db_id"], r["barcode"], r["name"])],
                {"location": "Abholbereit", "abholbereitu005fat": ts}
            )
            self.frame.after(0, lambda: _done(result, ts))

        def _done(result, ts):
            if result["failed"]:
                self.status_lbl.config(
                    text=f"❌  Fehler: {result['failed'][0][:60]}", fg="#c0392b")
            else:
                # Lokal sofort aktualisieren
                for row in self._all_rows:
                    if row["barcode"] == r["barcode"]:
                        row["_abholbereit_bool"] = True
                        row["abholbereit_at"]    = ts[:10]
                        row["db_status"]         = "abholbereit"
                self._refresh_ui()
                self.status_lbl.config(
                    text=f"✓  '{r['barcode']}' auf Abholbereit gesetzt", fg="#27ae60")

        _thr.Thread(target=_worker, daemon=True).start()

    def _delete_from_db(self, r):
        """Zeile per Rechtsklick aus Abholer_DB und ggf. Tagesboten-Sheet löschen."""
        import threading as _thr
        from tkinter import messagebox

        has_db = bool(r.get("_db_id"))
        has_tb = bool(r.get("_tb_row_id"))

        sheets = []
        if has_db:
            sheets.append("Abholer_DB")
        if has_tb:
            sheets.append("Tagesboten-Sheet")
        sheets_txt = " und ".join(sheets)

        if not messagebox.askyesno(
            "Löschen bestätigen",
            f"Paket '{r['barcode']}' ({r['name']}) wirklich aus {sheets_txt} löschen?\n\n"
            "Diese Aktion kann nicht rückgängig gemacht werden."
        ):
            return

        self.status_lbl.config(
            text=f"⏳  Lösche '{r['barcode']}' …", fg="#e67e22")

        def _worker():
            errors = []
            if has_db:
                res = delete_rows_orca_bulk([r["_db_id"]], sheet_id=ORCA_ABHOLER_SHEET_ID)
                if res["failed"]:
                    errors.append(f"Abholer_DB: {res['errors'][0][:50] if res['errors'] else 'Fehler'}")
            if has_tb:
                res = delete_rows_orca_bulk([r["_tb_row_id"]], sheet_id=ORCA_TAGESBOTE_SHEET_ID)
                if res["failed"]:
                    errors.append(f"Tagesbote: {res['errors'][0][:50] if res['errors'] else 'Fehler'}")
            self.frame.after(0, lambda: _done(errors))

        def _done(errors):
            if errors:
                self.status_lbl.config(
                    text=f"❌  Fehler: {' | '.join(errors)}", fg="#c0392b")
            else:
                self._all_rows = [row for row in self._all_rows
                                  if row.get("_db_id") != r.get("_db_id")
                                  and row.get("barcode") != r["barcode"]]
                self._refresh_ui()
                self.status_lbl.config(
                    text=f"✓  '{r['barcode']}' aus {sheets_txt} gelöscht", fg="#27ae60")

        _thr.Thread(target=_worker, daemon=True).start()

    # ── Tour-Abfahrt Steuerung ────────────────────────────────────────────
    def _lese_netz_tour_zeiten(self):
        """Liest tour_zeiten JSON direkt vom Netzlaufwerk (synchron). Gibt {} zurueck wenn nicht gefunden."""
        import datetime as _dt, json as _json
        heute = (_dt.datetime.now()).strftime("%Y-%m-%d")
        pfad = TOURLISTEN_DIR / f"tour_zeiten_{heute}.json"
        try:
            if pfad.exists():
                return _json.loads(pfad.read_text(encoding="utf-8"))
        except Exception:
            pass
        return {}

    def _set_t1_abfahrt(self):
        import datetime as _dt
        # Zuerst Netzlaufwerk pruefen – wer zuerst klickt gewinnt
        netz = self._lese_netz_tour_zeiten()
        if netz.get("t1"):
            # anderer PC war schneller – dessen Daten uebernehmen
            _save_tour_zeiten(netz.get("t1"), netz.get("t2"),
                              netz.get("t1_barcodes", []), netz.get("t2_barcodes", []))
            self._restore_tour_buttons()
            self._recompute_tours_local()
            return
        tz = _load_tour_zeiten()
        if tz.get("t1"):
            return
        jetzt = (_dt.datetime.now()).strftime("%H:%M")
        t1_barcodes = [r["barcode"] for r in self._all_rows if r["tb_status"] == "Verpackt"]
        _save_tour_zeiten(t1=jetzt, t2=tz.get("t2"), t1_barcodes=t1_barcodes)
        self._restore_tour_buttons()
        self._recompute_tours_local()
        self._upload_tourliste("T1")
        self._upload_tour_zeiten_to_drive()

    def _set_t2_abfahrt(self):
        import datetime as _dt
        # Zuerst Netzlaufwerk pruefen – wer zuerst klickt gewinnt
        netz = self._lese_netz_tour_zeiten()
        if netz.get("t2"):
            # anderer PC war schneller – dessen Daten uebernehmen
            _save_tour_zeiten(netz.get("t1"), netz.get("t2"),
                              netz.get("t1_barcodes", []), netz.get("t2_barcodes", []))
            self._restore_tour_buttons()
            self._recompute_tours_local()
            return
        tz = _load_tour_zeiten()
        if tz.get("t2"):
            return
        jetzt = (_dt.datetime.now()).strftime("%H:%M")
        # t1_barcodes vom Netzlaufwerk verwenden falls verfuegbar (konsistente Basis)
        t1_bc = set(netz.get("t1_barcodes") or tz.get("t1_barcodes") or [])
        t2_barcodes = [r["barcode"] for r in self._all_rows if r["barcode"] not in t1_bc]
        _save_tour_zeiten(t1=tz.get("t1"), t2=jetzt, t2_barcodes=t2_barcodes)
        self._restore_tour_buttons()
        self._recompute_tours_local()
        self._upload_tourliste("T2")
        self._upload_tour_zeiten_to_drive()

    def _upload_tourliste(self, tour: str):
        import datetime as _dt, pandas as _pd
        rows_all = getattr(self, "_all_rows", [])
        tour_rows = [r for r in rows_all if r.get("tour") == tour]
        if not tour_rows:
            return
        def _last4(v):
            s = str(v).strip()
            return s[-4:] if len(s) >= 4 else s
        df = _pd.DataFrame([{
            "Paket-Barcode":     r["barcode"],
            "Bestellnummer":     _last4(r["barcode"]),
            "Datum":             r["scan_datum"],
            "Vorname":           "",
            "Name":              r["name"],
            "Ziel-Kiosk":        r["zielkiosk"],
            "Status":            r["tb_status"],
            "Bestellwert":       "",
            "Versicher.":        "",
            "Email":             "",
            "Lieferung-Adresse": "",
            "Lieferung":         "",
            "Notizen":           "",
            "Kontrollstatus":    r["tb_status"],
            "Zahlung":           "",
            "Rezept":            "",
        } for r in tour_rows])
        heute = (_dt.datetime.now()).strftime("%d.%m.%Y")
        tour_suffix = "A" if tour == "T1" else "B"
        filename = f"Orca_Abholer_{heute}{tour_suffix}.xlsx"

        def worker():
            try:
                TOURLISTEN_DIR.mkdir(parents=True, exist_ok=True)
                pfad = TOURLISTEN_DIR / filename
                from openpyxl import Workbook as _WB
                from openpyxl.styles import Font as _Fnt, PatternFill as _Fill, Alignment as _Aln, Border as _Brd, Side as _Side
                from openpyxl.utils import get_column_letter as _gcl
                import math as _math
                wb = _WB()
                ws = wb.active
                cols = list(df.columns)
                _thin = _Side(style='thin', color='BFBFBF')
                _border = _Brd(left=_thin, right=_thin, top=_thin, bottom=_thin)
                _col_w = [22, 14, 16, 10, 28, 13, 16, 13, 13, 28, 28, 12, 16, 15, 12, 10]
                for ci, cn in enumerate(cols, 1):
                    c = ws.cell(row=1, column=ci, value=cn)
                    c.fill = _Fill('solid', start_color='4F81BD', end_color='4F81BD')
                    c.font = _Fnt(name='Arial', bold=True, color='FFFFFF', size=10)
                    c.alignment = _Aln(horizontal='center', vertical='center', wrap_text=True)
                    c.border = _border
                    ws.column_dimensions[_gcl(ci)].width = _col_w[ci - 1] if ci <= len(_col_w) else 15
                ws.row_dimensions[1].height = 30
                for ri, (_, row) in enumerate(df.iterrows(), 2):
                    for ci, cn in enumerate(cols, 1):
                        val = row[cn]
                        if isinstance(val, float) and _math.isnan(val):
                            val = None
                        c = ws.cell(row=ri, column=ci, value=val)
                        c.font = _Fnt(name='Arial', size=10)
                        c.alignment = _Aln(vertical='center')
                        c.border = _border
                ws.freeze_panes = 'A2'
                wb.save(pfad)
                self.frame.after(0, lambda: self.status_lbl.config(text=f"✅ Tourliste {tour} gespeichert: {pfad}"))
            except Exception as e:
                self.frame.after(0, lambda err=str(e): self.status_lbl.config(text=f"⚠ Tourliste-Speichern fehlgeschlagen: {err}"))
        threading.Thread(target=worker, daemon=True).start()

    def _restore_tour_buttons(self):
        """Liest gespeicherte Abfahrtszeiten und aktualisiert Button-Farbe + Info-Label."""
        tz = _load_tour_zeiten()
        t1, t2 = tz.get("t1"), tz.get("t2")

        # Tour-1-Button
        if t1:
            self.b_t1_los.config(text=f"🚐  Tour 1  \u2705 {t1}", bg="#4ea874",
                                  activebackground="#3d8f63")
        else:
            self.b_t1_los.config(text="🚐  Tour 1 abgefahren", bg="#4a90d9",
                                  activebackground="#3a7ac9")

        # Tour-2-Button
        if t2:
            self.b_t2_los.config(text=f"🚐  Tour 2  \u2705 {t2}", bg="#4ea874",
                                  activebackground="#3d8f63")
        else:
            self.b_t2_los.config(text="🚐  Tour 2 abgefahren", bg="#4a90d9",
                                  activebackground="#3a7ac9")

        # Info-Label
        lbl_parts = []
        if t1:
            lbl_parts.append(f"T1 ab {t1}")
        if t2:
            lbl_parts.append(f"T2 ab {t2}")
        self.tour_zeit_lbl.config(
            text="  \u00b7  ".join(lbl_parts) if lbl_parts else "Noch keine Abfahrt markiert",
            fg="#27ae60" if (t1 or t2) else "#888"
        )

    # \u2500\u2500 Tour-Zeiten Drive-Sync \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    def _upload_tour_zeiten_to_drive(self):
        """L\u00e4dt die heutigen tour_zeiten als JSON nach Google Drive hoch (Hintergrund-Thread).
        So sehen andere Bombadil-Instanzen sofort welche Tour abgefahren ist."""
        import datetime as _dt, threading as _th, json as _json
        heute    = (_dt.datetime.now()).strftime("%Y-%m-%d")
        filename = f"tour_zeiten_{heute}.json"
        tz       = _load_tour_zeiten()
        def worker():
            try:
                TOURLISTEN_DIR.mkdir(parents=True, exist_ok=True)
                pfad = TOURLISTEN_DIR / filename
                pfad.write_text(_json.dumps(tz, ensure_ascii=False), encoding="utf-8")
            except Exception as e:
                self.frame.after(0, lambda err=str(e): self.status_lbl.config(text=f"\u26a0 Tour-Sync fehlgeschlagen: {err}"))
        _th.Thread(target=worker, daemon=True).start()

    def _sync_tour_zeiten_from_drive(self):
        """Liest tour_zeiten-JSON vom Netzlaufwerk. Remote-Daten gewinnen immer."""
        import datetime as _dt, threading as _th, json as _json
        heute    = (_dt.datetime.now()).strftime("%Y-%m-%d")
        filename = f"tour_zeiten_{heute}.json"
        def worker():
            try:
                pfad = TOURLISTEN_DIR / filename
                if not pfad.exists():
                    return
                remote = _json.loads(pfad.read_text(encoding="utf-8"))
                if not remote.get("t1") and not remote.get("t2"):
                    return
                # Remote-Daten immer anwenden (ueberschreiben lokale)
                _save_tour_zeiten(
                    remote.get("t1"), remote.get("t2"),
                    remote.get("t1_barcodes", []), remote.get("t2_barcodes", []),
                )
                self.frame.after(0, self._restore_tour_buttons)
                self.frame.after(0, self._recompute_tours_local)
                self.frame.after(0, lambda t1=remote.get("t1"), t2=remote.get("t2"): self.status_lbl.config(
                    text=f"\u2705 Tour-Sync: T1={t1 or '-'}  T2={t2 or '-'}"))
            except Exception as e:
                self.frame.after(0, lambda err=str(e): self.status_lbl.config(text=f"\u26a0 Sync-Fehler: {err}"))
        _th.Thread(target=worker, daemon=True).start()

    def _recompute_tours_local(self):
        """Weist Touren per gespeicherter Barcode-Liste zu und aktualisiert die UI."""
        if not self._all_rows:
            return
        tz    = _load_tour_zeiten()
        t1_bc = set(tz.get("t1_barcodes") or [])
        t2_bc = set(tz.get("t2_barcodes") or [])
        for r in self._all_rows:
            bc = r["barcode"]
            if bc in t1_bc:
                r["tour"] = "T1"
            elif bc in t2_bc:
                r["tour"] = "T2"
            else:
                r["tour"] = ""
        self._refresh_ui()

    # ── Daten laden ──────────────────────────────────────────────────────
    def _run(self):
        if self._loading:
            return
        self._loading = True
        self.b_laden.config(state="disabled")
        self.status_lbl.config(text="⏳  Lade Tagesbote aus OrcaScan …", fg="#555")
        import threading
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            abholer_df = self.get_abholer_df() if self.get_abholer_df else None
            if abholer_df is None or abholer_df.empty:
                self.frame.after(0, lambda: self.status_lbl.config(
                    text="⏳  Lade Abholer_DB aus OrcaScan …"))
                abholer_df = fetch_abholer_orca()

            self.frame.after(0, lambda: self.status_lbl.config(
                text="⏳  Lade Tagesbote-Sheet …"))
            tagesbote_df = fetch_sheet_orca(ORCA_TAGESBOTE_SHEET_ID)

            self.frame.after(0, lambda: self.status_lbl.config(
                text="⏳  Verarbeite …"))
            _tz = _load_tour_zeiten()
            # Erst ohne Cutoff laden um Roh-Timestamps zu bekommen
            rows, diag = compute_pickup_heute(abholer_df, tagesbote_df, t2_cutoff=None)

            # Tour-Zuweisung nur per gespeicherter Barcode-Liste (kein Scan-Zeit-Vergleich)
            _t1_bc = set(_tz.get("t1_barcodes") or [])
            _t2_bc = set(_tz.get("t2_barcodes") or [])
            for _r in rows:
                _bc = _r["barcode"]
                if _bc in _t1_bc:
                    _r["tour"] = "T1"
                elif _bc in _t2_bc:
                    _r["tour"] = "T2"

            self.frame.after(0, lambda r=rows, d=diag: self._apply(r, d))
        except Exception as e:
            self.frame.after(0, lambda err=e: self._on_error(err))

    def _apply(self, rows, diag=None):
        import datetime as _dt
        self._loading = False
        self.b_laden.config(state="normal")

        heute_str = (_dt.datetime.now()).strftime("%d.%m.%Y")
        self.title_lbl.config(text=f"🚐  PU heute – {heute_str}")

        if not rows:
            import datetime as _dt2
            today_local = (_dt2.datetime.now()).date()
            same_day = (self._last_load_date == today_local)
            had_data = bool(getattr(self, "_last_rows", None)) and same_day
            if had_data:
                # API-Blip am gleichen Tag – alten Zähler behalten, nur Status-Hinweis
                self.status_lbl.config(
                    text=f"⚠  Tagesbote-Sheet leer – zeige letzten Stand ({heute_str})", fg="#e67e22")
            else:
                # Neuer Arbeitstag oder erster Load – Zähler zurücksetzen
                self.status_lbl.config(
                    text=f"Keine Pickups für heute ({heute_str}) gefunden.", fg="#555")
                self.count_lbl.config(text="")
                if self._sheet:
                    self._sheet.set_sheet_data([])
                if self.on_count_change:
                    self.on_count_change(0)
            return

        n_gesamt       = len(rows)
        if self.on_count_change:
            self.on_count_change(n_gesamt)
        n_abholbereit  = sum(1 for r in rows if r["_abholbereit_bool"]
                             or r.get("db_status") == "abgeholt")
        n_verpackt     = sum(1 for r in rows if r["tb_status"].lower() == "verpackt"
                             and not r["_abholbereit_bool"]
                             and r.get("db_status") != "abgeholt")
        n_offen        = sum(1 for r in rows if r["tb_status"].lower() == "offen")
        n_t1           = sum(1 for r in rows if r.get("tour") == "T1")
        n_t2           = sum(1 for r in rows if r.get("tour") == "T2")
        _parts = [f"🚐 Tour 1: {n_t1}  Tour 2: {n_t2}",
                  f"{n_gesamt} PUs", f"{n_abholbereit} angekommen",
                  f"{n_verpackt} verpackt", f"{n_offen} offen"]
        self.count_lbl.config(text="  ·  ".join(_parts))
        _uhrzeit = (_dt.datetime.now()).strftime("%H:%M")
        _status_text = f"✅  Zuletzt geladen: {_uhrzeit} Uhr"
        _status_fg   = "#28a745"

        # Diagnose: Doppelte / leere Barcodes anzeigen
        if diag:
            _warn_parts = []
            if diag.get("n_empty_bc", 0):
                _warn_parts.append(f"{diag['n_empty_bc']} leere(r) Barcode(s) übersprungen")
            if diag.get("n_dedup_drop", 0):
                _dups_str = ", ".join(diag.get("dup_barcodes", []))
                _warn_parts.append(
                    f"{diag['n_dedup_drop']} Duplikat(e) entfernt"
                    + (f": {_dups_str}" if _dups_str else "")
                )
            if _warn_parts:
                _status_text = f"⚠  {' · '.join(_warn_parts)}  (Sheet: {diag['n_raw']}, angezeigt: {n_gesamt})"
                _status_fg   = "#e67e22"

        self.status_lbl.config(text=_status_text, fg=_status_fg)
        self._restore_tour_buttons()   # Button-Status nach Laden aktualisieren
        self._sync_tour_zeiten_from_drive()  # Kollegen-Sync: Tour-Status von Drive holen

        if self._sheet:
            self._all_rows  = rows   # vollständige Basis für Suche / Filter
            self._last_rows = rows   # für Report-Kachel-Detail (Panel D)
            import datetime as _dt2
            self._last_load_date = (_dt2.datetime.now()).date()
            self._refresh_ui()
            if self._on_pu_loaded:
                self.frame.after(0, self._on_pu_loaded)
        self._schedule_refresh()

    def _refresh_ui(self):
        """Wendet Suchtext und Statusfilter auf _all_rows an und aktualisiert die Tabelle."""
        if not self._sheet:
            return
        rows = self._all_rows

        # ── Suchfilter ───────────────────────────────────────────────────
        q = self._search_var.get().strip().lower()
        if q:
            rows = [r for r in rows
                    if q in str(r["barcode"]).lower() or q in str(r["name"]).lower()]

        # ── Statusfilter ─────────────────────────────────────────────────
        f = self._filter_var.get()
        if f == "Am Standort":
            rows = [r for r in rows if r["_abholbereit_bool"]]
        elif f == "Abgeholt":
            rows = [r for r in rows if r.get("db_status") == "abgeholt"]
        elif f == "Tour 1":
            rows = [r for r in rows if r.get("tour") == "T1"]
        elif f == "Tour 2":
            rows = [r for r in rows if r.get("tour") == "T2"]
        elif f == "Verpackt":
            rows = [r for r in rows if r["tb_status"].lower() == "verpackt"
                    and not r["_abholbereit_bool"]]
        elif f == "Offen":
            rows = [r for r in rows if r["tb_status"].lower() == "offen"]

        # ── Sortierung ───────────────────────────────────────────────────
        _none_last = "ÿ"   # leere Strings nach hinten sortieren
        _col_keys = {
            0: lambda r: str(r.get("tour", "") or "").lower(),
            1: lambda r: str(r["barcode"] or "").lower(),
            2: lambda r: str(r["name"] or "").lower() or _none_last,
            3: lambda r: str(r["tb_status"] or "").lower(),
            5: lambda r: str(r["verpackt_at"] or _none_last),
            6: lambda r: str(r["abholbereit_at"] or _none_last),
            7: lambda r: str(r["zielkiosk"] or "").lower() or _none_last,
        }
        if self._sort_col is not None and self._sort_dir > 0 and self._sort_col in _col_keys:
            rows = sorted(rows, key=_col_keys[self._sort_col], reverse=(self._sort_dir == 2))

        # Standard: erst nach Prozessschritt (Offen → Verpackt → …), dann Tour
        def _tour_order(r):
            t = r.get("tour", "")
            return 0 if t == "T1" else (1 if t == "T2" else 2)

        def _stage(r):
            ds = r.get("db_status", "")
            if ds == "abgeholt":                        return 3
            if ds == "retoure":                         return 4
            if r["_abholbereit_bool"]:                  return 2
            if r["tb_status"].lower() == "verpackt":    return 1
            return 0  # offen / nicht verpackt → ganz oben

        if self._sort_col is None or self._sort_dir == 0:
            rows = sorted(rows, key=lambda r: (_stage(r), _tour_order(r)))

        self._displayed_rows = rows   # für Rechtsklick-Menü merken

        data = [
            [r.get("tour", ""), r["barcode"], r["name"], r["tb_status"], r["in_db"],
             r["verpackt_at"], r["abholbereit_at"], r["zielkiosk"]]
            for r in rows
        ]
        self._last_data = data
        self._sheet.set_sheet_data(data)

        # Kopfzeilen-Zaehler aktualisieren (Tour 1/2 immer aus _all_rows)
        if self._all_rows:
            _ar = self._all_rows
            _n_ges = len(_ar)
            _n_ab  = sum(1 for _r in _ar if _r["_abholbereit_bool"] or _r.get("db_status") == "abgeholt")
            _n_vp  = sum(1 for _r in _ar if _r["tb_status"].lower() == "verpackt"
                         and not _r["_abholbereit_bool"] and _r.get("db_status") != "abgeholt")
            _n_of  = sum(1 for _r in _ar if _r["tb_status"].lower() == "offen")
            _n_t1  = sum(1 for _r in _ar if _r.get("tour") == "T1")
            _n_t2  = sum(1 for _r in _ar if _r.get("tour") == "T2")
            self.count_lbl.config(text="  ·  ".join([
                f"🚐 Tour 1: {_n_t1}  Tour 2: {_n_t2}",
                f"{_n_ges} PUs", f"{_n_ab} angekommen",
                f"{_n_vp} verpackt", f"{_n_of} offen",
            ]))

        try:
            self._sheet.dehighlight_all(redraw=False)
        except Exception:
            pass

        for i, r in enumerate(rows):
            # Zeilenfarbe nach Prozessstatus
            if r["_abholbereit_bool"]:
                bg = self._COLOR_ABHOLBEREIT
            elif r.get("db_status") == "abgeholt":
                bg = self._COLOR_ABGEHOLT
            elif r.get("db_status") == "retoure":
                bg = self._COLOR_RETOURE
            elif r["tb_status"].lower() == "verpackt":
                bg = self._COLOR_VERPACKT
            else:
                bg = self._COLOR_OFFEN
            self._sheet.highlight_rows(rows=[i], bg=bg, redraw=False)

            # Tour-Spalte (Spalte 0) mit eigener Farbe hervorheben
            t = r.get("tour", "")
            tc = self._COLOR_TOUR1 if t == "T1" else (self._COLOR_TOUR2 if t == "T2" else None)
            if tc:
                try:
                    self._sheet.highlight_cells(row=i, column=0, bg=tc, redraw=False)
                except Exception:
                    pass

        self._sheet.refresh()

    def _on_error(self, err):
        self._loading = False
        self.b_laden.config(state="normal")
        self.status_lbl.config(text=f"❌  Fehler: {err}", fg="#dc3545")
        self._schedule_refresh()

    def _schedule_refresh(self):
        """Plant den nächsten automatischen Reload in 2 Minuten."""
        if self._refresh_job is not None:
            try:
                self.frame.after_cancel(self._refresh_job)
            except Exception:
                pass
        self._refresh_job = self.frame.after(self.PU_REFRESH_INTERVAL_MS, self._run)

    def _export_xlsx(self, tour: str):
        """Exportiert Am-Standort-Pakete der angegebenen Tour + Fehlerliste als Excel."""
        import datetime as _dt
        from tkinter import filedialog, messagebox
        import openpyxl
        from openpyxl.styles import Font

        rows_all = getattr(self, "_all_rows", [])
        if not rows_all:
            messagebox.showinfo("Export", "Keine Daten geladen – bitte zuerst laden.")
            return

        def _is_fehler(r):
            return (r.get("tour") in ("T1", "T2")
                    and not r["_abholbereit_bool"]
                    and r.get("db_status") not in ("abgeholt", "retoure"))

        angekommen = [r for r in rows_all if r.get("tour") == tour and r["_abholbereit_bool"]]
        fehler     = [r for r in rows_all if r.get("tour") == tour and _is_fehler(r)]

        if not angekommen and not fehler:
            messagebox.showinfo("Export", f"Keine Pakete für {tour} vorhanden.")
            return

        heute = (_dt.datetime.now()).strftime("%Y-%m-%d")
        tour_nr = tour.replace("T", "")
        default_name = f"PU_Tour{tour_nr}_{heute}.xlsx"

        pfad = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel-Datei", "*.xlsx")],
            initialfile=default_name,
            title=f"Tour {tour_nr} exportieren",
        )
        if not pfad:
            return

        # Alle 27 Spalten exakt wie Abholer_DB
        HEADERS = [
            "Paket-Barcode", "Reservierungsnr.", "Best.-Nr.", "Bestellnummer",
            "Datum", "Name", "Ankunftsort", "Ziel-Kiosk", "Lieferung-Name",
            "Standort", "Paketstatus", "Lieferung-Adresse", "Abgeholt_At",
            "Abholbereit_At", "Lieferung-Zusatz", "Verpackt_At", "Ziel Kiosk",
            "Unterschrift", "Scan-Datum", "Status", "Bestellwert", "Versicher.",
            "Lieferung", "Zahlung", "Rezept", "Notizen", "Paketfoto",
        ]
        HDR_FONT     = Font(bold=True)
        COL_WIDTHS   = [30, 18, 12, 16, 14, 30, 16, 20, 20,
                        14, 16, 20, 20, 20, 18, 20, 20,
                        14, 18, 14, 14, 12, 14, 12, 10, 14, 14]

        def _row_data(r):
            return [
                r["barcode"],   # Paket-Barcode
                None,           # Reservierungsnr.
                None,           # Best.-Nr.
                None,           # Bestellnummer
                r["scan_datum"],# Datum
                r["name"],      # Name
                None,           # Ankunftsort
                r["zielkiosk"], # Ziel-Kiosk
                None,           # Lieferung-Name
                None,           # Standort
                r["db_status"], # Paketstatus
                None,           # Lieferung-Adresse
                None,           # Abgeholt_At
                r["abholbereit_at"],  # Abholbereit_At
                None,           # Lieferung-Zusatz
                r["verpackt_at"],     # Verpackt_At
                r["zielkiosk"], # Ziel Kiosk (Duplikat)
                None,           # Unterschrift
                r["scan_datum"],# Scan-Datum
                r["tb_status"], # Status (Kontrollstatus Tagesbote)
                None,           # Bestellwert
                None,           # Versicher.
                None,           # Lieferung
                None,           # Zahlung
                None,           # Rezept
                None,           # Notizen
                None,           # Paketfoto
            ]

        def _fill_sheet(ws, rows, sheet_title):
            ws.title = sheet_title
            ws.append(HEADERS)
            for i, cell in enumerate(ws[1]):
                cell.font = HDR_FONT
            for r in rows:
                ws.append(_row_data(r))
            for i, w in enumerate(COL_WIDTHS, start=1):
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(i)].width = w

        wb = openpyxl.Workbook()
        _fill_sheet(wb.active, angekommen, f"Tour {tour_nr}")
        _fill_sheet(wb.create_sheet(), fehler, "Fehlerliste PU")

        try:
            wb.save(pfad)
            messagebox.showinfo(
                "Export erfolgreich",
                f"Gespeichert:\n{pfad}\n\n"
                f"Tour {tour_nr} (Am Standort): {len(angekommen)} Pakete\n"
                f"Fehlerliste PU: {len(fehler)} Pakete"
            )
        except Exception as e:
            messagebox.showerror("Export Fehler", str(e))


# ============================================================
# GUI – Hauptfenster
# ============================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bombadil")
        self.geometry("1260x860")
        self.minsize(1000, 640)
        self.configure(bg="#f0f2f5")

        # ---- ttk Theme (clam bevorzugt – unterstützt Tab-Farben vollständig)
        style = ttk.Style(self)
        for theme in ("clam", "vista", "default"):
            if theme in style.theme_names():
                style.theme_use(theme)
                break
        style.configure("TNotebook",        background="#f0f2f5", borderwidth=0)
        style.configure("TNotebook.Tab",    padding=[12, 5], font=("Segoe UI", 9),
                        background="#d0d4db", foreground="#333333")
        style.map("TNotebook.Tab",
                  background=[("selected", "#1a3a5c"), ("!selected", "#d0d4db")],
                  foreground=[("selected", "white"),   ("!selected", "#333333")])
        style.configure("Treeview",         rowheight=23, background="white",
                        fieldbackground="white", borderwidth=0)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"),
                        background="#1a3a5c", foreground="white")
        style.map("Treeview.Heading",       background=[("active", "#2a5a8c")])

        self.watch_folder          = str(Path.home() / "Downloads")
        self.last_file: str | None        = None
        self._last_failed_file: str | None = None
        self.last_abholer_df = None
        self._n_dhl_heute    = 0   # DHL Normal für "Pakete heute"-Kachel
        self._n_dhl_express  = 0   # DHL Express für "Pakete heute"-Kachel
        self._n_pu_heute     = 0   # PU heute für "Pakete heute"-Kachel
        self._n_pu_verpackt  = 0   # Davon Verpackt (PU) für Prozentanzeige
        self._last_load_time = 0.0 # Zeitpunkt letzter erfolgreicher Abholer_DB-Laden
        self._load_settings()   # export_folder aus settings.json laden
        if ADMIN_MODE:
            self.after(5_000, self._schedule_backup_check)

        # ---- Menüleiste
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        einst_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Einstellungen", menu=einst_menu)
        einst_menu.add_command(label="Exportordner wählen …", command=self._choose_export_folder)
        if ADMIN_MODE:
            einst_menu.add_separator()
            einst_menu.add_command(label="Backup jetzt erstellen", command=lambda: self._run_backup(manual=True))
        einst_menu.add_separator()
        einst_menu.add_command(label="Einstellungen zurücksetzen", command=self._reset_settings)

        hilfe_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Hilfe", menu=hilfe_menu)
        hilfe_menu.add_command(label="Funktionsübersicht",  command=self._show_help)
        hilfe_menu.add_command(label="Tastenkürzel",        command=self._show_shortcuts)
        hilfe_menu.add_separator()
        hilfe_menu.add_command(label="Über Bombadil",       command=self._show_about)

        # ---- Header
        HDR_BG = "#1a3a5c"
        header = tk.Frame(self, bg=HDR_BG)
        header.pack(fill="x")

        # Logo + Titel (immer links, fixiert)
        self.logo_img = None
        self.logo_lbl = tk.Label(header, bg=HDR_BG)
        self.logo_lbl.pack(side="left", padx=(10, 0), pady=6)
        self.load_logo()

        tk.Label(header, text="Bombadil", font=("Segoe UI", 20, "bold"),
                 fg="white", bg=HDR_BG).pack(side="left", padx=(8, 12), pady=8)

        # ── Kompakte Status-Indikatoren (rechts im Header) ───────────────
        _hdr_right = tk.Frame(header, bg=HDR_BG)
        _hdr_right.pack(side="right", padx=16, pady=6)

        # DB Indikator
        _db_box = tk.Frame(_hdr_right, bg=HDR_BG)
        _db_box.pack(side="left", padx=(0, 16))
        self.refresh_on  = False
        self.refresh_job = None
        self.refresh_btn = tk.Button(
            _db_box, text="⟳  Abholer DB", command=self.toggle_refresh,
            bg="#1c6e3d", fg="white", relief="flat",
            font=("Segoe UI", 9), cursor="hand2",
            activebackground="#17572f", activeforeground="white",
            padx=10, pady=5)
        self.refresh_btn.pack()
        tk.Label(_db_box, text="Abholer-Datenbank", fg="#6a8fad", bg=HDR_BG,
                 font=("Segoe UI", 7)).pack()
        self.refresh_lbl = tk.Label(_db_box, text="", fg="#8aafc8", bg=HDR_BG,
                                    font=("Segoe UI", 7))
        self.refresh_lbl.pack()

        # DHL Indikator
        _dhl_box = tk.Frame(_hdr_right, bg=HDR_BG)
        _dhl_box.pack(side="left")
        self.dhl_refresh_on  = False
        self.dhl_refresh_job = None
        self.dhl_refresh_btn = tk.Button(
            _dhl_box, text="⟳  DHL", command=self.toggle_dhl_refresh,
            bg="#956a0a", fg="white", relief="flat",
            font=("Segoe UI", 9), cursor="hand2",
            activebackground="#7a5508", activeforeground="white",
            padx=10, pady=5)
        self.dhl_refresh_btn.pack()
        tk.Label(_dhl_box, text="DHL-Pakete", fg="#6a8fad", bg=HDR_BG,
                 font=("Segoe UI", 7)).pack()
        self.dhl_refresh_lbl = tk.Label(_dhl_box, text="", fg="#8aafc8", bg=HDR_BG,
                                        font=("Segoe UI", 7))
        self.dhl_refresh_lbl.pack()

        self.watch_on = False   # wird in _build_sidebar_buttons initialisiert

        # ---- Statusleiste (unten)
        statusframe = tk.Frame(self, bg="#1a2744")
        statusframe.pack(fill="x", side="bottom")
        self.statusbar = tk.Label(
            statusframe, text="Bereit.", anchor="w",
            bg="#1a2744", fg="#8aafc8",
            font=("Segoe UI", 8), padx=10, pady=3
        )
        self.statusbar.pack(side="left")
        self._progress = ttk.Progressbar(
            statusframe, mode="indeterminate", length=160
        )
        self._progress.pack(side="left", padx=(6, 8), pady=2)

        # ---- Hauptbereich: Sidebar (links) + Inhalt (rechts)
        SIDE_BG = "#1a2744"
        main_area = tk.Frame(self, bg=SIDE_BG)
        main_area.pack(expand=True, fill="both")

        # Sidebar
        self.sidebar = tk.Frame(main_area, bg=SIDE_BG, width=210)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        # Trennlinie
        tk.Frame(main_area, bg="#0d1b35", width=1).pack(side="left", fill="y")

        # Inhaltsbereich
        content_area = tk.Frame(main_area, bg="#f0f2f5")
        content_area.pack(side="left", expand=True, fill="both")

        # Notebook – Tab-Leiste per Clip-Trick verstecken (Windows-sicher)
        # Der Clip-Frame schneidet die obere Tab-Leiste ab
        _NB_CLIP = 30   # Tab-Leiste auf Windows ~25-28px – 30 reicht sicher
        clip = tk.Frame(content_area, bg="#f0f2f5")
        clip.pack(expand=True, fill="both")
        self.nb = ttk.Notebook(clip)
        self.nb.place(x=0, y=-_NB_CLIP, relwidth=1.0, relheight=1.0, height=_NB_CLIP)

        # 1. Report – Kachel-Ansicht (Hauptansicht)
        self.report_tab = tk.Frame(self.nb, bg="#f0f2f5")
        self.nb.add(self.report_tab, text="  📊 Report  ")

        # Titelzeile
        _title_row = tk.Frame(self.report_tab, bg="#f0f2f5")
        _title_row.pack(fill="x", padx=20, pady=(14, 4))
        tk.Label(_title_row, text="Übersicht", font=("Segoe UI", 14, "bold"),
                 bg="#f0f2f5", fg="#2c3e50").pack(side="left")
        self._report_date_lbl = tk.Label(
            _title_row, text="Noch keine Daten geladen",
            font=("Segoe UI", 9), bg="#f0f2f5", fg="#7f8c8d")
        self._report_date_lbl.pack(side="right", padx=(0, 4))

        # Kacheln – zwei Reihen mit unterschiedlicher Spaltenanzahl
        self._tiles_frame = tk.Frame(self.report_tab, bg="#f0f2f5")
        self._tiles_frame.pack(fill="x", padx=14, pady=(4, 8))

        _row1_frame = tk.Frame(self._tiles_frame, bg="#f0f2f5")
        _row1_frame.pack(fill="x")
        for c in range(4):
            _row1_frame.columnconfigure(c, weight=1)

        _row2_frame = tk.Frame(self._tiles_frame, bg="#f0f2f5")
        _row2_frame.pack(fill="x")
        for c in range(3):
            _row2_frame.columnconfigure(c, weight=1)

        self._tile_lbls = {}         # key → (count_lbl, inner, frame, base_color, mode, trend_lbl, pct_lbl)
        self._prev_tile_counts = {}  # speichert letzte Zählwerte für Trend-Pfeil

        _ROW1 = [
            ("pakete_heute", "Pakete heute", "📊", "#1a237e", "neutral",
             "Gesamtzahl aller heute eingegangenen Pakete:\nDHL Normal + DHL Express + PU heute\n\nKlicken für Aufteilung"),
            ("pickup_heute", "PU heute",     "🚐", "#117a65", "neutral",
             "Heutige Pickups (PUs) aus dem Tagesbote-Sheet.\nZeigt alle PUs mit aktuellem Status.\n\nKlicken für Detailansicht"),
            ("abholbereit",  "Abholbereit",  "📦", "#1abc9c", "neutral",
             "Pakete die abholbereit sind und\nauf Abholung warten.\n\nKlicken für Detailliste"),
            ("yesterday",    "PU gestern", "📅", "#5d6d7e", "neutral",
             "Anzahl aller gestern abgeholten Pakete.\n\nKlicken für Detailliste"),
        ]
        _ROW2 = [
            ("pay",    "Zahlung offen", "💰", "#2c3e50", "neutral",
             "Pakete bei denen die Zahlung\nnoch aussteht.\n\nKlicken für Detailliste"),
            ("older7", "> 7 Tage",     "⏰", "#e67e22", "warn50",
             "Abholbereite Pakete die seit mehr\nals 7 Tagen warten.\nWird rot bei mehr als 50 Einträgen.\n\nKlicken für Detailliste"),
            ("kissel", "Kissel > 3W",  "🏪", "#8e44ad", "warn",
             "Pakete im Kissel-Kiosk die\nmehr als 3 Wochen alt sind.\nWird rot sobald Einträge vorhanden.\n\nKlicken für Detailliste"),
        ]
        for col, (key, label, icon, color, mode, tip) in enumerate(_ROW1):
            self._make_tile(_row1_frame, key, label, icon, color, mode, row=0, col=col, tooltip=tip)
        for col, (key, label, icon, color, mode, tip) in enumerate(_ROW2):
            self._make_tile(_row2_frame, key, label, icon, color, mode, row=0, col=col, tooltip=tip)

        ttk.Separator(self.report_tab).pack(fill="x", padx=20, pady=(2, 0))

        # ── Wechselbereich (Chart ↔ Detail-Tabelle) ──────────────────────
        self._bottom_container = tk.Frame(self.report_tab, bg="#f0f2f5")
        self._bottom_container.pack(fill="both", expand=True)

        # ── Panel A: Balkendiagramm (Standard) ───────────────────────────
        self._chart_panel = tk.Frame(self._bottom_container, bg="#f0f2f5")
        self._chart_panel.pack(fill="both", expand=True)

        _bar_hdr = tk.Frame(self._chart_panel, bg="#f0f2f5")
        _bar_hdr.pack(fill="x", padx=20, pady=(8, 0))
        tk.Label(_bar_hdr, text="Abholungen letzte 7 Tage",
                 font=("Segoe UI", 11, "bold"),
                 bg="#f0f2f5", fg="#2c3e50").pack(side="left")

        self._chart_canvas = tk.Canvas(self._chart_panel, bg="#f0f2f5",
                                       height=240, highlightthickness=0)
        self._chart_canvas.pack(fill="x", padx=20, pady=(6, 14))
        self._last_daily7 = []
        self._chart_canvas.bind("<Configure>", lambda e: self._redraw_chart())
        self._chart_canvas.bind("<Motion>",    self._on_chart_hover)
        self._chart_canvas.bind("<Leave>",     lambda e: self._hide_chart_tooltip())
        # Schwebendes Tooltip-Label (anfangs versteckt)
        self._chart_tip = tk.Label(
            self._chart_canvas, bg="#2c3e50", fg="white",
            font=("Segoe UI", 9), padx=6, pady=3, relief="flat",
        )
        self._chart_tip.place_forget()
        self._chart_bar_areas: list = []  # [(x1, y1, x2, y2, day_str, count), ...]

        # ── Panel B: Detail-Tabelle (nach Kachel-Klick) ──────────────────
        self._detail_panel = tk.Frame(self._bottom_container, bg="#f0f2f5")
        # erst bei Klick eingeblendet

        _det_hdr = tk.Frame(self._detail_panel, bg="#f0f2f5")
        _det_hdr.pack(fill="x", padx=14, pady=(8, 4))

        tk.Button(
            _det_hdr, text="← Übersicht",
            command=self._show_report_chart,
            bg="#3d566e", fg="white", relief="flat",
            font=("Segoe UI", 9), cursor="hand2",
            activebackground="#2c4055", activeforeground="white",
            padx=10, pady=4,
        ).pack(side="left")

        self._detail_title_lbl = tk.Label(
            _det_hdr, text="",
            font=("Segoe UI", 11, "bold"),
            bg="#f0f2f5", fg="#2c3e50",
        )
        self._detail_title_lbl.pack(side="left", padx=14)

        # Suchzeile
        _det_search_frame = tk.Frame(self._detail_panel, bg="#f0f2f5")
        _det_search_frame.pack(fill="x", padx=14, pady=(0, 4))
        tk.Label(_det_search_frame, text="🔍", bg="#f0f2f5",
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 4))
        self._detail_search_var = tk.StringVar()
        _det_entry = tk.Entry(_det_search_frame, textvariable=self._detail_search_var,
                              font=("Segoe UI", 10), relief="solid", bd=1, width=30)
        _det_entry.pack(side="left")
        self._detail_search_var.trace_add("write", lambda *_: self._filter_detail_sheet())
        self._detail_all_data: list = []   # ungefilterte Daten für Suche

        try:
            from tksheet import Sheet as _Sheet
            self._detail_sheet = _Sheet(
                self._detail_panel,
                headers=[],
                theme="light green",
                show_row_index=False,
                show_x_scrollbar=True,
                show_y_scrollbar=True,
            )
            self._detail_sheet.enable_bindings("single_select", "column_select",
                                               "row_select", "copy")
            self._detail_sheet.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        except ImportError:
            self._detail_sheet = None
            tk.Label(self._detail_panel, text="tksheet nicht installiert",
                     bg="#f0f2f5").pack()

        # ── Panel C: Pakete-heute Breakdown ──────────────────────────────
        self._breakdown_panel = tk.Frame(self._bottom_container, bg="#f0f2f5")
        # erst bei Klick eingeblendet

        _bk_hdr = tk.Frame(self._breakdown_panel, bg="#f0f2f5")
        _bk_hdr.pack(fill="x", padx=14, pady=(8, 4))
        tk.Button(
            _bk_hdr, text="← Übersicht",
            command=self._show_report_chart,
            bg="#3d566e", fg="white", relief="flat",
            font=("Segoe UI", 9), cursor="hand2",
            activebackground="#2c4055", activeforeground="white",
            padx=10, pady=4,
        ).pack(side="left")
        tk.Label(_bk_hdr, text="Pakete heute – Aufteilung",
                 font=("Segoe UI", 11, "bold"),
                 bg="#f0f2f5", fg="#2c3e50").pack(side="left", padx=14)

        # 3 farbige Karten nebeneinander
        _bk_cards = tk.Frame(self._breakdown_panel, bg="#f0f2f5")
        _bk_cards.pack(fill="x", padx=20, pady=14)
        for c in range(3):
            _bk_cards.columnconfigure(c, weight=1)

        def _bk_card(parent, col, title, color):
            outer = tk.Frame(parent, bg=color, padx=3, pady=3)
            outer.grid(row=0, column=col, sticky="nsew", padx=6)
            inner = tk.Frame(outer, bg=color)
            inner.pack(expand=True, fill="both", padx=14, pady=14)
            cnt = tk.Label(inner, text="–", font=("Segoe UI", 36, "bold"),
                           bg=color, fg="white")
            cnt.pack()
            tk.Label(inner, text=title, font=("Segoe UI", 11),
                     bg=color, fg="white").pack()
            return cnt

        self._bk_dhl_normal_lbl  = _bk_card(_bk_cards, 0, "DHL Normal",  "#2471a3")
        self._bk_dhl_express_lbl = _bk_card(_bk_cards, 1, "DHL Express", "#1a5276")
        self._bk_pu_lbl          = _bk_card(_bk_cards, 2, "PU",          "#16a085")

        # ── Panel E: PU heute Breakdown ──────────────────────────────────
        self._pu_breakdown_panel = tk.Frame(self._bottom_container, bg="#f0f2f5")

        _pub_hdr = tk.Frame(self._pu_breakdown_panel, bg="#f0f2f5")
        _pub_hdr.pack(fill="x", padx=14, pady=(8, 4))
        tk.Button(
            _pub_hdr, text="← Übersicht",
            command=self._show_report_chart,
            bg="#3d566e", fg="white", relief="flat",
            font=("Segoe UI", 9), cursor="hand2",
            activebackground="#2c4055", activeforeground="white",
            padx=10, pady=4,
        ).pack(side="left")
        tk.Label(_pub_hdr, text="PU heute – Aufteilung",
                 font=("Segoe UI", 11, "bold"),
                 bg="#f0f2f5", fg="#2c3e50").pack(side="left", padx=14)

        _pub_cards = tk.Frame(self._pu_breakdown_panel, bg="#f0f2f5")
        _pub_cards.pack(fill="x", padx=20, pady=14)
        for c in range(3):
            _pub_cards.columnconfigure(c, weight=1)

        def _pub_card(parent, col, title, color):
            outer = tk.Frame(parent, bg=color, padx=3, pady=3)
            outer.grid(row=0, column=col, sticky="nsew", padx=6)
            inner = tk.Frame(outer, bg=color)
            inner.pack(expand=True, fill="both", padx=14, pady=14)
            cnt = tk.Label(inner, text="–", font=("Segoe UI", 36, "bold"),
                           bg=color, fg="white")
            cnt.pack()
            tk.Label(inner, text=title, font=("Segoe UI", 11),
                     bg=color, fg="white").pack()
            return cnt

        self._pub_tagesbote_lbl = _pub_card(_pub_cards, 0, "PU Tagesbote", "#117a65")
        self._pub_verpackt_lbl  = _pub_card(_pub_cards, 1, "Verpackt (PU)",    "#16a085")
        self._pub_abhol_lbl     = _pub_card(_pub_cards, 2, "Abholbereit (PU)", "#1abc9c")

        # Tagesboten Abgleich ist jetzt als Seitenpanel in PU heute eingebettet
        # (kein separater Tab mehr nötig)

        # 2b. Statistik (PU + DHL)
        self.tab_statistik = StatistikTab(
            self.nb,
            start_loading=self._start_loading,
            stop_loading=self._stop_loading,
        )
        self.nb.add(self.tab_statistik.frame, text="  Statistik  ")

        # 3. PU heute
        self.tab_pickup_heute = PickupHeuteTab(
            self.nb,
            get_abholer_df    = lambda: self.last_abholer_df,
            on_count_change   = self._on_pu_count_change,
            get_export_folder = lambda: self.export_folder,
            on_pu_loaded      = self._enrich_pay_tours,
        )
        self.nb.add(self.tab_pickup_heute.frame, text="  Pickup heute  ")

        # 4. Abholbereit
        def _abhol_color(row):
            wt = row[4] if len(row) > 4 else ""
            if wt == "?":
                return "#fff3cd"   # gelb – kein Timestamp
            try:
                days = int(wt.split()[0])
                if days > 7:
                    return "#f8d7da"   # rot – älter als 7 Tage
                if days > 3:
                    return "#fef3e2"   # orange – älter als 3 Tage
            except Exception:
                pass
            return None
        self.tab_abhol = TableTab(
            self.nb, "Abholbereit",
            [("barcode", "Paket-Barcode", 230), ("dt", "Abholbereit_At", 175),
             ("name", "Name", 400), ("kiosk", "Ziel-Kiosk", 130), ("wt", "Wartezeit", 90)],
            row_color_fn=_abhol_color,
            editable_col_map={3: "zielu002dkiosk"},
            orca_sheet_id=ORCA_ABHOLER_SHEET_ID,
            orca_id_idx=5,
        )
        self.nb.add(self.tab_abhol.frame, text="  Abholbereit  ")

        _b = tk.Button(self.tab_abhol.btn_frame, text="↩  Retoure",
                       command=self._abhol_set_retoure,
                       bg="#e8a44a", fg="white", activebackground="#d08830",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(8, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Retoure' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")
        _b = tk.Button(self.tab_abhol.btn_frame, text="✗  Storno",
                       command=self._abhol_set_storno,
                       bg="#d96b6b", fg="white", activebackground="#c04848",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(4, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Storno' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")
        _b = tk.Button(self.tab_abhol.btn_frame, text="✓  Abgeholt",
                       command=self._abhol_set_abgeholt,
                       bg="#4ea874", fg="white", activebackground="#37855a",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(4, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Abgeholt' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")

        # Farblegende in der Button-Zeile des Abholbereit-Tabs
        _legend_frame = tk.Frame(self.tab_abhol.btn_frame)
        _legend_frame.pack(side="right", padx=(12, 0))
        tk.Label(_legend_frame, text="Legende:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 6))
        for _bg, _txt in [("#fff3cd", "kein Datum"), ("#fef3e2", "> 3 Tage"), ("#f8d7da", "> 7 Tage")]:
            _box = tk.Frame(_legend_frame, bg=_bg, width=14, height=14,
                            relief="solid", bd=1)
            _box.pack(side="left", padx=(0, 2))
            _box.pack_propagate(False)
            tk.Label(_legend_frame, text=_txt, font=("Segoe UI", 8)).pack(side="left", padx=(0, 8))

        # Verpackt – als versteckter Tab (kein Sidebar-Eintrag, nur für Kachel-Detail)
        def _verpackt_color(row):
            wt = row[4] if len(row) > 4 else ""
            if wt == "?":
                return "#fff3cd"
            try:
                days = int(wt.split()[0])
                if days > 7:
                    return "#f8d7da"
                if days > 3:
                    return "#fef3e2"
            except Exception:
                pass
            return None
        self.tab_verpackt = TableTab(
            self.nb, "Verpackt – wartet auf Abholung",
            [("barcode", "Paket-Barcode", 230), ("dt", "Verpackt_At", 175),
             ("name", "Name", 400), ("kiosk", "Ziel-Kiosk", 130), ("wt", "Wartezeit", 90)],
            row_color_fn=_verpackt_color,
        )
        self.nb.add(self.tab_verpackt.frame, text="  Verpackt  ")

        # 5. Zahlung offen
        _PAY_OFFEN       = "#f0b8b0"   # altrosa     – noch nicht verpackt / offen
        _PAY_VERPACKT    = "#f0e4a0"   # buttergelb  – verpackt, wartet auf Tour
        _PAY_ABHOLBEREIT = "#bdd8c8"   # salbeigrün  – am Standort
        _PAY_ABGEHOLT    = "#bdd0e8"   # stahlblau   – abgeholt

        def _pay_color(row):
            s = str(row[3]).lower() if len(row) > 3 else ""
            if s == "abgeholt":    return _PAY_ABGEHOLT
            if s == "abholbereit": return _PAY_ABHOLBEREIT
            if "verpackt" in s:    return _PAY_VERPACKT
            return _PAY_OFFEN

        self.tab_pay = TableTab(
            self.nb,
            "Zahlung offen",
            [("barcode",  "Paket-Barcode",  200),
             ("name",     "Name",           320),
             ("bw",       "Bestellwert",    110),
             ("status",   "Status",         130),
             ("dt",       "Abholbereit_At", 155),
             ("wt",       "Wartezeit",       80)],
            row_color_fn=_pay_color,
            editable_col_map={2: "bestellwert"},
            orca_sheet_id=ORCA_ABHOLER_SHEET_ID,
            orca_id_idx=6,
            legend_items=[
                (_PAY_OFFEN,       "Offen / noch nicht verpackt"),
                (_PAY_VERPACKT,    "Verpackt – wartet auf Tour"),
                (_PAY_ABHOLBEREIT, "Abholbereit – am Standort"),
                (_PAY_ABGEHOLT,    "Abgeholt"),
            ],
        )
        self.nb.add(self.tab_pay.frame, text="  Zahlung offen  ")

        # Standard-Buttons entfernen (Kopieren, Suche leeren werden nicht gebraucht)
        for _w in list(self.tab_pay.btn_frame.winfo_children()):
            _w.destroy()

        # Rechtsklick → Kontextmenü
        if self.tab_pay.sheet:
            self.tab_pay.sheet.bind("<Button-3>", self._pay_right_click)

        # 6. > 7 Tage
        def _older7_color(row):
            wt = row[4] if len(row) > 4 else ""
            try:
                days = int(wt.split()[0])
                if days > 30:
                    return "#a93226"   # dunkelrot – älter als 30 Tage
                if days > 14:
                    return "#f8d7da"   # rot – 15–30 Tage
                return "#fef3e2"       # orange – 7–14 Tage
            except Exception:
                return None

        self.tab_older = TableTab(
            self.nb, "> 7 Tage abholbereit",
            [("barcode", "Paket-Barcode", 230), ("name", "Name", 400),
             ("dt", "Abholbereit_At", 175), ("kiosk", "Ziel-Kiosk", 130), ("wt", "Wartezeit", 90)],
            row_color_fn=_older7_color,
        )
        self.nb.add(self.tab_older.frame, text="  > 7 Tage  ")

        _b = tk.Button(self.tab_older.btn_frame, text="↩  Retoure",
                       command=self._older_set_retoure,
                       bg="#e8a44a", fg="white", activebackground="#d08830",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(8, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Retoure' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")
        _b = tk.Button(self.tab_older.btn_frame, text="✗  Storno",
                       command=self._older_set_storno,
                       bg="#d96b6b", fg="white", activebackground="#c04848",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(4, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Storno' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")
        _b = tk.Button(self.tab_older.btn_frame, text="✓  Abgeholt",
                       command=self._older_set_abgeholt,
                       bg="#4ea874", fg="white", activebackground="#37855a",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(4, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Abgeholt' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")

        # Farblegende > 7 Tage
        _leg7 = tk.Frame(self.tab_older.btn_frame)
        _leg7.pack(side="right", padx=(12, 0))
        tk.Label(_leg7, text="Legende:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 6))
        for _bg, _fg, _txt in [("#fef3e2", "black", "7–14 Tage"),
                                ("#f8d7da", "black", "15–30 Tage"),
                                ("#a93226", "white", "> 30 Tage")]:
            _box = tk.Frame(_leg7, bg=_bg, width=14, height=14, relief="solid", bd=1)
            _box.pack(side="left", padx=(0, 2))
            _box.pack_propagate(False)
            tk.Label(_leg7, text=_txt, font=("Segoe UI", 8), fg=_fg).pack(side="left", padx=(0, 8))

        # 7. Kissel > 3W
        def _kissel_color(row):
            wt = row[4] if len(row) > 4 else ""
            try:
                days = int(wt.split()[0])
                if days > 90:
                    return "#a93226"   # dunkelrot – älter als 3 Monate
                if days > 42:
                    return "#f8d7da"   # rot – älter als 6 Wochen
                return "#fef3e2"       # orange – 3–6 Wochen
            except Exception:
                return None

        self.tab_kissel = TableTab(
            self.nb, "Kissel: länger als 3 Wochen abholbereit",
            [("barcode", "Paket-Barcode", 230), ("name", "Name", 400),
             ("dt", "Abholbereit_At", 175), ("kiosk", "Ziel-Kiosk", 130), ("wt", "Wartezeit", 90)],
            row_color_fn=_kissel_color,
        )
        self.nb.add(self.tab_kissel.frame, text="  Kissel > 3W  ")

        _b = tk.Button(self.tab_kissel.btn_frame, text="↩  Retoure",
                       command=self._kissel_set_retoure,
                       bg="#e8a44a", fg="white", activebackground="#d08830",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(8, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Retoure' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")
        _b = tk.Button(self.tab_kissel.btn_frame, text="✗  Storno",
                       command=self._kissel_set_storno,
                       bg="#d96b6b", fg="white", activebackground="#c04848",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(4, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Storno' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")
        _b = tk.Button(self.tab_kissel.btn_frame, text="✓  Abgeholt",
                       command=self._kissel_set_abgeholt,
                       bg="#4ea874", fg="white", activebackground="#37855a",
                       activeforeground="white", font=("Segoe UI", 9, "bold"),
                       relief="flat", padx=10, pady=3, cursor="hand2")
        _b.pack(side="left", padx=(4, 0))
        add_tooltip(_b, "Markierte Pakete in OrcaScan auf 'Abgeholt' setzen.\n"
                        "Zeile(n) in der Tabelle markieren, dann klicken.")

        # Farblegende Kissel
        _leg_kissel = tk.Frame(self.tab_kissel.btn_frame)
        _leg_kissel.pack(side="right", padx=(12, 0))
        tk.Label(_leg_kissel, text="Legende:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 6))
        for _bg, _fg, _txt in [("#fef3e2", "black", "3–6 Wochen"),
                                ("#f8d7da", "black", "> 6 Wochen"),
                                ("#a93226", "white", "> 3 Monate")]:
            _box = tk.Frame(_leg_kissel, bg=_bg, width=14, height=14, relief="solid", bd=1)
            _box.pack(side="left", padx=(0, 2))
            _box.pack_propagate(False)
            tk.Label(_leg_kissel, text=_txt, font=("Segoe UI", 8), fg=_fg).pack(side="left", padx=(0, 8))

        # 8. Unstimmigkeiten (Verpackt aber nicht im Tagesbote)
        self._unstimmigkeiten_updates = []   # (row_id, barcode, name) für OrcaScan
        self._unstimm_container = tk.Frame(self.nb, bg="#f0f2f5")
        self.nb.add(self._unstimm_container, text="  Unstimmigkeiten  ")

        _ustimm_btn_frame = tk.Frame(self._unstimm_container, bg="#f0f2f5")
        _ustimm_btn_frame.pack(fill="x", padx=10, pady=(8, 0))
        self._unstimmig_btn = tk.Button(
            _ustimm_btn_frame,
            text="→ Abholbereit setzen",
            command=self._unstimmig_set_abholbereit,
            bg="#e67e22", fg="white", relief="flat",
            font=("Segoe UI", 9, "bold"), cursor="hand2",
            activebackground="#d35400", activeforeground="white",
            padx=10, pady=5, state="disabled",
        )
        self._unstimmig_btn.pack(side="left")
        add_tooltip(self._unstimmig_btn,
                    "Setzt alle aufgelisteten Pakete in OrcaScan auf 'Abholbereit'.\n"
                    "Nur aktiv wenn Unstimmigkeiten vorhanden.\n\n"
                    "Unstimmigkeit = Status 'Verpackt' in Abholer_DB,\n"
                    "aber kein Eintrag im heutigen Tagesboten.")

        self.tab_unstimmig = TableTab(
            self._unstimm_container, "Verpackt – nicht im heutigen Tagesbote",
            [("barcode", "Paket-Barcode", 260), ("name", "Name", 400),
             ("vp", "Verpackt_At", 180), ("kiosk", "Ziel-Kiosk", 150)],
        )
        self.tab_unstimmig.frame.pack(fill="both", expand=True)

        # 9. Gestern
        self.tab_yest = TableTab(
            self.nb, "Gestern abgeholt",
            [("barcode", "Paket-Barcode", 230), ("name", "Name", 400),
             ("dt", "Abgeholt_At", 175), ("kiosk", "Ziel-Kiosk", 130)],
        )
        self.nb.add(self.tab_yest.frame, text="  Gestern  ")

        # 9. DHL_Express (heute)
        self.tab_dhl = TableTab(
            self.nb, "DHL_Express – Scans von heute",
            [("barcode", "Package Barcode", 420), ("dt", "Date Of Scan", 260)],
            today_header=True,
        )
        self.nb.add(self.tab_dhl.frame, text="  DHL_Express  ")

        # 10. DHL (heute) Merge
        self.tab_dhl_merge = DHLMergeTab(self.nb, get_export_folder=lambda: self.export_folder)
        self.nb.add(self.tab_dhl_merge.frame, text="  DHL (heute)  ")


        # ---- Tab-Tooltips
        _TAB_TIPS = {
            "  📊 Report  ":
                "Startansicht mit Echtzeit-Kacheln:\n"
                "• Pakete heute (DHL + PU gesamt)\n"
                "• PU heute (verpackt · offen · %)\n"
                "• Abholbereit (warten auf Abholung)\n"
                "• Zahlung offen · > 7 Tage · Kissel > 3W\n"
                "Klick auf Kachel → Detailliste\n"
                "Balkendiagramm: Abholungen letzte 7 Tage",

            "  Tagesboten Abgleich  ":
                "Vergleicht den heutigen Tagesboten (Google Drive)\n"
                "mit der Abholer_DB aus OrcaScan.\n\n"
                "Fehlerliste: In Abholer_DB vorhanden,\n"
                "aber Abholbereit_At fehlt.\n"
                "→ Direkt auf Abholbereit setzen möglich.\n\n"
                "Errorliste: Barcode gar nicht in Abholer_DB.\n"
                "→ Neue Einträge anlegen möglich.\n\n"
                "Nur Einträge mit Kontrollstatus 'Verpackt'\n"
                "werden geprüft.",

            "  Statistik  ":
                "Statistiken für PU und DHL:\n\n"
                "PU: Lieferungen & Kundenabholungen\n"
                "nach Woche / Monat, aufgeteilt nach Kiosk.\n\n"
                "DHL: Gesamt, Normal, Express, Abholung\n"
                "nach Woche / Monat.\n\n"
                "Daten aus OrcaScan + Google Drive Archiv.",

            "  Pickup heute  ":
                "Heutige Pickup-Aufträge aus dem Tagesboten.\n\n"
                "Tour-Zuweisung (automatisch):\n"
                "• Verpackt vor 11:16 Uhr → Tour 1\n"
                "• Verpackt ab 11:16 Uhr → Tour 2\n"
                "• Paket vom Vortag → immer Tour 1\n"
                "• Offen nach 11:16 Uhr → Tour 2\n\n"
                "Export: Tour 1 / Tour 2 als Excel\n"
                "(Sheet 1: Am Standort, Sheet 2: Fehlerliste)\n\n"
                "Farben: Altrosa=Offen · Gelb=Verpackt\n"
                "Grün=Am Standort · Blau=Abgeholt · Apricot=Retoure",

            "  Abholbereit  ":
                "Alle Pakete mit Status 'Abholbereit',\n"
                "die noch nicht vom Kunden abgeholt wurden.\n\n"
                "Sortierung: Älteste zuerst.\n\n"
                "Farben: Gelb=kein Datum · Orange=>3 Tage · Rot=>7 Tage\n\n"
                "Aktionen (Zeilen markieren, dann Button):\n"
                "• ↩ Retoure – Paket zurückschicken\n"
                "• ✗ Storno – Auftrag stornieren\n"
                "• ✓ Abgeholt – Abholung bestätigen",

            "  Verpackt  ":
                "Alle Pakete mit Status 'Verpackt',\n"
                "die noch kein Abholbereit_At haben\n"
                "(= noch nicht beim Kiosk angekommen).\n\n"
                "Sortierung: Älteste zuerst.",

            "  Zahlung offen  ":
                "Abholbereite Pakete bei denen\n"
                "die Zahlung noch aussteht.\n\n"
                "Zahlungsstatus: unbezahlt / offen / vor Ort\n\n"
                "Sortierung: Älteste zuerst.\n\n"
                "Aktion: Markierte Pakete → 'bezahlt' setzen.",

            "  > 7 Tage  ":
                "Abholbereite Pakete die seit mehr als\n"
                "7 Tagen nicht abgeholt wurden.\n\n"
                "Farben: Orange=7–14 Tage\n"
                "Rot=15–30 Tage · Dunkelrot=>30 Tage\n\n"
                "Aktionen (Zeilen markieren, dann Button):\n"
                "• ↩ Retoure · ✗ Storno · ✓ Abgeholt",

            "  Kissel > 3W  ":
                "Pakete am Standort Kissel die seit\n"
                "mehr als 3 Wochen nicht abgeholt wurden.\n\n"
                "Farben: Orange=3–6 Wochen\n"
                "Rot=6W–3 Monate · Dunkelrot=>3 Monate\n\n"
                "Aktionen (Zeilen markieren, dann Button):\n"
                "• ↩ Retoure · ✗ Storno · ✓ Abgeholt",

            "  Unstimmigkeiten  ":
                "Pakete mit Status 'Verpackt' in der Abholer_DB,\n"
                "die NICHT im heutigen Tagesboten stehen.\n\n"
                "Mögliche Ursache: Paket wurde gescannt,\n"
                "aber nicht als PU-Auftrag angelegt.\n\n"
                "Aktion: Alle Pakete direkt auf\n"
                "'Abholbereit' setzen.",

            "  Gestern  ":
                "Alle Pakete die gestern vom Kunden\n"
                "abgeholt wurden.\n\n"
                "Sortierung: Neueste zuerst.",

            "  DHL_Express  ":
                "DHL Express Scans von heute.\n\n"
                "Zeigt: Package Barcode + Scan-Datum.\n"
                "Nur Einträge vom heutigen Tag.",

            "  DHL (heute)  ":
                "DHL Normal Scans von heute\n"
                "(aus OrcaScan DHL_Normal Sheet).\n\n"
                "Export als YYMMDD.xlsx in den Exportordner.\n"
                "Barcode-Spalte wird als Text exportiert.",
        }

        _nb_tip = ToolTip(self.nb)

        def _on_nb_motion(event):
            try:
                idx = self.nb.index(f"@{event.x},{event.y}")
                tab_text = self.nb.tab(idx, "text")
                tip = _TAB_TIPS.get(tab_text, "")
                if tip:
                    _nb_tip.showtip(tip, event.x_root, event.y_root)
                else:
                    _nb_tip.hidetip()
            except Exception:
                _nb_tip.hidetip()

        def _on_nb_leave(event):
            _nb_tip.hidetip()

        self.nb.bind("<Motion>",  _on_nb_motion)
        self.nb.bind("<Leave>",   _on_nb_leave)
        self.nb.bind("<Button-1>", lambda e: _nb_tip.hidetip())

        # ---- Keyboard-Shortcuts
        self.bind("<Control-o>", lambda e: self.open_file())
        self.bind("<Control-O>", lambda e: self.open_file())
        self.bind("<Control-d>", lambda e: self.open_dhl_file())
        self.bind("<Control-D>", lambda e: self.open_dhl_file())
        self.bind("<F5>",        lambda e: self.reload())

        self.after(POLL_MS, self.poll)

        # Konfiguration für Detail-Ansicht im Report-Tab
        self._tile_detail_cfg = {
            "abholbereit": (self.tab_abhol,    "📦 Abholbereit",
                            [("Paket-Barcode", 230), ("Abholbereit_At", 175), ("Name", 400), ("Ziel-Kiosk", 130), ("Wartezeit", 90)]),
            "verpackt":    (self.tab_verpackt, "📫 Verpackt",
                            [("Paket-Barcode", 230), ("Verpackt_At", 175), ("Name", 400), ("Ziel-Kiosk", 130), ("Wartezeit", 90)]),
            "pay":         (self.tab_pay,      "💰 Zahlung offen",
                            [("Paket-Barcode", 220), ("Abholbereit_At", 175), ("Name", 380), ("Zahlung", 160), ("Ziel-Kiosk", 120), ("Wartezeit", 90)]),
            "older7":      (self.tab_older,    "⏰ > 7 Tage",
                            [("Paket-Barcode", 230), ("Name", 400), ("Abholbereit_At", 175), ("Ziel-Kiosk", 130), ("Wartezeit", 90)],
                            _older7_color),
            "kissel":      (self.tab_kissel,   "🏪 Kissel > 3W",
                            [("Paket-Barcode", 230), ("Name", 400), ("Abholbereit_At", 175), ("Ziel-Kiosk", 130), ("Wartezeit", 90)],
                            _kissel_color),
            "yesterday":   (self.tab_yest,     "✅ Gestern abgeholt",
                            [("Paket-Barcode", 230), ("Name", 400), ("Abgeholt_At", 175), ("Ziel-Kiosk", 130)]),
        }

        # Sidebar aufbauen + Report als Startansicht
        self._active_sidebar_key = "report"
        self._sidebar_btns = {}
        self._build_sidebar_buttons()
        self._select_tab(self.report_tab, "report")

        # Beim Start automatisch OrcaScan + DHL laden + Auto-Refresh aktivieren
        self.after(200,    self.load_main_orca)
        self.after(400,    self.load_dhl_orca)
        self.after(600,    self.toggle_refresh)
        self.after(800,    self.toggle_dhl_refresh)
        self.after(3_000,  self.tab_statistik.load_archive_async)  # Archiv nach Hauptladen
        self.after(5_000,  self.tab_statistik.load_dhl_async)      # DHL Statistik
        self.after(60_000, self._check_stale)  # Stale-Data Indikator starten

    # ------------------------------------------------------------------ helpers

    # ------------------------------------------------------------------ report tiles

    def _make_tile(self, parent, key, label, icon, base_color, mode, row, col, tooltip=None):
        """Erstellt eine einzelne klickbare Kachel im Report-Tab."""
        outer = tk.Frame(parent, bg=base_color, cursor="hand2", bd=0)
        outer.grid(row=row, column=col, padx=8, pady=8, sticky="ew")

        inner = tk.Frame(outer, bg=base_color, padx=18, pady=14)
        inner.pack(expand=True, fill="both")

        count_lbl = tk.Label(inner, text="—",
                             font=("Segoe UI", 28, "bold"),
                             bg=base_color, fg="white", anchor="center")
        count_lbl.pack(fill="x")

        # Trend-Pfeil (▲ grün / ▼ rot / leer beim ersten Laden)
        trend_lbl = tk.Label(inner, text="",
                             font=("Segoe UI", 11, "bold"),
                             bg=base_color, fg="white", anchor="center")
        trend_lbl.pack(fill="x")

        # Prozentzahl (z.B. "15%  aller aktiven Pakete")
        pct_lbl = tk.Label(inner, text="",
                           font=("Segoe UI", 8),
                           bg=base_color, fg="#c8e6e0", anchor="center")
        pct_lbl.pack(fill="x")

        name_lbl = tk.Label(inner, text=f"{icon}  {label}",
                            font=("Segoe UI", 10),
                            bg=base_color, fg="white", anchor="center")
        name_lbl.pack(fill="x")

        # Hover-Effekt
        def _on_enter(e, w_list=(outer, inner, count_lbl, trend_lbl, pct_lbl, name_lbl), c=base_color):
            darker = "#" + "".join(f"{max(0, int(c[i:i+2], 16) - 28):02x}"
                                   for i in (1, 3, 5))
            for w in w_list:
                try:
                    w.config(bg=darker)
                except Exception:
                    pass

        def _on_leave(e, w_list=(outer, inner, count_lbl, trend_lbl, pct_lbl, name_lbl)):
            bg = self._tile_lbls[key][3]   # aktuell gesetzte Farbe
            for w in w_list:
                try:
                    w.config(bg=bg)
                except Exception:
                    pass

        for w in (outer, inner, count_lbl, trend_lbl, pct_lbl, name_lbl):
            w.bind("<Button-1>", lambda e, k=key: self._navigate_tile(k))
            w.bind("<Enter>",    _on_enter)
            w.bind("<Leave>",    _on_leave)

        if tooltip:
            _tip = add_tooltip(outer, tooltip)
            for w in (inner, count_lbl, trend_lbl, pct_lbl, name_lbl):
                w.bind("<Enter>", lambda e, t=_tip: t._schedule(), add="+")
                w.bind("<Leave>", lambda e, t=_tip: t._cancel(),   add="+")

        # (count_lbl, inner, outer, aktuelle_farbe, mode, trend_lbl, pct_lbl)
        self._tile_lbls[key] = (count_lbl, inner, outer, base_color, mode, trend_lbl, pct_lbl)

    def _navigate_tile(self, key: str):
        """Kachel-Klick: Detail-Tabelle im Report-Tab oder direkt zum Tab navigieren."""
        if key == "pickup_heute":
            self._show_pu_breakdown()
            return
        if key == "pakete_heute":
            self._show_pakete_breakdown()
            return
        self._show_tile_detail(key)

    def _show_pakete_breakdown(self):
        """Zeigt die Pakete-heute Aufteilung im unteren Bereich."""
        n_dhl_normal  = self._n_dhl_heute
        n_dhl_express = self._n_dhl_express
        n_pu = 0
        if "pickup_heute" in self._tile_lbls:
            try:
                n_pu = int(self._tile_lbls["pickup_heute"][0].cget("text") or 0)
            except Exception:
                pass
        self._bk_dhl_normal_lbl.config(text=str(n_dhl_normal))
        self._bk_dhl_express_lbl.config(text=str(n_dhl_express))
        self._bk_pu_lbl.config(text=str(n_pu))
        self._chart_panel.pack_forget()
        self._detail_panel.pack_forget()
        self._pu_breakdown_panel.pack_forget()
        self._breakdown_panel.pack(fill="both", expand=True)

    def _show_pu_breakdown(self):
        """Zeigt PU-heute Aufteilung: Tagesbote gesamt, Verpackt, Abholbereit."""
        rows = getattr(self.tab_pickup_heute, "_last_rows", None) or []
        n_total    = len(rows)
        n_abhol    = sum(1 for r in rows if r.get("abholbereit_at")
                         or r.get("db_status") == "abgeholt")
        n_verpackt = sum(1 for r in rows if r.get("tb_status", "").lower() == "verpackt"
                         and not r.get("_abholbereit_bool")
                         and r.get("db_status") != "abgeholt")
        self._pub_tagesbote_lbl.config(text=str(n_total))
        self._pub_verpackt_lbl.config(text=str(n_verpackt))
        self._pub_abhol_lbl.config(text=str(n_abhol))
        # Speichern für Prozentanzeige in PU-heute-Kachel
        self._n_pu_verpackt = n_verpackt
        self._update_pu_pct_label()
        self._chart_panel.pack_forget()
        self._detail_panel.pack_forget()
        self._breakdown_panel.pack_forget()
        self._pu_breakdown_panel.pack(fill="both", expand=True)

    def _refresh_unstimmigkeiten(self):
        """Vergleicht Verpackt-Pakete (Abholer_DB) mit heutigem Tagesbote → zeigt Differenz."""
        df = self.last_abholer_df
        if df is None:
            return
        c_status = first_existing(df, COL_STATUS)
        c_abhol  = first_existing(df, COL_ABHOLBEREIT)
        c_bc     = first_existing(df, COL_BARCODE)
        c_name   = first_existing(df, COL_NAME)
        c_zk     = first_existing(df, COL_ZIELKIOSK)
        c_vp     = first_existing(df, COL_VERPACKT_AT)
        if not all([c_status, c_bc, c_name]):
            return

        status_norm = df[c_status].astype(str).str.strip().str.lower()
        abhol_na    = df[c_abhol].isna() if c_abhol else pd.Series([True] * len(df), index=df.index)
        verpackt_df = df[status_norm.eq("verpackt") & abhol_na].copy()

        # Barcodes aus PU heute (Tagesbote)
        pu_rows     = getattr(self.tab_pickup_heute, "_last_rows", None) or []
        pu_barcodes = {str(r.get("barcode", "")).strip().upper() for r in pu_rows if r.get("barcode")}

        # Unstimmigkeiten: in Abholer_DB verpackt, aber NICHT im Tagesbote
        bc_norm = verpackt_df[c_bc].astype(str).str.strip().str.upper()
        unstimm_df = verpackt_df[~bc_norm.isin(pu_barcodes)].copy()

        self._unstimmigkeiten_updates = []
        rows = []
        for _, r in unstimm_df.iterrows():
            bc  = "" if pd.isna(r[c_bc])   else str(r[c_bc]).strip()
            if not bc:
                continue
            nm  = "" if pd.isna(r[c_name]) else str(r[c_name]).strip()
            vp  = fmt_dt(r[c_vp]) if c_vp and not pd.isna(r[c_vp]) else ""
            zk  = "" if (c_zk is None or pd.isna(r[c_zk])) else str(r[c_zk]).strip()
            rid = str(r.get("_id", "")).strip()
            rows.append((bc, nm, vp, zk))
            self._unstimmigkeiten_updates.append((rid, bc, nm))

        self.tab_unstimmig.set_rows(rows)
        n = len(rows)
        if n > 0:
            self._unstimmig_btn.config(
                state="normal",
                text=f"→ Alle {n} Pakete → Abholbereit setzen",
            )
        else:
            self._unstimmig_btn.config(state="disabled", text="→ Abholbereit setzen")

    def _unstimmig_set_abholbereit(self):
        """Setzt alle Unstimmigkeiten-Pakete in OrcaScan auf Abholbereit."""
        import threading as _th
        updates = list(self._unstimmigkeiten_updates)
        if not updates:
            return
        from datetime import datetime as _dt
        ts = _dt.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        extra = {"location": "Abholbereit", "abholbereitu005fat": ts}
        self._unstimmig_btn.config(state="disabled", text="Wird gesetzt …")

        def _worker():
            result = update_rows_orca_bulk(updates, extra)
            self.after(0, lambda: self._after_unstimmig(result))

        _th.Thread(target=_worker, daemon=True).start()

    def _after_unstimmig(self, result):
        ok     = result.get("ok", [])
        failed = result.get("failed", [])
        if failed:
            messagebox.showwarning(
                "Teilweise fehlgeschlagen",
                f"{len(ok)} gesetzt, {len(failed)} fehlgeschlagen.",
            )
        else:
            messagebox.showinfo("Fertig", f"{len(ok)} Pakete → Abholbereit gesetzt.")
        self.reload()

    # ------------------------------------------------------------------ Abholbereit-Aktionen

    def _abhol_set_retoure(self):
        self._abhol_set_status("Retoure", self.tab_abhol, barcode_idx=0, name_idx=2)

    def _abhol_set_storno(self):
        self._abhol_set_status("Storno", self.tab_abhol, barcode_idx=0, name_idx=2)

    def _abhol_set_abgeholt(self):
        self._abhol_set_status("Abgeholt", self.tab_abhol, barcode_idx=0, name_idx=2)

    def _older_set_retoure(self):
        self._abhol_set_status("Retoure", self.tab_older, barcode_idx=0, name_idx=1)

    def _older_set_storno(self):
        self._abhol_set_status("Storno", self.tab_older, barcode_idx=0, name_idx=1)

    def _older_set_abgeholt(self):
        self._abhol_set_status("Abgeholt", self.tab_older, barcode_idx=0, name_idx=1)

    def _kissel_set_retoure(self):
        self._abhol_set_status("Retoure", self.tab_kissel, barcode_idx=0, name_idx=1)

    def _kissel_set_storno(self):
        self._abhol_set_status("Storno", self.tab_kissel, barcode_idx=0, name_idx=1)

    def _kissel_set_abgeholt(self):
        self._abhol_set_status("Abgeholt", self.tab_kissel, barcode_idx=0, name_idx=1)

    def _abhol_set_status(self, new_status: str, tab, barcode_idx: int = 0, name_idx: int = 2):
        import threading as _th
        selected = tab.get_selected_rows()
        if not selected:
            messagebox.showinfo("Hinweis", "Keine Zeile markiert.")
            return
        df = self.last_abholer_df
        if df is None:
            messagebox.showwarning("Fehler", "Abholer_DB nicht geladen.")
            return

        c_bc   = first_existing(df, COL_BARCODE)
        c_name = first_existing(df, COL_NAME)
        if not c_bc:
            messagebox.showwarning("Fehler", "Barcode-Spalte nicht gefunden.")
            return

        updates = []
        missing = []
        for row in selected:
            barcode = str(row[barcode_idx]).strip()
            mask = df[c_bc].astype(str).str.strip() == barcode
            hits = df[mask]
            if hits.empty:
                missing.append(barcode)
                continue
            r      = hits.iloc[0]
            row_id = str(r.get("_id", "")).strip()
            name   = "" if (c_name is None or pd.isna(r[c_name])) else str(r[c_name]).strip()
            updates.append((row_id, barcode, name))

        if missing:
            messagebox.showwarning(
                "Nicht gefunden",
                "Folgende Barcodes nicht in Abholer_DB:\n" + "\n".join(missing),
            )
        if not updates:
            return

        def _worker():
            result = update_rows_orca_bulk(updates, {"location": new_status})
            self.after(0, lambda: self._after_abhol_set(result, new_status))

        _th.Thread(target=_worker, daemon=True).start()
        self._set_status(f"Setze {len(updates)} Paket(e) → {new_status} …")

    def _after_abhol_set(self, result, new_status: str):
        ok     = result.get("ok", [])
        failed = result.get("failed", [])
        if failed:
            messagebox.showwarning(
                "Teilweise fehlgeschlagen",
                f"{len(ok)} gesetzt, {len(failed)} fehlgeschlagen.",
            )
        else:
            messagebox.showinfo("Fertig", f"{len(ok)} Paket(e) → {new_status} gesetzt.")
        self.reload()

    def _show_tile_detail(self, key: str):
        """Blendet die Detail-Tabelle für eine Kachel im unteren Bereich ein."""
        cfg = self._tile_detail_cfg.get(key)
        if not cfg:
            return
        tab, title, cols = cfg[:3]
        color_fn = cfg[3] if len(cfg) > 3 else None
        rows = list(tab.rows)

        self._detail_title_lbl.config(
            text=f"{title}  –  {len(rows)} Einträge"
        )

        if self._detail_sheet is not None:
            headers = [c[0] for c in cols]
            widths   = [c[1] for c in cols]
            data     = [[str(v) if v is not None else "" for v in row]
                        for row in rows]
            self._detail_all_data = data
            self._detail_col_widths = widths
            self._detail_search_var.set("")          # Suchfeld leeren beim Tab-Wechsel
            self._detail_sheet.headers(headers)
            self._detail_sheet.set_sheet_data(data, reset_col_positions=True)
            self._detail_sheet.set_column_widths(widths)
            # Farbcodierung anwenden falls vorhanden
            if color_fn:
                try:
                    self._detail_sheet.dehighlight_all()
                    for i, row in enumerate(rows):
                        bg = color_fn(row)
                        if bg:
                            fg = "white" if bg == "#a93226" else "black"
                            self._detail_sheet.highlight_rows(
                                rows=[i], bg=bg, fg=fg, redraw=False)
                    self._detail_sheet.refresh()
                except Exception:
                    pass

        # Chart-Panel ausblenden, Detail-Panel einblenden
        self._chart_panel.pack_forget()
        self._breakdown_panel.pack_forget()
        self._pu_breakdown_panel.pack_forget()
        self._detail_panel.pack(fill="both", expand=True)

    def _filter_detail_sheet(self):
        """Filtert die Detail-Tabelle nach dem Suchbegriff."""
        if self._detail_sheet is None:
            return
        q = self._detail_search_var.get().strip().lower()
        data = self._detail_all_data
        if q:
            data = [row for row in data if any(q in cell.lower() for cell in row)]
        self._detail_sheet.set_sheet_data(data, reset_col_positions=False)
        widths = getattr(self, "_detail_col_widths", [])
        if widths:
            self._detail_sheet.set_column_widths(widths)

    def _show_report_chart(self):
        """Zurück zum Balkendiagramm (← Übersicht Button)."""
        self._detail_panel.pack_forget()
        self._breakdown_panel.pack_forget()
        self._pu_breakdown_panel.pack_forget()
        self._chart_panel.pack(fill="both", expand=True)

    def _select_tab(self, frame, key: str):
        """Wählt einen Tab aus und aktualisiert die Sidebar-Markierung."""
        if key != "report":
            self._show_report_chart()
        self.nb.select(frame)
        self._active_sidebar_key = key

        SIDE_BG    = "#1a2744"
        ACTIVE_BG  = "#243d5c"
        INDICATOR  = "#4a9fd4"

        for k, (row, ind, lbl) in self._sidebar_btns.items():
            if k == key:
                row.config(bg=ACTIVE_BG)
                ind.config(bg=INDICATOR)
                lbl.config(bg=ACTIVE_BG, fg="white",
                           font=("Segoe UI", 10, "bold"))
            else:
                row.config(bg=SIDE_BG)
                ind.config(bg=SIDE_BG)
                lbl.config(bg=SIDE_BG, fg="#8aafc8",
                           font=("Segoe UI", 10))

    def _build_sidebar_buttons(self):
        """Erstellt die Navigations-Buttons in der Sidebar."""
        SIDE_BG   = "#1a2744"
        HOVER_BG  = "#243a5e"

        # Abschnitts-Überschrift
        def _section_label(text):
            tk.Label(self.sidebar, text=text,
                     font=("Segoe UI", 7, "bold"),
                     bg=SIDE_BG, fg="#4a7a9b",
                     anchor="w", padx=18,
                     ).pack(fill="x", pady=(8, 2))

        # Trennlinie
        def _sep():
            tk.Frame(self.sidebar, bg="#243a5e", height=1
                     ).pack(fill="x", padx=14, pady=4)

        self._sidebar_btns = {}

        ITEMS = [
            "ÜBERSICHT",
            ("report",        "📊  Report",         self.report_tab,                "Übersicht: Kacheln, Statistiken\nund Abholungen der letzten 7 Tage"),
            ("statistik",     "📊  Statistik",        self.tab_statistik.frame,      "Auswertung für PU, DHL Normal und DHL Express\n(Tages-, Wochen-, Monatsansicht)"),
            "PAKETE",
            ("pickup_heute", "🚐  PU heute",        self.tab_pickup_heute.frame,
             "Heutige Pickup-Aufträge.\n"
             "Tour 1: verpackt vor 11:16 Uhr\n"
             "Tour 2: verpackt ab 11:16 Uhr\n"
             "Export: Excel mit Fehlerliste (Tour 1 / Tour 2)"),
            ("abholbereit", "📦  Abholbereit",     self.tab_abhol.frame,
             "Alle abholbereiten Pakete (vollständige Liste).\n"
             "Älteste zuerst. Aktionen:\n"
             "↩ Retoure · ✗ Storno · ✓ Abgeholt"),
            ("pay",         "💰  Zahlung offen",   self.tab_pay.frame,
             "Abholbereite Pakete mit\n"
             "offener Zahlung (unbezahlt / offen / vor Ort).\n"
             "Aktion: Markierte → 'bezahlt' setzen"),
            ("older7",      "⚠️  > 7 Tage",         self.tab_older.frame,
             "Pakete seit mehr als 7 Tagen abholbereit.\n"
             "Orange=7–14 Tage · Rot=15–30 Tage\n"
             "Dunkelrot=>30 Tage\n"
             "Aktionen: ↩ Retoure · ✗ Storno · ✓ Abgeholt"),
            ("kissel",      "🏪  Kissel > 3W",     self.tab_kissel.frame,
             "Kissel-Pakete seit mehr als 3 Wochen.\n"
             "Orange=3–6 Wo · Rot=6Wo–3 Mon\n"
             "Dunkelrot=>3 Monate\n"
             "Aktionen: ↩ Retoure · ✗ Storno · ✓ Abgeholt"),
            ("unstimmig",   "🔍  Unstimmigkeiten", self._unstimm_container,
             "Status 'Verpackt' in Abholer_DB,\n"
             "aber NICHT im heutigen Tagesboten.\n"
             "Aktion: Alle → Abholbereit setzen"),
            ("yesterday",   "📅  Gestern",          self.tab_yest.frame,
             "Pakete die gestern abgeholt wurden.\n"
             "Sortierung: Neueste zuerst."),
            "DHL",
            ("dhl_express", "🚚  DHL Express",     self.tab_dhl.frame,
             "DHL Express Scans von heute.\n"
             "Zeigt Barcode + Scan-Zeitpunkt."),
            ("dhl_heute",   "🚛  DHL heute",       self.tab_dhl_merge.frame,
             "DHL Normal Scans von heute\n"
             "(OrcaScan DHL_Normal Sheet).\n"
             "Export als YYMMDD.xlsx möglich."),
        ]

        INDICATOR   = "#4a9fd4"   # helles Blau – aktiver Markierungsreiter
        ACTIVE_BG   = "#243d5c"   # Hintergrund aktiver Eintrag

        tk.Frame(self.sidebar, bg=SIDE_BG, height=8).pack()  # Abstand oben

        for item in ITEMS:
            if item is None:
                _sep()
                continue
            if isinstance(item, str):
                _section_label(item)
                continue

            key, text, frame, tooltip = item

            # Zeilen-Container
            row = tk.Frame(self.sidebar, bg=SIDE_BG, cursor="hand2")
            row.pack(fill="x")

            # Markierungsstreifen links (4 px breit, zunächst unsichtbar)
            indicator = tk.Frame(row, bg=SIDE_BG, width=4)
            indicator.pack(side="left", fill="y")
            indicator.pack_propagate(False)

            # Text-Label
            lbl = tk.Label(
                row,
                text=f" {text}",
                font=("Segoe UI", 10),
                bg=SIDE_BG, fg="#8aafc8",
                anchor="w", cursor="hand2",
                padx=6, pady=9,
            )
            lbl.pack(side="left", fill="both", expand=True)

            def _click(e, f=frame, k=key):
                self._select_tab(f, k)

            def _enter(e, r=row, ind=indicator, lbl=lbl, k=key):
                if k != self._active_sidebar_key:
                    r.config(bg=HOVER_BG)
                    ind.config(bg=HOVER_BG)
                    lbl.config(bg=HOVER_BG)

            def _leave(e, r=row, ind=indicator, lbl=lbl, k=key):
                if k == self._active_sidebar_key:
                    r.config(bg=ACTIVE_BG); ind.config(bg=INDICATOR); lbl.config(bg=ACTIVE_BG)
                else:
                    r.config(bg=SIDE_BG);   ind.config(bg=SIDE_BG);   lbl.config(bg=SIDE_BG)

            for w in (row, indicator, lbl):
                w.bind("<Button-1>", _click)
                w.bind("<Enter>",    _enter)
                w.bind("<Leave>",    _leave)

            add_tooltip(lbl, tooltip, delay=700)

            # (row, indicator, lbl) speichern für _select_tab
            self._sidebar_btns[key] = (row, indicator, lbl)

        # ── Steuerung (unterer Bereich der Sidebar) ─────────────────────
        # Platzhalter der die Steuerung nach unten schiebt
        spacer = tk.Frame(self.sidebar, bg=SIDE_BG)
        spacer.pack(fill="both", expand=True)

        tk.Frame(self.sidebar, bg="#243a5e", height=1).pack(fill="x", padx=14, pady=(0, 4))

        def _ctrl_btn(text, cmd, color):
            b = tk.Button(
                self.sidebar, text=text, command=cmd,
                bg=color, fg="white", relief="flat",
                font=("Segoe UI", 9, "bold"), cursor="hand2",
                activebackground=color, activeforeground="white",
                padx=8, pady=5, anchor="w",
            )
            b.pack(fill="x", padx=10, pady=2)
            return b

        _ctrl_btn("🔄  Neu laden (F5)", self.reload, "#3d566e")
        if ADMIN_MODE:
            _ctrl_btn("🧹  Cleanup", lambda: self.run_cleanup_async(dry_run=False), "#8e44ad")

    def _flash_tile(self, count_lbl):
        """Lässt den Zähler kurz hell aufleuchten wenn sich der Wert geändert hat."""
        count_lbl.config(fg="#f0e040")
        self.after(450, lambda: count_lbl.config(fg="white"))

    def _update_tiles(self, report_data: dict):
        """Aktualisiert alle Kacheln und das Balkendiagramm."""
        from datetime import datetime as _dt
        today = report_data.get("today")
        ts = _dt.now().strftime("%d.%m.%Y  %H:%M")
        self._report_date_lbl.config(text=f"Stand: {ts}")

        WARN_COLOR    = "#c0392b"   # rot wenn Warnung + count > 0
        GREEN_OK      = "#27ae60"   # grün wenn Warnung + count = 0

        # Prozentzahl-Konfiguration: welche Kachel bekommt % von welchem Gesamtwert?
        total_aktiv = report_data.get("abholbereit", 0) + report_data.get("verpackt", 0)
        n_abhol     = report_data.get("abholbereit", 0)
        PCT_CONFIG = {
            "abholbereit": (total_aktiv, "aller aktiven Pakete"),
            "verpackt":    (total_aktiv, "aller aktiven Pakete"),
        }

        for key, tile_data in self._tile_lbls.items():
            count_lbl, inner, outer, _cur_color, mode = tile_data[:5]
            trend_lbl = tile_data[5] if len(tile_data) > 5 else None
            pct_lbl   = tile_data[6] if len(tile_data) > 6 else None

            if key not in report_data:
                continue          # Kacheln wie "pickup_heute" werden separat via Callback gesetzt
            n = report_data.get(key, 0)
            old_val = count_lbl.cget("text")
            count_lbl.config(text=str(n))
            if old_val != str(n):
                self._flash_tile(count_lbl)

            # Trend-Pfeil berechnen
            if trend_lbl is not None:
                if key == "yesterday":
                    # Vergleich: gestern vs. vorgestern
                    prev = report_data.get("day_before_yest")
                    tip  = f"Vorgestern: {prev}" if prev is not None else ""
                else:
                    prev = self._prev_tile_counts.get(key)
                    tip  = ""
                if prev is None:
                    trend_lbl.config(text="")
                elif n > prev:
                    trend_lbl.config(text="▲", fg="#a8e6cf")   # hellgrün
                    if tip:
                        add_tooltip(trend_lbl, tip)
                elif n < prev:
                    trend_lbl.config(text="▼", fg="#ffb3b3")   # hellrot
                    if tip:
                        add_tooltip(trend_lbl, tip)
                else:
                    trend_lbl.config(text="=")

            # Grundfarbe aus der Original-Definition holen
            orig = {
                "abholbereit":  "#1abc9c",
                "pickup_heute": "#117a65",
                "verpackt":     "#2980b9",
                "pakete_heute": "#1a237e",
                "pay":          "#2c3e50",
                "older7":       "#e67e22",
                "kissel":       "#8e44ad",
                "yesterday":    "#5d6d7e",
            }.get(key, _cur_color)

            # Farbe berechnen
            if mode == "warn":
                new_color = WARN_COLOR if n > 0 else GREEN_OK
            elif mode == "warn50":
                new_color = WARN_COLOR if n > 50 else orig
            else:
                new_color = orig

            # Farbe auf alle Widgets setzen
            for w in (outer, inner, count_lbl):
                w.config(bg=new_color)
            for w in inner.winfo_children():
                try:
                    w.config(bg=new_color)
                except Exception:
                    pass

            # Prozentzahl aktualisieren
            if pct_lbl is not None:
                if key in PCT_CONFIG:
                    denom, suffix = PCT_CONFIG[key]
                    if denom > 0:
                        pct_lbl.config(text=f"{round(n / denom * 100)}%  {suffix}", bg=new_color)
                    else:
                        pct_lbl.config(text="", bg=new_color)
                else:
                    pct_lbl.config(text="", bg=new_color)

            # gespeicherte Farbe aktualisieren (für Hover-Leave)
            self._tile_lbls[key] = (count_lbl, inner, outer, new_color, mode, trend_lbl, pct_lbl)

        # Zählwerte für nächsten Trend-Vergleich speichern
        for key in report_data:
            if key in self._tile_lbls:
                self._prev_tile_counts[key] = report_data[key]

        # Balkendiagramm aktualisieren
        self._last_daily7 = report_data.get("daily7", [])
        self._redraw_chart()

    def _on_chart_hover(self, event):
        """Zeigt Tooltip beim Hover über einen Balken."""
        for (x1, y1, x2, y2, day_str, count) in getattr(self, "_chart_bar_areas", []):
            if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                self._chart_tip.config(text=f"{day_str}: {count} Pakete")
                tip_x = min(event.x + 12, self._chart_canvas.winfo_width() - 120)
                tip_y = event.y - 24
                self._chart_tip.place(x=tip_x, y=tip_y)
                return
        self._hide_chart_tooltip()

    def _hide_chart_tooltip(self):
        self._chart_tip.place_forget()

    def _redraw_chart(self):
        """Zeichnet das Canvas-Balkendiagramm neu (wird bei Resize und Daten-Update aufgerufen)."""
        daily_7 = getattr(self, "_last_daily7", [])
        c = self._chart_canvas
        c.delete("all")
        self._chart_bar_areas = []
        if not daily_7:
            return

        DAY_DE     = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
        PAD_LEFT   = 82    # Platz für Tag-Label links
        PAD_RIGHT  = 52    # Platz für Zahl rechts
        BAR_H      = 26    # Balkenhöhe
        BAR_GAP    = 10    # Abstand zwischen Balken
        PAD_TOP    = 12    # Abstand oben

        canvas_w = c.winfo_width()
        if canvas_w < 200:
            canvas_w = 800  # Fallback vor erstem Render

        max_cnt = max((cnt for _, cnt in daily_7), default=1) or 1
        bar_area_w = canvas_w - PAD_LEFT - PAD_RIGHT
        n = len(daily_7)
        needed_h = PAD_TOP + n * (BAR_H + BAR_GAP)
        c.config(height=max(needed_h + 10, 80))

        for i, (d, count) in enumerate(daily_7):
            y_top = PAD_TOP + i * (BAR_H + BAR_GAP)
            y_mid = y_top + BAR_H // 2
            y_bot = y_top + BAR_H

            # Tag-Label
            day_str = f"{DAY_DE[d.weekday()]}  {d.strftime('%d.%m.')}"
            c.create_text(PAD_LEFT - 8, y_mid, text=day_str,
                          anchor="e", font=("Segoe UI", 10), fill="#2c3e50")

            # Hintergrundbalken (hellgrau)
            c.create_rectangle(PAD_LEFT, y_top, PAD_LEFT + bar_area_w, y_bot,
                                fill="#e4e8ec", outline="", tags="bar_bg")

            # Farbbalken
            if count:
                bar_w = max(int(count / max_cnt * bar_area_w), 4)
                # Abgerundete Optik: kleines Rechteck + normales Rechteck
                c.create_rectangle(PAD_LEFT, y_top, PAD_LEFT + bar_w, y_bot,
                                   fill="#27ae60", outline="", tags="bar")

            # Zahl rechts vom Balken
            bar_w_actual = max(int(count / max_cnt * bar_area_w), 0) if count else 0
            num_x = PAD_LEFT + bar_w_actual + 8
            c.create_text(num_x, y_mid, text=str(count),
                          anchor="w", font=("Segoe UI", 10, "bold"), fill="#2c3e50")

            # Hit-Area für Hover-Tooltip (volle Zeile)
            self._chart_bar_areas.append(
                (PAD_LEFT, y_top, PAD_LEFT + bar_area_w, y_bot, day_str, count)
            )

    def _set_report(self, text: str):
        """Veraltet – nur noch für Kompatibilität."""
        pass

    def _set_status(self, text: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.statusbar.config(text=f"[{ts}]  {text}")

    def _start_loading(self, text: str = "Lade Daten …"):
        self._set_status(text)
        self._progress.start(12)

    def _on_pu_count_change_pakete(self):
        """Aktualisiert die Pakete-heute-Kachel: DHL Normal + DHL Express + PU."""
        if "pakete_heute" not in self._tile_lbls:
            return
        n_pu = self._n_pu_heute
        total = self._n_dhl_heute + self._n_dhl_express + n_pu
        count_lbl = self._tile_lbls["pakete_heute"][0]
        old_val = count_lbl.cget("text")
        count_lbl.config(text=str(total))
        if old_val != str(total):
            self._flash_tile(count_lbl)

    def _on_pu_count_change(self, n: int):
        """Wird von PickupHeuteTab aufgerufen wenn Daten geladen werden → Kachel aktualisieren."""
        self._n_pu_heute = n
        if "pickup_heute" not in self._tile_lbls:
            return
        tile_data = self._tile_lbls["pickup_heute"]
        count_lbl, inner, outer, cur_color, mode = tile_data[:5]
        trend_lbl = tile_data[5] if len(tile_data) > 5 else None
        pct_lbl   = tile_data[6] if len(tile_data) > 6 else None

        old_val = count_lbl.cget("text")
        count_lbl.config(text=str(n))
        if old_val != str(n):
            self._flash_tile(count_lbl)

        # Trend-Pfeil
        if trend_lbl is not None:
            prev = self._prev_tile_counts.get("pickup_heute")
            if prev is None:
                trend_lbl.config(text="")
            elif n > prev:
                trend_lbl.config(text="▲", fg="#a8e6cf")
            elif n < prev:
                trend_lbl.config(text="▼", fg="#ffb3b3")
            else:
                trend_lbl.config(text="")
        self._prev_tile_counts["pickup_heute"] = n

        # Prozentzahl: Verpackt (PU) Anteil – wird über _update_pu_pct_label gesetzt
        self._update_pu_pct_label()

        self._on_pu_count_change_pakete()
        self._refresh_unstimmigkeiten()

    def _update_pu_pct_label(self):
        """Aktualisiert die Prozentanzeige der PU-heute-Kachel: Verpackt / Gesamt PU."""
        if "pickup_heute" not in self._tile_lbls:
            return
        tile_data = self._tile_lbls["pickup_heute"]
        pct_lbl = tile_data[6] if len(tile_data) > 6 else None
        if pct_lbl is None:
            return
        n_total    = self._n_pu_heute
        n_verpackt = self._n_pu_verpackt
        if n_total > 0:
            pct = round(n_verpackt / n_total * 100)
            n_nicht_verpackt = n_total - n_verpackt
            pct_lbl.config(text=f"{n_verpackt} verpackt  ·  {n_nicht_verpackt} offen  ({pct}%)")
        else:
            pct_lbl.config(text="")

    def _stop_loading(self, text: str = "Bereit."):
        self._progress.stop()
        self._set_status(text)

    def _check_stale(self):
        """Prüft ob die letzte Ladung mehr als 10 Minuten her ist → Button orange."""
        import time as _time
        if self._last_load_time > 0:
            age = _time.time() - self._last_load_time
            if age > 600:  # 10 Minuten
                self.refresh_btn.config(bg="#e67e22", activebackground="#e67e22")
        self.after(60_000, self._check_stale)

    def load_logo(self):
        if not LOGO_PATH.exists():
            return
        try:
            img = tk.PhotoImage(file=str(LOGO_PATH))
            self.logo_img = img
            self.logo_lbl.config(image=img)
        except Exception as e:
            messagebox.showwarning("Logo", f"Logo konnte nicht geladen werden:\n{e}")

    # ------------------------------------------------------------------ actions

    # ── Zahlung offen – Rechtsklick-Menü ─────────────────────────────────────

    def _pay_right_click(self, event):
        """Rechtsklick auf eine Zeile im Zahlung-offen-Tab öffnet Kontextmenü."""
        try:
            row_idx = self.tab_pay.sheet.identify_row(event, allow_end=False)
        except Exception:
            return
        if row_idx is None or row_idx >= len(self.tab_pay.filtered):
            return
        row = self.tab_pay.filtered[row_idx]
        oid = row[6] if len(row) > 6 else ""
        if not oid:
            return
        menu = tk.Menu(self.nb, tearoff=0)
        menu.add_command(
            label="✔  Als bezahlt markieren",
            command=lambda: self._set_pay_bezahlt(row)
        )
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _set_pay_bezahlt(self, row):
        """Setzt Zahlung für eine Zeile auf 'Bezahlt' in OrcaScan und entfernt sie lokal."""
        oid, bc, nm = row[6], row[0], row[1]
        self._start_loading(f"Setze '{bc}' auf bezahlt …")
        def _worker():
            result = update_rows_orca_bulk(
                [(oid, bc, nm)], {"zahlung": "Bezahlt"},
                sheet_id=ORCA_ABHOLER_SHEET_ID)
            self.after(0, lambda: _done(result))
        def _done(result):
            failed = result.get("failed", [])
            if failed:
                self._stop_loading(f"❌ Fehler: {failed[0][:60]}")
            else:
                # Zeile lokal aus der Liste entfernen
                self.tab_pay.rows = [r for r in self.tab_pay.rows if r[0] != bc]
                self.tab_pay.refresh()
                self._stop_loading(f"✓ '{bc}' als bezahlt markiert")
        threading.Thread(target=_worker, daemon=True).start()

    def _enrich_pay_tours(self):
        """Ergänzt den Status im Zahlung-offen-Tab mit Tournummer aus PU heute."""
        if not hasattr(self, "tab_pay"):
            return
        pu_rows  = getattr(self.tab_pickup_heute, "_all_rows", [])
        tour_map = {r["barcode"]: r.get("tour", "") for r in pu_rows if r.get("tour")}
        if not tour_map:
            return
        new_rows = []
        for row in self.tab_pay.rows:
            bc, status = row[0], row[3]
            tour = tour_map.get(bc, "")
            if tour and "verpackt" in status.lower():
                row = list(row)
                row[3] = f"Verpackt ({tour})"
                row = tuple(row)
            new_rows.append(row)
        self.tab_pay.rows = new_rows
        self.tab_pay.refresh()

    def open_file(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.load_main(p)

    def load_main(self, path: str):
        try:
            df = pd.read_excel(path)
            result = compute_all_rows(df)
            self.last_file = path
            self._apply_main(result, Path(path).name, abholer_df=df)
        except Exception as e:
            messagebox.showerror("Fehler", str(e))
            self._set_status(f"Fehler beim Laden: {e}")

    def load_main_orca(self):
        self._start_loading("Lade Abholer_DB aus OrcaScan …")
        import threading
        def _worker():
            try:
                df = fetch_abholer_orca()
                result = compute_all_rows(df)
                self.after(0, lambda d=df: self._apply_main(result, "OrcaScan", abholer_df=d))
            except Exception as e:
                self.after(0, lambda err=e: (
                    self._stop_loading(f"Fehler: {err}"),
                    messagebox.showerror("OrcaScan Fehler", str(err))
                ))
        threading.Thread(target=_worker, daemon=True).start()

    def _apply_main(self, result, label: str, abholer_df=None):
        report_data, r_abhol, r_older, r_yest, r_pay, r_kissel, r_verpackt = result
        self.tab_abhol.set_rows(r_abhol)
        self.tab_older.set_rows(r_older)
        self.tab_yest.set_rows(r_yest)
        self.tab_pay.set_rows(r_pay)
        self._enrich_pay_tours()
        self.tab_kissel.set_rows(r_kissel)
        self.tab_verpackt.set_rows(r_verpackt)
        # Pakete heute = DHL Normal + DHL Express + PU heute
        report_data["pakete_heute"] = (self._n_dhl_heute + self._n_dhl_express
                                       + self._n_pu_heute)
        self._update_tiles(report_data)
        self._refresh_unstimmigkeiten()
        self.title(f"Bombadil  –  {label}")
        self._stop_loading(f"Geladen: {label}")
        import time as _time
        self._last_load_time = _time.time()
        self.refresh_btn.config(bg="#27ae60", activebackground="#27ae60")
        if abholer_df is not None:
            self.last_abholer_df = abholer_df
            self.tab_statistik.update_main(abholer_df)

    def open_dhl_file(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.load_dhl(p)

    def load_dhl(self, path: str):
        try:
            rows = compute_dhl_today_rows(path)
            self.tab_dhl.set_rows(rows)
            self._select_tab(self.tab_dhl.frame, "dhl_express")
            self._set_status(f"DHL_Express: {Path(path).name}  –  {len(rows)} Scans heute")
        except Exception as e:
            messagebox.showerror("Fehler DHL_Express", str(e))
            self._set_status(f"Fehler DHL_Express: {e}")

    def load_dhl_orca(self):
        self._start_loading("Lade DHL_Normal, DHL_Express aus OrcaScan …")
        import threading
        def _worker():
            try:
                self.after(0, lambda: self._set_status("Lade DHL_Normal …"))
                df_normal = fetch_sheet_orca(ORCA_DHL_NORMAL_SHEET_ID)
                self.after(0, lambda: self._set_status("Lade DHL_Express …"))
                df_ex     = fetch_sheet_orca(ORCA_DHL_EX_SHEET_ID)
                rows_ex               = compute_dhl_today_rows(df_ex)
                merged_df, rows_merge = compute_dhl_normal_today(df_normal)
                self.after(0, lambda: self._apply_dhl_orca(rows_ex, rows_merge, merged_df))
            except Exception as e:
                err_msg = str(e)
                self.after(0, lambda m=err_msg: self._stop_loading(f"Fehler DHL OrcaScan: {m}"))
                self.after(10, lambda m=err_msg: messagebox.showerror("DHL OrcaScan Fehler", m))
        threading.Thread(target=_worker, daemon=True).start()

    def _apply_dhl_orca(self, rows_ex, rows_merge, merged_df=None):
        self.tab_dhl.set_rows(rows_ex)
        self.tab_dhl_merge.set_rows_direct(rows_merge, merged_df)
        self._stop_loading(
            f"DHL OrcaScan geladen  –  Express: {len(rows_ex)} | Merge: {len(rows_merge)} | Gesamt: {len(rows_ex) + len(rows_merge)} Scans heute"
        )
        # Pakete-heute-Kachel aktualisieren: DHL Normal + DHL Express + PU
        self._n_dhl_heute    = len(rows_merge)
        self._n_dhl_express  = len(rows_ex)
        self._on_pu_count_change_pakete()

    # ------------------------------------------------------------------ Hilfe
    def _show_help(self):
        win = tk.Toplevel(self)
        win.title("Bombadil – Funktionsübersicht")
        win.geometry("620x560")
        win.resizable(False, False)
        txt = tk.Text(win, wrap="word", font=("Segoe UI", 10),
                      padx=16, pady=12, relief="flat", bg="#fafafa")
        txt.pack(fill="both", expand=True)
        sb = ttk.Scrollbar(win, command=txt.yview)
        sb.pack(side="right", fill="y")
        txt.configure(yscrollcommand=sb.set)
        txt.tag_configure("h", font=("Segoe UI", 11, "bold"), foreground="#1a3a5c", spacing1=10)
        txt.tag_configure("b", font=("Segoe UI", 10, "bold"))
        content = [
            ("h", "📋  Report\n"),
            ("",  "Gesamtübersicht: Anzahl abholbereiter Pakete, >7 Tage, gestern abgeholt, verpackt ohne Abholbereit_At.\n\n"),
            ("h", "📦  Abholbereit\n"),
            ("",  "Alle Pakete mit Status 'Abholbereit' – vollständige Liste ohne Datumsfilter.\n\n"),
            ("h", "⏳  > 7 Tage\n"),
            ("",  "Pakete die seit mehr als 7 Tagen abholbereit sind – Handlungsbedarf!\n\n"),
            ("h", "📅  Gestern\n"),
            ("",  "Pakete die gestern abgeholt wurden.\n\n"),
            ("h", "🔧  Verpackt\n"),
            ("",  "Pakete mit Status 'Verpackt' aber noch kein Abholbereit_At – warten auf Freigabe.\n\n"),
            ("h", "💳  Zahlung offen\n"),
            ("",  "Pakete bei denen die Zahlung noch aussteht.\n\n"),
            ("h", "🏪  Kissel > 3W\n"),
            ("",  "Pakete am Kiosk Kissel die seit mehr als 3 Wochen nicht abgeholt wurden.\n\n"),
            ("h", "🚚  DHL_Express (heute)\n"),
            ("",  "Heutige DHL Express Scans – aus OrcaScan oder lokaler Excel-Datei.\n\n"),
            ("h", "📬  DHL (heute)\n"),
            ("",  "DHL_Normal Sheet – nur heutige Scans.\n\n"),
            ("h", "🔍  Tagesboten Abgleich\n"),
            ("",  "Vergleicht die Tagesboten-Liste (Google Drive) mit der Abholer_DB.\n"
                  "• Fehlerliste: Barcode vorhanden, aber Abholbereit_At fehlt\n"
                  "• Errorliste: Barcode gar nicht in der DB (Auto-Export)\n"),
        ]
        txt.configure(state="normal")
        for tag, line in content:
            txt.insert("end", line, tag if tag else ())
        txt.configure(state="disabled")

    def _show_shortcuts(self):
        win = tk.Toplevel(self)
        win.title("Tastenkürzel")
        win.geometry("360x280")
        win.resizable(False, False)
        rows = [
            ("Strg+O",  "Abholer_DB Excel auswählen"),
            ("Strg+D",  "DHL_Express Excel auswählen"),
            ("F5",      "Aktuelle Daten neu laden"),
            ("Klick",   "Zellwert in Zwischenablage kopieren"),
        ]
        frm = tk.Frame(win, padx=20, pady=16)
        frm.pack(fill="both", expand=True)
        tk.Label(frm, text="Tastenkürzel", font=("Segoe UI", 12, "bold"),
                 fg="#1a3a5c").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0,12))
        for i, (key, desc) in enumerate(rows, start=1):
            tk.Label(frm, text=key, font=("Segoe UI", 10, "bold"),
                     fg="#2980b9", width=10, anchor="w").grid(row=i, column=0, sticky="w", pady=3)
            tk.Label(frm, text=desc, font=("Segoe UI", 10),
                     anchor="w").grid(row=i, column=1, sticky="w", pady=3)

    def _show_about(self):
        messagebox.showinfo(
            "Über Bombadil",
            "Bombadil\n\n"
            "Paketverwaltung & Abgleich-Tool\n"
            "Cannabis Apotheke Frankfurt / Kissel Apotheke\n\n"
            "Funktionen:\n"
            "• Abholer_DB via OrcaScan API (live)\n"
            "• DHL_Express & DHL A/B via OrcaScan\n"
            "• Tagesboten-Abgleich via Google Drive\n"
            "• Auto-Import & Auto-Refresh\n"
        )

    # ------------------------------------------------------------------ Einstellungen
    def _load_settings(self):
        import json as _json
        defaults = {"export_folder": str(Path.home() / "Downloads")}
        try:
            if SETTINGS_FILE.exists():
                with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                    data = _json.load(f)
                defaults.update(data)
        except Exception:
            pass
        self.export_folder = Path(defaults["export_folder"])
        self.last_backup_date = defaults.get("last_backup_date", "")
        self.tourlisten_folder_id = defaults.get("tourlisten_folder_id", "")

    def _save_settings(self):
        import json as _json
        data = {"export_folder": str(self.export_folder),
                "last_backup_date": self.last_backup_date,
                "tourlisten_folder_id": self.tourlisten_folder_id}
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                _json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("Einstellungen", f"Speichern fehlgeschlagen:\n{e}")

    def _choose_export_folder(self):
        p = filedialog.askdirectory(title="Exportordner wählen",
                                    initialdir=str(self.export_folder))
        if p:
            self.export_folder = Path(p)
            self._save_settings()
            messagebox.showinfo("Einstellungen",
                                f"Exportordner gesetzt:\n{self.export_folder}")

    def _reset_settings(self):
        if messagebox.askyesno("Einstellungen zurücksetzen",
                               "Alle Einstellungen auf Standard zurücksetzen?"):
            self.export_folder = Path.home() / "Downloads"
            self._save_settings()
            messagebox.showinfo("Einstellungen", "Zurückgesetzt auf:\n" + str(self.export_folder))

    # ── Backup ────────────────────────────────────────────────────────────

    _BACKUP_HOUR = 20  # 20:00 Uhr

    def _schedule_backup_check(self):
        now = datetime.now()
        today_str = now.strftime("%Y-%m-%d")
        if now.hour >= self._BACKUP_HOUR and self.last_backup_date != today_str:
            self._run_backup(manual=False)
        self.after(60_000, self._schedule_backup_check)

    def _run_backup(self, manual=False):
        def worker():
            try:
                path = backup_abholer_db(self.export_folder)
                today_str = datetime.now().strftime("%Y-%m-%d")
                self.after(0, lambda p=path, d=today_str: self._on_backup_done(p, d, manual))
            except Exception as e:
                self.after(0, lambda err=str(e): self._on_backup_error(err, manual))
        threading.Thread(target=worker, daemon=True).start()

    def _on_backup_done(self, path: Path, today_str: str, manual: bool):
        self.last_backup_date = today_str
        self._save_settings()
        self._set_status(f"Backup erstellt: {path}")
        if manual:
            messagebox.showinfo("Backup", f"Backup erfolgreich gespeichert:\n{path}")

    def _on_backup_error(self, err: str, manual: bool):
        self._set_status(f"⚠ Backup fehlgeschlagen: {err}")
        if manual:
            messagebox.showerror("Backup fehlgeschlagen", err)

    def choose_folder(self):
        p = filedialog.askdirectory()
        if p:
            self.watch_folder = p
            self.folder_lbl.config(text=f"  📂 {p}")

    REFRESH_INTERVAL_MS = 5 * 60 * 1000   # 5 Minuten

    def toggle_refresh(self):
        self.refresh_on = not self.refresh_on
        if self.refresh_on:
            self.refresh_btn.config(text="Auto-Refresh: AN",  bg="#27ae60", activebackground="#27ae60")
            self.load_main_orca()          # sofort einmal laden
            self._schedule_refresh()
        else:
            self.refresh_btn.config(text="Auto-Refresh: AUS", bg="#e74c3c", activebackground="#e74c3c")
            self.refresh_lbl.config(text="")
            if self.refresh_job:
                try:
                    self.after_cancel(self.refresh_job)
                except Exception:
                    pass
                finally:
                    self.refresh_job = None

    def _schedule_refresh(self):
        if not self.refresh_on:
            return
        self.refresh_job = self.after(self.REFRESH_INTERVAL_MS, self._auto_refresh)

    def _auto_refresh(self):
        if not self.refresh_on:
            return
        import datetime as _dt
        self.load_main_orca()
        now = _dt.datetime.now().strftime("%H:%M")
        self.refresh_lbl.config(text=f"Zuletzt: {now}")
        self._schedule_refresh()   # nächsten Refresh planen

    DHL_REFRESH_INTERVAL_MS = 5 * 60 * 1000   # 5 Minuten

    def toggle_dhl_refresh(self):
        self.dhl_refresh_on = not self.dhl_refresh_on
        if self.dhl_refresh_on:
            self.dhl_refresh_btn.config(text="Auto-Refresh: AN",  bg="#27ae60", activebackground="#27ae60")
            self.load_dhl_orca()
            self._schedule_dhl_refresh()
        else:
            self.dhl_refresh_btn.config(text="Auto-Refresh: AUS", bg="#e74c3c", activebackground="#e74c3c")
            self.dhl_refresh_lbl.config(text="")
            if self.dhl_refresh_job:
                try:
                    self.after_cancel(self.dhl_refresh_job)
                except Exception:
                    pass
                finally:
                    self.dhl_refresh_job = None

    def _schedule_dhl_refresh(self):
        if not self.dhl_refresh_on:
            return
        self.dhl_refresh_job = self.after(self.DHL_REFRESH_INTERVAL_MS, self._auto_dhl_refresh)

    def _auto_dhl_refresh(self):
        if not self.dhl_refresh_on:
            return
        import datetime as _dt
        self.load_dhl_orca()
        now = _dt.datetime.now().strftime("%H:%M")
        self.dhl_refresh_lbl.config(text=f"Zuletzt: {now}")
        self._schedule_dhl_refresh()

    def toggle_watch(self):
        self.watch_on = not self.watch_on
        if self.watch_on:
            self.watch_btn.config(text="Auto-Import: AN",  bg="#27ae60", activebackground="#27ae60")
        else:
            self.watch_btn.config(text="Auto-Import: AUS", bg="#e74c3c", activebackground="#e74c3c")
        if self.watch_on:
            nf = newest_excel(self.watch_folder)
            if nf:
                self.load_main(nf)

    def reload(self):
        if self.last_file:
            self.load_main(self.last_file)
        else:
            self._set_status("Kein Reload: noch keine Datei geladen.")

    # ------------------------------------------------------------------ Cleanup

    def run_cleanup_async(self, dry_run=False):
        """Startet den Cleanup-Prozess im Hintergrund-Thread."""
        import threading

        def _on_status(text):
            self.after(0, lambda t=text: self._set_status(t))

        def _on_done(text, errors=None):
            def _show(t=text, errs=errors):
                self._set_status(t)
                self._stop_loading(t)
                if dry_run:
                    messagebox.showinfo("Cleanup – Dry Run Ergebnis", t)
                elif errs:
                    detail = "\n".join(errs)
                    messagebox.showwarning(
                        "Cleanup – Fehler beim Löschen",
                        f"{t}\n\nFehlerdetails (bis zu 10 eindeutige Fehler):\n{detail}"
                    )
            self.after(0, _show)

        def _on_preview(df):
            self.after(0, lambda d=df: self._show_cleanup_preview(d))

        label = "DRY RUN Cleanup …" if dry_run else "Cleanup läuft …"
        self._start_loading(label)
        threading.Thread(
            target=run_cleanup,
            kwargs={
                "on_status":  _on_status,
                "on_done":    _on_done,
                "on_preview": _on_preview if dry_run else None,
                "dry_run":    dry_run,
            },
            daemon=True,
        ).start()

    def _show_cleanup_preview(self, df: "pd.DataFrame"):
        """Zeigt die Cleanup-Treffer in einem Popup-Fenster (Dry Run)."""
        win = tk.Toplevel(self)
        win.title("Cleanup – Dry Run Vorschau")
        win.geometry("1100x600")

        tk.Label(win,
                 text=f"Dry Run – {len(df)} Einträge würden gelöscht (noch nichts passiert)",
                 font=("Segoe UI", 10, "bold"), fg="#8e44ad").pack(fill="x", padx=10, pady=(10, 4))

        # Spalten für Anzeige auswählen
        show_cols = [c for c in [
            "Paket-Barcode", "Name", "Paketstatus",
            "Abgeholt_At", "Abholbereit_At", "Scan-Datum",
        ] if c in df.columns]

        frame = tk.Frame(win)
        frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        tree = ttk.Treeview(frame, columns=show_cols, show="headings", selectmode="none")
        for col in show_cols:
            tree.heading(col, text=col)
            tree.column(col, width=160, anchor="w", minwidth=80)
        tree.tag_configure("even", background="#f2f2f2")
        tree.tag_configure("odd",  background="#ffffff")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        for i, (_, row) in enumerate(df[show_cols].iterrows()):
            vals = [str(v) if not pd.isna(v) else "" for v in row]
            tree.insert("", "end", values=vals, tags=("even" if i % 2 == 0 else "odd",))

        tk.Button(win, text="Schließen", command=win.destroy).pack(pady=(0, 10))

    def poll(self):
        try:
            if self.watch_on:
                nf = newest_excel(self.watch_folder)
                if nf and nf != self.last_file and nf != self._last_failed_file:
                    time.sleep(0.3)
                    try:
                        self.load_main(nf)
                        self._last_failed_file = None   # Erfolg → Reset
                    except Exception as e:
                        self._last_failed_file = nf     # Datei merken → nicht nochmal versuchen
                        self._set_status(f"Auto-Import Fehler: {e}")
        finally:
            self.after(POLL_MS, self.poll)


# ============================================================
# Start-sicher
# ============================================================
if __name__ == "__main__":
    try:
        App().mainloop()
    except Exception as e:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Startfehler", f"{type(e).__name__}: {e}")
        raise
